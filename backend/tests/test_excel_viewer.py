"""Data Source workbook reader: windowing, merges and the style guards.

Hermetic on purpose -- every test builds a tiny workbook in tmp_path rather
than touching the 10 MB file in resources/, which takes ~13s to parse.
"""

from __future__ import annotations

import json

import pytest

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Color,
    Font,
    PatternFill,
)

from src.excel import workbook as reader
from src.excel.formatting import (
    MAX_COLUMN_PX,
    column_width_px,
    format_value,
    rgb_hex,
)


@pytest.fixture(autouse=True)
def _clear_cache():
    reader._CACHE = None
    yield
    reader._CACHE = None


@pytest.fixture
def sample_path(tmp_path, monkeypatch):
    book = Workbook()

    sheet = book.active
    sheet.title = "Grid"

    sheet["A1"] = "Banner"
    sheet["A1"].font = Font(bold=True, color="FFFFFF00")
    sheet["A1"].fill = PatternFill(
        fill_type="solid",
        fgColor="FF1E3A5F",
    )
    sheet["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    sheet.merge_cells("A1:C1")

    sheet["A2"] = "Label"
    # A theme colour: openpyxl hands back a descriptor object here, not a
    # string, and a naive guard would leak it into the payload or raise.
    sheet["A2"].font = Font(color=Color(theme=1, type="theme"))

    sheet["B2"] = 1234567
    sheet["B2"].number_format = "#,##0"
    sheet["C2"] = 0.1234
    sheet["C2"].number_format = "0.0%"

    for row in range(3, 13):
        sheet.cell(row=row, column=1, value=f"r{row}")

    # A merge that a page boundary will split, so the clipped-anchor path runs.
    sheet["A6"] = "Split banner"
    sheet.merge_cells("A6:C8")

    sheet.column_dimensions["A"].width = 20
    sheet.column_dimensions["B"].width = 98

    book.create_sheet("Second")

    path = tmp_path / "sample.xlsx"
    book.save(path)

    monkeypatch.setattr(
        reader,
        "workbook_path",
        lambda: path,
    )

    return path


def test_lists_every_sheet_with_dimensions(sample_path) -> None:
    payload = reader.list_sheets()

    assert payload["workbook"] == "sample.xlsx"
    assert payload["count"] == 2
    assert [sheet["name"] for sheet in payload["sheets"]] == [
        "Grid",
        "Second",
    ]

    grid = payload["sheets"][0]

    assert grid["index"] == 0
    assert grid["row_count"] == 12
    assert grid["column_count"] == 3


def test_reads_a_window_and_reports_more(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=0, limit=5)

    assert page["returned_rows"] == 5
    assert page["row_count"] == 12
    assert page["has_more"] is True
    assert [row["row"] for row in page["rows"]] == [1, 2, 3, 4, 5]

    # Dense rows: one entry per column, null where the cell is empty.
    assert all(
        len(row["cells"]) == page["column_count"]
        for row in page["rows"]
    )
    assert page["rows"][2]["cells"][1] is None


def test_last_page_has_no_more(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=10, limit=100)

    assert page["returned_rows"] == 2
    assert page["has_more"] is False
    assert page["rows"][0]["row"] == 11


def test_offset_past_the_end_is_an_empty_page(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=500, limit=100)

    assert page["rows"] == []
    assert page["returned_rows"] == 0
    assert page["has_more"] is False


def test_unknown_sheet_raises(sample_path) -> None:
    with pytest.raises(reader.UnknownSheet):
        reader.read_sheet("Nope")


def test_missing_workbook_raises(tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(
        reader,
        "workbook_path",
        lambda: tmp_path / "absent.xlsx",
    )

    with pytest.raises(reader.WorkbookMissing):
        reader.list_sheets()


def test_cell_carries_the_workbooks_own_formatting(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=0, limit=2)

    banner = page["rows"][0]["cells"][0]

    assert banner["v"] == "Banner"
    assert banner["b"] is True
    assert banner["fg"] == "#FFFF00"
    assert banner["bg"] == "#1E3A5F"
    assert banner["a"] == "center"
    assert banner["va"] == "middle"
    assert banner["w"] is True

    thousands = page["rows"][1]["cells"][1]

    assert thousands["v"] == "1,234,567"
    assert thousands["t"] == "n"

    percent = page["rows"][1]["cells"][2]

    assert percent["v"] == "12.3%"


def test_theme_colours_and_blank_fills_never_reach_the_payload(
    sample_path,
) -> None:
    page = reader.read_sheet("Grid", offset=0, limit=12)
    encoded = json.dumps(page)

    # str() of an unresolved theme colour, which a len()/slice-only guard
    # would either leak here or raise on.
    assert "Values must be of type" not in encoded

    themed = page["rows"][1]["cells"][0]

    assert themed["v"] == "Label"
    assert "fg" not in themed

    # An unfilled cell reports fgColor 00000000; emitting it paints the whole
    # grid black.
    assert all(
        "bg" not in cell
        for row in page["rows"]
        for cell in row["cells"]
        if cell and row["row"] != 1
    )


def test_merges_are_clipped_to_the_window(sample_path) -> None:
    first = reader.read_sheet("Grid", offset=0, limit=6)
    banner = next(
        merge for merge in first["merges"] if merge["row"] == 1
    )

    assert banner["column"] == 1
    assert banner["colspan"] == 3
    assert banner["rowspan"] == 1
    assert banner["clipped"] is False

    split = next(
        merge for merge in first["merges"] if merge["row"] == 6
    )

    # A6:C8 seen from a page that ends at row 6.
    assert split["rowspan"] == 1
    assert split["clipped"] is False


def test_a_merge_clipped_from_above_keeps_its_anchor(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=6, limit=6)

    split = next(
        merge for merge in page["merges"] if merge["column"] == 1
    )

    assert split["row"] == 7
    assert split["rowspan"] == 2
    assert split["clipped"] is True
    # Without the carry-over the banner would render blank on this page.
    assert split["anchor"]["v"] == "Split banner"


def test_column_widths_expand_and_clamp(sample_path) -> None:
    page = reader.read_sheet("Grid", offset=0, limit=1)

    widths = {
        column["letter"]: column["width_px"]
        for column in page["columns"]
    }

    assert widths["A"] == column_width_px(20)
    assert widths["B"] == MAX_COLUMN_PX
    # Unset columns still get a width, from the sheet default.
    assert widths["C"] >= 32


def test_rgb_hex_guards() -> None:
    assert rgb_hex(None) == ""
    assert rgb_hex(Color(theme=1, type="theme")) == ""
    assert rgb_hex(Color(rgb="00000000")) == ""
    assert rgb_hex(Color(rgb="FF1E3A5F")) == "1E3A5F"


def test_general_format_keeps_numbers_readable() -> None:
    assert format_value(7.45, "General") == "7.45"
    assert format_value(12.0, "General") == "12"
    assert format_value(3, None) == "3"
    assert format_value(True, "General") == "TRUE"
    assert format_value(None, "General") == ""
    # An unsupported format falls back to General rather than failing.
    assert format_value(1.5, "[$-409]0.00;;") == "1.5"
