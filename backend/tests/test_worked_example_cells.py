"""The worked examples, checked down to the workbook cell.

`test_formulas.py` proves the stored expressions reproduce the workbook's
answers. These tests prove the *inputs* are what the workbook actually holds:
every example now cites the source cell for each of its parameters, and a
citation that points at the wrong cell is worse than no citation at all.

Two independent guards, because they fail for different reasons:

* the parity test re-parses `resources/formula.md` and asserts the generated
  `workedExamples.json` is still byte-identical -- it catches a hand-edit to
  either file and needs nothing but the two text files;
* the cell test opens the real workbook and asserts each cited cell holds the
  documented value -- it catches a citation that drifted off a row when the
  workbook changed, and skips when the 10 MB file is not deployed.

The workbook is opened `read_only=True, data_only=True` (~0.1 s) rather than
through `src.excel.workbook`, whose cached singleton loads styles and merges
for the viewer and costs ~13 s and ~220 MB. Cells are collected in one
streaming pass per sheet: random access on a read-only worksheet re-walks the
rows on every lookup.
"""

from __future__ import annotations

import json
import math

from typing import Any

import openpyxl
import pytest

from openpyxl.utils.cell import coordinate_to_tuple

from src.excel import workbook as viewer
from src.formulas import repository
from src.formulas.verification_pack import (
    EXAMPLES_PATH,
    PACK_PATH,
    build,
    render,
)


def load_examples() -> dict[str, list[dict[str, Any]]]:
    return json.loads(EXAMPLES_PATH.read_text(encoding="utf-8"))


def corpus() -> list[tuple[str, dict[str, Any]]]:
    """Every stored rule that has a verification-pack entry, cell by cell.

    `examples.get(...)` rather than `examples[...]`: the catalogue also holds
    the `fc*` rules from `resources/custom_formulas.json`, which are derived
    or come from v8.5 `*_live` sheets and have no worked examples to cite. A
    plain lookup raises `KeyError` during collection and takes the whole
    module down with it. Nothing is silently lost by skipping them -- that
    every formula in the workbook transcript has exactly five examples is
    asserted in `test_formulas.py::test_corpus_covers_every_documented_example`.
    """
    examples = load_examples()
    return [
        (formula["id"], case)
        for formula in repository.load()
        for case in examples.get(formula["id"], [])
    ]


CORPUS = corpus()


# -- the pack and the corpus agree -------------------------------------


def test_generated_corpus_matches_the_checked_in_file() -> None:
    """workedExamples.json is generated, never hand-edited."""
    assert render(build()) == EXAMPLES_PATH.read_text(encoding="utf-8"), (
        f"{EXAMPLES_PATH.name} is out of date with {PACK_PATH.name}. "
        "Run: python -m src.formulas.verification_pack --write"
    )


def test_every_parameter_cites_a_source_cell() -> None:
    examples = load_examples()

    for formula in repository.load():
        keys = {parameter["key"] for parameter in formula["parameters"]}

        for case in examples[formula["id"]]:
            assert set(case["sources"]) == keys, (
                f"{formula['id']} {case['address']}: cited cells "
                f"{sorted(case['sources'])} do not cover {sorted(keys)}"
            )
            assert set(case["values"]) == keys

            for key, address in case["sources"].items():
                sheet, separator, cell = address.partition("!")
                assert separator and sheet and cell, (
                    f"{formula['id']} {key}: '{address}' is not Sheet!Cell"
                )


# -- the cited cells hold the documented values ------------------------


def _cited_addresses() -> set[str]:
    addresses = set()

    for _formula_id, case in CORPUS:
        addresses.add(case["address"])
        addresses.update(case["sources"].values())

    return addresses


@pytest.fixture(scope="module")
def saved_values() -> dict[str, Any]:
    """Every cited cell's saved value, read in one pass per sheet."""
    path = viewer.workbook_path()

    if not path.exists():
        pytest.skip(f"workbook not deployed at {path}")

    wanted: dict[str, dict[tuple[int, int], str]] = {}

    for address in _cited_addresses():
        sheet, _, cell = address.partition("!")
        wanted.setdefault(sheet, {})[coordinate_to_tuple(cell)] = address

    book = openpyxl.load_workbook(
        path,
        data_only=True,
        read_only=True,
    )

    try:
        # Seeded with every address so a cell the pass never reaches reads as
        # empty in the failure message rather than raising a KeyError.
        values: dict[str, Any] = dict.fromkeys(_cited_addresses())

        for sheet, positions in wanted.items():
            assert sheet in book.sheetnames, (
                f"the pack cites sheet '{sheet}', which the workbook "
                "does not have"
            )

            worksheet = book[sheet]
            last_row = max(row for row, _column in positions)

            # Positions come from enumerate, not cell.row/cell.column: a
            # read-only pass yields EmptyCell for blanks, which carries no
            # coordinates.
            for offset, row in enumerate(
                worksheet.iter_rows(min_row=1, max_row=last_row),
                start=1,
            ):
                for column, cell in enumerate(row, start=1):
                    address = positions.get((offset, column))
                    if address is not None:
                        values[address] = cell.value

        return values
    finally:
        book.close()


def _matches(actual: Any, documented: Any) -> bool:
    try:
        # The pack prints what Excel displays; the cache keeps more precision
        # (ENGINE_STORE!J4 is 29.1668846784 for a documented 29.166885).
        return math.isclose(
            float(actual),
            float(documented),
            rel_tol=1e-6,
            abs_tol=1e-9,
        )
    except (TypeError, ValueError):
        return str(actual).strip() == str(documented).strip()


@pytest.mark.parametrize(
    ("formula_id", "case"),
    CORPUS,
    ids=[f"{formula_id}:{case['address']}" for formula_id, case in CORPUS],
)
def test_cited_cells_hold_the_documented_values(
    saved_values: dict[str, Any],
    formula_id: str,
    case: dict[str, Any],
) -> None:
    for key, address in case["sources"].items():
        documented = case["values"][key]
        actual = saved_values[address]

        assert _matches(actual, documented), (
            f"{formula_id} {key}: {address} holds {actual!r}, "
            f"the pack says {documented!r}"
        )

    result = saved_values[case["address"]]

    assert _matches(result, case["expected"]), (
        f"{formula_id}: {case['address']} holds {result!r}, "
        f"the pack says {case['expected']!r}"
    )
