"""Cached read-only access to the Data Source workbook.

The viewer needs merged ranges, column widths and per-cell styles, which means
`load_workbook(read_only=False)`. On the shipped 10 MB workbook that parse
costs ~9-13 seconds and holds ~220 MB, so it happens once and the result lives
in a module-level singleton; a windowed read against the cached object is then
6-19 ms.

Two consequences worth keeping:

* The lock is held for the whole read, not just the load. openpyxl lazily
  materialises `Cell` objects into `ws._cells` as you iterate, which is not
  thread-safe, and FastAPI runs sync endpoints on a threadpool -- two
  concurrent requests genuinely race. Serialising a 6-19 ms read costs
  nothing, so do not narrow this lock.
* Sheet dimensions are computed once at load. `ws.max_row` and
  `ws.max_column` are O(cells) and recompute on *every* access -- together
  135 ms per touch on ENGINE_STORE.

Callers must be plain `def` endpoints, never `async def`: a cold load inside
the event loop would stall health checks and the chat SSE streams for the full
13 seconds.
"""

from __future__ import annotations

import logging
import threading

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from openpyxl.utils import get_column_letter

from src.common.constants import AppPaths
from src.common.env import config

from src.excel.formatting import (
    cell_payload,
    column_width_px,
)

logger = logging.getLogger(__name__)

DEFAULT_WORKBOOK_NAME = (
    "Copy of AI_360_Retail_Dataset_v8.2_General_20260806.xlsx"
)

# Matches how main.py resolves frontend/dist: the repo root is the backend
# package's parent, both locally and at /app in the container.
DEFAULT_WORKBOOK_PATH = (
    AppPaths.BACKEND_ROOT.parent
    / "resources"
    / DEFAULT_WORKBOOK_NAME
)

MAX_PAGE_ROWS = 500
DEFAULT_PAGE_ROWS = 100


class WorkbookMissing(RuntimeError):
    """The configured workbook is not on disk."""


class WorkbookUnreadable(RuntimeError):
    """The workbook exists but openpyxl could not parse it."""


class UnknownSheet(LookupError):
    """No sheet in the workbook carries that name."""


@dataclass(frozen=True)
class _Cached:
    stamp: tuple[int, int]
    workbook: Any
    # {title: (row_count, column_count)}, computed once -- see module docstring.
    dimensions: dict[str, tuple[int, int]]


_LOCK = threading.Lock()
_CACHE: _Cached | None = None


def workbook_path() -> Path:
    """The configured workbook, resolved against the repo rather than the CWD.

    `EXCEL_WORKBOOK_PATH` is written relative in `.env.example` and in every
    `.env` that follows it. Returning it unresolved made the answer depend on
    where the process was started: uvicorn runs from `backend/` and found
    nothing, and so did pytest -- which is why 110 tests in
    `test_worked_example_cells.py` skipped themselves as "workbook not
    deployed" while the 10 MB file sat exactly where it was configured to be.
    A test that cannot reach its subject and reports success is worse than one
    that fails.

    `retail_data_bootstrap/paths.py:resolve_workbook_path` already did this
    correctly. It cannot simply be called from here -- it imports
    `DEFAULT_WORKBOOK_PATH` from this module -- so the rule is repeated, and
    the two must move together.
    """
    configured = (
        config.EXCEL_WORKBOOK_PATH or ""
    ).strip()

    if not configured:
        return DEFAULT_WORKBOOK_PATH

    candidate = Path(configured).expanduser()

    if not candidate.is_absolute():
        candidate = AppPaths.BACKEND_ROOT.parent / candidate

    return candidate


def _stamp(
    path: Path,
) -> tuple[int, int]:
    """Cheap identity for the file, so replacing it reloads without a restart."""
    info = path.stat()

    return (
        info.st_mtime_ns,
        info.st_size,
    )


def _load(
    path: Path,
) -> _Cached:
    logger.info(
        "Loading Data Source workbook: %s",
        path,
    )

    try:
        book = openpyxl.load_workbook(
            path,
            data_only=True,
        )
    except Exception as error:
        raise WorkbookUnreadable(
            str(error)
        ) from error

    dimensions = {
        sheet.title: (
            sheet.max_row or 0,
            sheet.max_column or 0,
        )
        for sheet in book.worksheets
    }

    logger.info(
        "Data Source workbook ready: %d sheets",
        len(dimensions),
    )

    return _Cached(
        stamp=_stamp(path),
        workbook=book,
        dimensions=dimensions,
    )


def _cached() -> _Cached:
    """The parsed workbook, loading or reloading it if needed.

    Callers must already hold _LOCK.
    """
    global _CACHE

    path = workbook_path()

    if not path.is_file():
        raise WorkbookMissing(str(path))

    stamp = _stamp(path)

    if _CACHE is None or _CACHE.stamp != stamp:
        _CACHE = _load(path)

    return _CACHE


def warm() -> None:
    """Parse the workbook ahead of the first request.

    Called fire-and-forget from the app lifespan on a worker thread, so a
    missing or corrupt workbook is a logged warning rather than a boot failure.
    """
    try:
        with _LOCK:
            _cached()
    except WorkbookMissing as error:
        logger.warning(
            "Data Source workbook not found, skipping pre-warm: %s",
            error,
        )
    except Exception:
        logger.exception(
            "Data Source workbook pre-warm failed"
        )


def list_sheets() -> dict[str, Any]:
    with _LOCK:
        cache = _cached()

        sheets = [
            {
                "index": index,
                "name": sheet.title,
                "row_count": cache.dimensions[sheet.title][0],
                "column_count": cache.dimensions[sheet.title][1],
            }
            for index, sheet in enumerate(
                cache.workbook.worksheets
            )
        ]

    return {
        "workbook": workbook_path().name,
        "count": len(sheets),
        "sheets": sheets,
    }


def read_sheet(
    name: str,
    offset: int = 0,
    limit: int = DEFAULT_PAGE_ROWS,
) -> dict[str, Any]:
    """One window of a sheet, with the workbook's own formatting attached.

    `offset` is 0-based over the sheet's rows; an offset past the last row is
    an empty page rather than an error, so a stale pager cannot break the view.
    """
    with _LOCK:
        cache = _cached()

        if name not in cache.workbook.sheetnames:
            raise UnknownSheet(name)

        sheet = cache.workbook[name]
        row_count, column_count = cache.dimensions[name]

        first_row = offset + 1
        last_row = min(
            offset + limit,
            row_count,
        )

        rows = _read_rows(
            sheet,
            first_row,
            last_row,
            column_count,
        )

        merges = _read_merges(
            sheet,
            first_row,
            last_row,
            column_count,
        )

        columns = _read_columns(
            sheet,
            column_count,
        )

        index = cache.workbook.sheetnames.index(name)

    return {
        "sheet": name,
        "index": index,
        "offset": offset,
        "limit": limit,
        "row_count": row_count,
        "column_count": column_count,
        "returned_rows": len(rows),
        "has_more": last_row < row_count,
        "columns": columns,
        "merges": merges,
        "rows": rows,
    }


def _read_rows(
    sheet: Any,
    first_row: int,
    last_row: int,
    column_count: int,
) -> list[dict[str, Any]]:
    if column_count < 1 or last_row < first_row:
        return []

    rows: list[dict[str, Any]] = []

    row_number = first_row

    for row in sheet.iter_rows(
        min_row=first_row,
        max_row=last_row,
        min_col=1,
        max_col=column_count,
    ):
        cells = [
            cell_payload(cell)
            for cell in row
        ]

        # Dense: exactly column_count entries, null for "empty <td>". Padding
        # matters because openpyxl can hand back a short row.
        if len(cells) < column_count:
            cells.extend(
                [None] * (column_count - len(cells))
            )

        rows.append(
            {
                "row": row_number,
                "cells": cells,
            }
        )

        row_number += 1

    return rows


def _read_merges(
    sheet: Any,
    first_row: int,
    last_row: int,
    column_count: int,
) -> list[dict[str, Any]]:
    """Merged ranges clipped to the rendered window.

    Sent once at the top level rather than duplicated onto cells; the renderer
    derives the covered positions from this list. A range that starts above the
    window is re-anchored to the first visible row and marked `clipped`, and
    the caller copies the true anchor's payload into it -- otherwise a banner
    split by pagination renders blank.
    """
    merges: list[dict[str, Any]] = []

    for merged in sheet.merged_cells.ranges:
        if merged.max_row < first_row or merged.min_row > last_row:
            continue

        if merged.min_col > column_count:
            continue

        row = max(merged.min_row, first_row)
        rowspan = min(merged.max_row, last_row) - row + 1
        colspan = (
            min(merged.max_col, column_count)
            - merged.min_col
            + 1
        )

        if rowspan < 1 or colspan < 1:
            continue

        if rowspan == 1 and colspan == 1:
            continue

        clipped = row != merged.min_row

        entry: dict[str, Any] = {
            "row": row,
            "column": merged.min_col,
            "rowspan": rowspan,
            "colspan": colspan,
            "clipped": clipped,
        }

        if clipped:
            anchor = cell_payload(
                sheet.cell(
                    row=merged.min_row,
                    column=merged.min_col,
                )
            )

            if anchor:
                entry["anchor"] = anchor

        merges.append(entry)

    return merges


def _read_columns(
    sheet: Any,
    column_count: int,
) -> list[dict[str, Any]]:
    """Per-column pixel widths for the table's <colgroup>.

    A column_dimensions entry spans a *range*, not one letter: ENGINE_STORE
    stores B as min=2,max=3 and LISTING stores D as min=4,max=16384. Reading
    them per letter silently loses widths, so expand each range.
    """
    default_width = getattr(
        sheet.sheet_format,
        "defaultColWidth",
        None,
    )

    widths: dict[int, float] = {}

    for dimension in sheet.column_dimensions.values():
        if dimension.width is None:
            continue

        start = max(1, dimension.min or 1)
        end = min(
            column_count,
            dimension.max or start,
        )

        for index in range(start, end + 1):
            widths[index] = dimension.width

    return [
        {
            "index": index,
            "letter": get_column_letter(index),
            "width_px": column_width_px(
                widths.get(index),
                default_width,
            ),
        }
        for index in range(1, column_count + 1)
    ]
