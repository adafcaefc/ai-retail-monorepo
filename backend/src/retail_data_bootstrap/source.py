from __future__ import annotations

import re
import unicodedata
from collections.abc import Iterator
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from .classification import SHEET_SPECS, validate_sheet_specs


def normalize_column_name(value: Any, index: int | None = None) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = text.encode("ascii", "ignore").decode("ascii").lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    if not text:
        return f"column_{index or 0}"
    if text[0].isdigit():
        return f"column_{text}"
    return text


def unique_columns(values: tuple[Any, ...]) -> list[str]:
    output: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(values, 1):
        base = normalize_column_name(value, index)
        seen[base] = seen.get(base, 0) + 1
        output.append(base if seen[base] == 1 else f"{base}_{seen[base]}")
    return output


def json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


class ExcelSourceAdapter:
    """Excel-specific extraction boundary returning plain normalized records."""

    def __init__(self, path: Path):
        self.path = path

    def open(self, *, data_only: bool = True, read_only: bool = True):
        return openpyxl.load_workbook(
            self.path,
            data_only=data_only,
            read_only=read_only,
        )

    def sheet_names(self) -> list[str]:
        workbook = self.open()
        try:
            names = list(workbook.sheetnames)
        finally:
            workbook.close()
        validate_sheet_specs(names)
        return names

    def rows(self, sheet_name: str) -> Iterator[dict[str, Any]]:
        spec = SHEET_SPECS[sheet_name]
        workbook = self.open(data_only=True)
        try:
            sheet = workbook[sheet_name]
            header_values = next(
                sheet.iter_rows(
                    min_row=spec.header_row,
                    max_row=spec.header_row,
                    values_only=True,
                )
            )
            columns = unique_columns(header_values)
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=spec.header_row + 1, values_only=True),
                spec.header_row + 1,
            ):
                if not any(value is not None and value != "" for value in values):
                    continue
                record = {
                    column: value
                    for column, value in zip(columns, values)
                    if not column.startswith("column_") or value is not None
                }
                record["_source_sheet"] = sheet_name
                record["_source_row"] = row_number
                yield record
        finally:
            workbook.close()
