from __future__ import annotations

import hashlib
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .classification import SHEET_SPECS
from .source import unique_columns


@dataclass(frozen=True)
class NormalizedDataset:
    workbook_name: str
    workbook_sha256: str
    tables: dict[str, list[dict[str, Any]]]
    issues: tuple[str, ...]

    @property
    def row_counts(self) -> dict[str, int]:
        return {name: len(rows) for name, rows in self.tables.items()}


SOURCE_SHEETS = (
    "Verticals",
    "Stores",
    "Categories",
    "Main Vendor",
    "Trade Agreement",
    "SKU_Master",
    "ENGINE",
    "ENGINE_STORE",
    "Replenishment Detail",
    "Promotion & Discount Detail",
    "Brand Events",
    "Workforce",
    "Time Series 24mo",
)

VERTICAL_ALIASES = {
    "Grocery": "GRC",
    "General Merch": "GMR",
    "Fashion": "FSH",
    "Health & Beauty": "HNB",
    "Electronics": "ELC",
    "Home & Living": "HNL",
    "Digital/Online": "DGT",
    "Omnichannel": "OMN",
}


def _records(path: Path) -> dict[str, list[dict[str, Any]]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        output: dict[str, list[dict[str, Any]]] = {}
        for sheet_name in SOURCE_SHEETS:
            sheet = workbook[sheet_name]
            spec = SHEET_SPECS[sheet_name]
            header = next(
                sheet.iter_rows(
                    min_row=spec.header_row,
                    max_row=spec.header_row,
                    values_only=True,
                )
            )
            columns = unique_columns(header)
            rows: list[dict[str, Any]] = []
            for row_number, values in enumerate(
                sheet.iter_rows(min_row=spec.header_row + 1, values_only=True),
                spec.header_row + 1,
            ):
                if not any(value is not None and value != "" for value in values):
                    continue
                row = {column: value for column, value in zip(columns, values)}
                row["source_sheet"] = sheet_name
                row["source_row"] = row_number
                rows.append(row)
            output[sheet_name] = rows
        return output
    finally:
        workbook.close()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return None if not result or result == "—" else result


def _flag(value: Any) -> bool:
    return str(value or "").strip().upper() in {"Y", "YES", "TRUE", "1"}


def _date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value).strip())


def _vendor_account(value: Any, vendor_by_code: dict[str, str]) -> str | None:
    code = _text(value)
    if code is None:
        return None
    if code not in vendor_by_code:
        raise ValueError(f"Vendor code is not present in Main Vendor: {code}")
    return vendor_by_code[code]


def normalize_workbook(path: Path) -> NormalizedDataset:
    source = _records(path)
    issues: list[str] = []

    entities = [
        {
            "legal_entity_id": row["id"],
            "legal_entity_name": row["vertical"],
            "short_name": row["short"],
            "workforce_base_per_size": row["wf_base_size"],
            "sales_per_fte": row["sales_per_fte"],
            "peak_season_factor": row["peak_season"],
            "total_store_size": row["store_size"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Verticals"]
    ]
    entity_ids = {row["legal_entity_id"] for row in entities}

    stores = [
        {
            "store_id": row["store_id"],
            "legal_entity_id": row["vertical"],
            "store_name": row["store_name"],
            "cluster": _text(row["cluster"]),
            "size_factor": row["size"],
            "health_factor": row["health"],
            "footfall_index": row["footfall_idx"],
            "channel": _text(row["channel"]),
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Stores"]
    ]
    categories = [
        {
            "category_id": row["cat_id"],
            "legal_entity_id": row["vertical"],
            "category_name": row["category"],
            "is_perishable": _flag(row["perishable"]),
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Categories"]
    ]
    vendors = [
        {
            "vendor_account": row["vendor_account"],
            "vendor_code": row["vendor"],
            "vendor_name": row["vendor_name"],
            "vendor_group": _text(row["group"]),
            "currency": _text(row["currency"]),
            "payment_terms": _text(row["payment_terms"]),
            "delivery_terms": _text(row["delivery_terms"]),
            "lead_time_days": row["lead_time_d"],
            "moq_units": row["moq_units"],
            "otif_pct": row["otif"],
            "fill_pct": row["fill"],
            "defect_pct": row["defect"],
            "lead_adherence_pct": row["lead_adherence"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Main Vendor"]
    ]
    vendor_by_code = {row["vendor_code"]: row["vendor_account"] for row in vendors}

    brands_first_row: dict[str, int] = {}
    for row in source["SKU_Master"]:
        brand = _text(row["brand"])
        if brand is not None:
            brands_first_row.setdefault(brand, row["source_row"])
    brands = [
        {
            "brand_name": brand,
            "source_sheet": "SKU_Master",
            "source_row": row_number,
        }
        for brand, row_number in sorted(brands_first_row.items())
    ]

    skus = [
        {
            "sku_id": row["sku_id"],
            "legal_entity_id": row["vertical"],
            "category_id": row["cat_id"],
            "item_name": row["item"],
            "is_perishable": _flag(row["perishable"]),
            "base_ads": row["base_ads"],
            "price": row["price"],
            "margin_pct": row["margin"],
            "cost": row["cost"],
            "lead_time_days": row["lead_d"],
            "on_hand_days": row["onhand_days"],
            "open_po_units": row["open_po"],
            "safety_days": row["safety_d"],
            "expiry_days": row["expiry_d"],
            "growth_factor": row["growth"],
            "elasticity": row["elasticity"],
            "competitor_index": row["comp_idx"],
            "funding_pct": row["fund"],
            "cannibalization_pct": row["cannib"],
            "is_promo": _flag(row["promo"]),
            "is_viral": _flag(row["viral"]),
            "sales_uom": _text(row["sales_uom"]),
            "buy_uom": _text(row["buy_uom"]),
            "pack_factor": row["pack_factor"],
            "channel": _text(row["channel"]),
            "seasonality_factor": row["seasonality"],
            "stock_factor": row["stockf"],
            "sales_per_fte": row["sales_fte"],
            "vendor_account": _vendor_account(row["vendor"], vendor_by_code),
            "brand_name": _text(row["brand"]),
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["SKU_Master"]
    ]

    agreements = [
        {
            "sku_id": row["item"],
            "vendor_account": row["vendor_account"],
            "valid_from": _date(row["valid_from"]),
            "min_quantity": row["min_qty_break"],
            "item_name": row["item_name"],
            "unit_price": row["unit_price"],
            "currency": row["currency"],
            "lead_time_days": row["lead_time_d"],
            "discount_pct": row["discount"],
            "valid_to": _date(row["valid_to"]),
            "is_designated": _flag(row["designated"]),
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Trade Agreement"]
    ]

    promotions: list[dict[str, Any]] = []
    for row in source["Promotion & Discount Detail"]:
        vertical_name = row["vertical"]
        if vertical_name not in VERTICAL_ALIASES:
            raise ValueError(f"Unmapped promotion vertical: {vertical_name}")
        promotions.append(
            {
                "promotion_id": row["promo_id"],
                "promotion_name": row["promo_name"],
                "discount_type": row["discount_type"],
                "scope": _text(row["scope"]),
                "legal_entity_id": VERTICAL_ALIASES[vertical_name],
                "target_category": _text(row["target_category"]),
                "season": _text(row["season"]),
                "peak_month": _text(row["peak_month"]),
                "mechanism": _text(row["mechanism"]),
                "discount_pct": row["discount"],
                "value_rule": _text(row["value_rule"]),
                "min_quantity_threshold": _text(row["min_qty_threshold"]),
                "supplier_funding_pct": row["supplier_funding"],
                "expected_uplift_pct": row["expected_uplift"],
                "prebuy_uplift_units": row["pre_buy_uplift_units"],
                "valid_from": _date(row["valid_from"]),
                "valid_to": _date(row["valid_to"]),
                "d365_construct": _text(row["d365_construct"]),
                "source_sheet": row["source_sheet"],
                "source_row": row["source_row"],
            }
        )

    inventory = [
        {
            "sku_id": row["sku"],
            "ads": row["ads"],
            "inventory_position": row["position"],
            "reorder_point": row["rop"],
            "max_inventory": row["max"],
            "days_of_supply": row["dos"],
            "inventory_state": _text(row["state"]),
            "price": row["price"],
            "inventory_value": row["inv_value"],
            "at_risk_value": row["at_risk"],
            "expiry_units": row["expiry_u"],
            "order_units": row["order_units"],
            "order_value": row["order_value"],
            "weekly_gmv": row["weekly_gmv"],
            "margin_amount": row["margin_rp"],
            "funding_amount": row["funding_rp"],
            "open_po_units": row["open_po"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["ENGINE"]
    ]

    store_sku = [
        {
            "sku_id": row["sku_id"],
            "store_id": row["store"],
            "ads": row["ads"],
            "on_hand_units": row["on_hand"],
            "open_po_units": row["open_po"],
            "inventory_position": row["position"],
            "reorder_point": row["rop"],
            "max_inventory": row["max"],
            "days_of_supply": row["dos"],
            "inventory_state": _text(row["state"]),
            "price": row["price"],
            "inventory_value": row["inv_value"],
            "at_risk_value": row.get("markdown_recoverable", row.get("at_risk_value", 0.0)),
            "forecast_7d": row["forecast_7d"],
            "order_sales_units": row["order_sales"],
            "pack_factor": row["pack"],
            "order_buy_units": row["order_buy"],
            "order_value": row["order_value"],
            "promo_incremental_margin": row["promo_incr_margin"],
            "contribution_per_day": row["contribution_day"],
            "labour_fte": row["labour_fte"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["ENGINE_STORE"]
    ]

    replenishment = [
        {
            "sku_id": row["item"],
            "reorder_required": str(row["reorder"]).upper() == "YES",
            "order_sales_units": row["order_qty_sales"],
            "buy_uom": _text(row["buy_uom"]),
            "order_buy_units": row["order_qty_buy"],
            "designated_vendor_account": _vendor_account(row["designated_vendor"], vendor_by_code),
            "designated_unit_price": row["unit_price_ta"],
            "amount": row["amount"],
            "best_price_vendor_account": _vendor_account(row["best_price_vendor"], vendor_by_code),
            "best_price": row["best_price"],
            "saving_vs_designated": row["saving_vs_designated"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Replenishment Detail"]
    ]

    events = [
        {
            "store_id": row["store"],
            "event_name": row["event"],
            "legal_entity_id": row["vertical"],
            "demand_lift": row["demand_lift"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in source["Brand Events"]
    ]

    workforce_source = source["Workforce"]
    workforce = [
        {
            "store_id": row["store"],
            "event_name": _text(row["event"]),
            "event_lift": row["event_lift"],
            "workforce_base": row["wf_base"],
            "peak_factor": row["peak"],
            "scheduled_fte": row["scheduled"],
            "required_fte": row["required"],
            "gap_fte": row["gap"],
            "surplus_fte": row["surplus"],
            "coverage_pct": row["coverage"],
            "source_sheet": row["source_sheet"],
            "source_row": row["source_row"],
        }
        for row in workforce_source
        if row["store"] != "TOTAL"
    ]
    if len(workforce) != len(workforce_source):
        issues.append(
            "Workforce row 166 is an aggregate TOTAL row and was intentionally excluded from the store-keyed table."
        )

    monthly_sales: list[dict[str, Any]] = []
    for row in source["Time Series 24mo"]:
        for entity_id in sorted(entity_ids):
            monthly_sales.append(
                {
                    "period_label": row["month"],
                    "legal_entity_id": entity_id,
                    "sales_amount": row[entity_id.lower()],
                    "source_sheet": row["source_sheet"],
                    "source_row": row["source_row"],
                }
            )

    tables = {
        "LegalEntity": entities,
        "Store": stores,
        "Category": categories,
        "Vendor": vendors,
        "Brand": brands,
        "Sku": skus,
        "TradeAgreement": agreements,
        "Promotion": promotions,
        "InventorySnapshot": inventory,
        "StoreSkuSnapshot": store_sku,
        "ReplenishmentProposal": replenishment,
        "BrandEvent": events,
        "WorkforceSnapshot": workforce,
        "MonthlySales": monthly_sales,
    }
    _validate_normalized(tables)
    return NormalizedDataset(
        workbook_name=path.name,
        workbook_sha256=hashlib.sha256(path.read_bytes()).hexdigest(),
        tables=tables,
        issues=tuple(issues),
    )


PRIMARY_KEYS: dict[str, tuple[str, ...]] = {
    "LegalEntity": ("legal_entity_id",),
    "Store": ("store_id",),
    "Category": ("category_id",),
    "Vendor": ("vendor_account",),
    "Brand": ("brand_name",),
    "Sku": ("sku_id",),
    "TradeAgreement": ("sku_id", "vendor_account", "valid_from", "min_quantity"),
    "Promotion": ("promotion_id",),
    "InventorySnapshot": ("sku_id",),
    "StoreSkuSnapshot": ("sku_id", "store_id"),
    "ReplenishmentProposal": ("sku_id",),
    "BrandEvent": ("store_id", "event_name"),
    "WorkforceSnapshot": ("store_id",),
    "MonthlySales": ("period_label", "legal_entity_id"),
}


def _validate_normalized(tables: dict[str, list[dict[str, Any]]]) -> None:
    for table, keys in PRIMARY_KEYS.items():
        seen: set[tuple[Any, ...]] = set()
        for row in tables[table]:
            key = tuple(row.get(column) for column in keys)
            if any(value is None or value == "" for value in key):
                raise ValueError(f"{table} has a null primary key: {key}")
            if key in seen:
                raise ValueError(f"{table} has a duplicate primary key: {key}")
            seen.add(key)

    relationships = (
        ("Store", "legal_entity_id", "LegalEntity", "legal_entity_id"),
        ("Category", "legal_entity_id", "LegalEntity", "legal_entity_id"),
        ("Sku", "legal_entity_id", "LegalEntity", "legal_entity_id"),
        ("Sku", "category_id", "Category", "category_id"),
        ("Sku", "vendor_account", "Vendor", "vendor_account"),
        ("Sku", "brand_name", "Brand", "brand_name"),
        ("TradeAgreement", "sku_id", "Sku", "sku_id"),
        ("TradeAgreement", "vendor_account", "Vendor", "vendor_account"),
        ("Promotion", "legal_entity_id", "LegalEntity", "legal_entity_id"),
        ("InventorySnapshot", "sku_id", "Sku", "sku_id"),
        ("StoreSkuSnapshot", "sku_id", "Sku", "sku_id"),
        ("StoreSkuSnapshot", "store_id", "Store", "store_id"),
        ("ReplenishmentProposal", "sku_id", "Sku", "sku_id"),
        ("BrandEvent", "store_id", "Store", "store_id"),
        ("BrandEvent", "legal_entity_id", "LegalEntity", "legal_entity_id"),
        ("WorkforceSnapshot", "store_id", "Store", "store_id"),
        ("MonthlySales", "legal_entity_id", "LegalEntity", "legal_entity_id"),
    )
    for child_table, child_column, parent_table, parent_column in relationships:
        parent_values = {row[parent_column] for row in tables[parent_table]}
        missing = {
            row[child_column]
            for row in tables[child_table]
            if row.get(child_column) is not None and row[child_column] not in parent_values
        }
        if missing:
            raise ValueError(
                f"{child_table}.{child_column} has values absent from "
                f"{parent_table}.{parent_column}: {sorted(missing, key=str)[:10]}"
            )
