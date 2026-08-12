from __future__ import annotations

import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from .classification import SHEET_SPECS, classification_counts, validate_sheet_specs
from .source import json_value, unique_columns


def _infer_type(values: list[Any]) -> str:
    present = [value for value in values if value is not None and value != ""]
    if not present:
        return "null"
    kinds: set[str] = set()
    for value in present:
        if isinstance(value, bool):
            kinds.add("boolean")
        elif isinstance(value, (datetime, date)):
            kinds.add("date")
        elif isinstance(value, int):
            kinds.add("integer")
        elif isinstance(value, float):
            kinds.add("number")
        else:
            kinds.add("string")
    if kinds <= {"integer", "number"}:
        return "number" if "number" in kinds else "integer"
    return next(iter(kinds)) if len(kinds) == 1 else "mixed"


def _candidate_result(
    columns: tuple[str, ...], rows: list[dict[str, Any]]
) -> dict[str, Any]:
    values: list[tuple[Any, ...]] = []
    null_rows = 0
    for row in rows:
        key = tuple(row.get(column) for column in columns)
        if any(value is None or value == "" for value in key):
            null_rows += 1
        else:
            values.append(key)
    counts = Counter(values)
    duplicates = [
        {"value": list(key), "count": count}
        for key, count in counts.items()
        if count > 1
    ]
    return {
        "columns": list(columns),
        "non_null_rows": len(values),
        "null_rows": null_rows,
        "unique": not duplicates and null_rows == 0 and bool(rows),
        "duplicate_count": sum(item["count"] - 1 for item in duplicates),
        "duplicate_examples": duplicates[:10],
    }


def inspect_workbook(path: Path) -> dict[str, Any]:
    import openpyxl

    formula_book = openpyxl.load_workbook(path, read_only=True, data_only=False)
    value_book = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        validate_sheet_specs(list(value_book.sheetnames))
        inventories: list[dict[str, Any]] = []
        value_sets: dict[str, dict[str, set[Any]]] = {}
        for sheet_name in value_book.sheetnames:
            spec = SHEET_SPECS[sheet_name]
            value_sheet = value_book[sheet_name]
            formula_sheet = formula_book[sheet_name]
            header = next(
                value_sheet.iter_rows(
                    min_row=spec.header_row,
                    max_row=spec.header_row,
                    values_only=True,
                )
            )
            columns = unique_columns(header)
            rows: list[dict[str, Any]] = []
            column_values: dict[str, list[Any]] = defaultdict(list)
            for row_number, values in enumerate(
                value_sheet.iter_rows(min_row=spec.header_row + 1, values_only=True),
                spec.header_row + 1,
            ):
                if not any(value is not None and value != "" for value in values):
                    continue
                row = {column: value for column, value in zip(columns, values)}
                row["_source_row"] = row_number
                rows.append(row)
                for column in columns:
                    column_values[column].append(row.get(column))

            formula_count = 0
            formula_examples: list[str] = []
            for formula_row in formula_sheet.iter_rows():
                for cell in formula_row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_count += 1
                        if len(formula_examples) < 5:
                            formula_examples.append(cell.coordinate)

            null_counts = {
                column: sum(value is None or value == "" for value in values)
                for column, values in column_values.items()
            }
            inferred_types = {
                column: _infer_type(values)
                for column, values in column_values.items()
            }
            candidates = [
                _candidate_result(key, rows)
                for key in spec.candidate_key_columns
                if all(column in columns for column in key)
            ]
            inventories.append(
                {
                    "sheet_name": sheet_name,
                    "worksheet_row_count": value_sheet.max_row,
                    "worksheet_column_count": value_sheet.max_column,
                    "header_row": spec.header_row,
                    "data_row_count": len(rows),
                    "column_names": columns,
                    "inferred_types": inferred_types,
                    "null_counts": null_counts,
                    "candidate_keys": candidates,
                    "formula_presence": formula_count > 0,
                    "formula_count": formula_count,
                    "formula_examples": formula_examples,
                    "representative_rows": [
                        {key: json_value(value) for key, value in row.items()}
                        for row in rows[:3]
                    ],
                    "classification": spec.classification,
                    "classification_reason": spec.reason,
                    "orientation": spec.orientation,
                }
            )
            value_sets[sheet_name] = {
                column: {
                    value
                    for value in values
                    if value is not None and value != ""
                    and isinstance(value, (str, int, float, bool, date, datetime))
                }
                for column, values in column_values.items()
            }

        relationships = _prove_relationships(value_sets)
        return {
            "workbook": path.name,
            "workbook_sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            "sheet_count": len(inventories),
            "classification_counts": classification_counts(),
            "sheets": inventories,
            "proven_relationships": relationships,
        }
    finally:
        formula_book.close()
        value_book.close()


RELATIONSHIP_CANDIDATES = (
    ("Stores", "vertical", "Verticals", "id"),
    ("Categories", "vertical", "Verticals", "id"),
    ("SKU_Master", "vertical", "Verticals", "id"),
    ("SKU_Master", "cat_id", "Categories", "cat_id"),
    ("Trade Agreement", "item", "SKU_Master", "sku_id"),
    ("Trade Agreement", "vendor_account", "Main Vendor", "vendor_account"),
    ("ENGINE", "sku", "SKU_Master", "sku_id"),
    ("ENGINE_STORE", "sku_id", "SKU_Master", "sku_id"),
    ("ENGINE_STORE", "store", "Stores", "store_id"),
    ("Brand Events", "store", "Stores", "store_id"),
    ("Workforce", "store", "Stores", "store_id"),
    ("Replenishment Detail", "item", "SKU_Master", "sku_id"),
)

RELATIONSHIP_EXCLUSIONS = {
    ("Workforce", "store"): {"TOTAL"},
}


def _prove_relationships(
    value_sets: dict[str, dict[str, set[Any]]]
) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for child_sheet, child_column, parent_sheet, parent_column in RELATIONSHIP_CANDIDATES:
        raw_child = value_sets.get(child_sheet, {}).get(child_column, set())
        excluded = raw_child & RELATIONSHIP_EXCLUSIONS.get((child_sheet, child_column), set())
        child = raw_child - excluded
        parent = value_sets.get(parent_sheet, {}).get(parent_column, set())
        missing = sorted(child - parent, key=str)
        output.append(
            {
                "child_sheet": child_sheet,
                "child_column": child_column,
                "parent_sheet": parent_sheet,
                "parent_column": parent_column,
                "child_distinct_count": len(child),
                "parent_distinct_count": len(parent),
                "matched": len(child & parent),
                "missing_count": len(missing),
                "missing_examples": [json_value(value) for value in missing[:10]],
                "excluded_aggregate_values": [json_value(value) for value in sorted(excluded, key=str)],
                "validated": bool(child) and not missing,
            }
        )
    return output


def write_inventory(report: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(
        json.dumps(report, ensure_ascii=False, indent=2, sort_keys=False) + "\n",
        encoding="utf-8",
    )


def readable_summary(report: dict[str, Any]) -> str:
    lines = [
        f"Workbook: {report['workbook']}",
        f"Sheets: {report['sheet_count']}",
        "Classification: "
        + ", ".join(
            f"{key}={value}"
            for key, value in report["classification_counts"].items()
        ),
        "",
    ]
    for sheet in report["sheets"]:
        keys = [
            "+".join(candidate["columns"])
            + (" (unique)" if candidate["unique"] else " (not unique)")
            for candidate in sheet["candidate_keys"]
        ]
        lines.append(
            f"- {sheet['sheet_name']}: {sheet['worksheet_row_count']}x"
            f"{sheet['worksheet_column_count']}, data={sheet['data_row_count']}, "
            f"{sheet['classification']}, formulas={sheet['formula_count']}, "
            f"keys={'; '.join(keys) or 'none declared'}"
        )
    validated = sum(r["validated"] for r in report["proven_relationships"])
    lines.extend(
        ["", f"Value-validated relationships: {validated}/{len(report['proven_relationships'])}"]
    )
    return "\n".join(lines)
