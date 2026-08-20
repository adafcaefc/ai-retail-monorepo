"""Generate the simplified 32-week demand dataset at SKU x Store grain.

The script is an offline/export step.  It performs a SELECT-only preflight and
read of the current v8.5 Azure SQL source, then writes the canonical wide CSV,
an XLSX review copy, a provenance manifest, and a generation report locally.
It never creates or changes Azure SQL objects or rows.

Run from the repository root, for example::

    .venv/bin/python backend/scripts/generate_demand_store_sku_32w.py \
        --test-result 'PASS: focused generator tests'

The pure generation and validation functions accept an in-memory
``SourceSnapshot`` so the generator can be tested without a database.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import statistics
import sys
from dataclasses import dataclass, replace
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT / "backend") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "backend"))


GENERATION_NAME = "demand_store_sku_32w_poc_v1"
GENERATOR_VERSION = "demand-store-sku-32w-generator-v1.0.0"
FIXED_SEED = 20260820
SOURCE_IMPORT_BATCH_ID = 23
SOURCE_SNAPSHOT_DATE = date(2026, 7, 1)
SOURCE_REVISION_SHA = "a6f4c7fabae4c27af9c4c035adb8252fa27cc062ea6b4bad4b0f81c976ec9510"
SOURCE_REVISION = f"v8.5:{SOURCE_REVISION_SHA}"
BUSINESS_TIMEZONE = "Asia/Jakarta"
DOW_SUM = Decimal("7.45")
QUANTITY_QUANTUM = Decimal("0.000001")
W1_RECONCILIATION_TOLERANCE = Decimal("0.000001")
EXPECTED_ROW_COUNT = 16_000
EXPECTED_STORE_COUNT = 160
EXPECTED_SKU_COUNT = 800
EXPECTED_ROWS_PER_STORE = 100
HISTORICAL_VALUE_COUNT = EXPECTED_ROW_COUNT * 16
SOURCE_W1_VALUE_COUNT = EXPECTED_ROW_COUNT
SYNTHETIC_FUTURE_VALUE_COUNT = EXPECTED_ROW_COUNT * 15
TOTAL_PERIOD_VALUE_COUNT = EXPECTED_ROW_COUNT * 32
MAX_ADJACENT_CHANGE = 0.20
MAX_ROW_TREND_ABS = 0.35
FLAT_HISTORY_CV_THRESHOLD = 0.001

ACTUAL_COLUMNS = tuple(f"actual_w{week}" for week in range(16, 0, -1))
FORECAST_COLUMNS = tuple(f"forecast_w{week}" for week in range(1, 17))
BUSINESS_COLUMNS = ("sku_id", "store_id", "cat", *ACTUAL_COLUMNS, *FORECAST_COLUMNS)


class GenerationError(RuntimeError):
    """Raised when generation or offline validation fails."""


class SourcePreflightError(GenerationError):
    """Raised when the live source is not the approved v8.5 revision."""


@dataclass(frozen=True)
class GeneratorParameters:
    """Versioned knobs for the deterministic, intentionally small POC model."""

    generation_name: str = GENERATION_NAME
    generator_version: str = GENERATOR_VERSION
    seed: int = FIXED_SEED
    source_snapshot_date: date = SOURCE_SNAPSHOT_DATE

    def manifest_dict(self) -> dict[str, Any]:
        return {
            "generation_name": self.generation_name,
            "generator_version": self.generator_version,
            "seed": self.seed,
            "source_snapshot_date": self.source_snapshot_date.isoformat(),
            "business_timezone": BUSINESS_TIMEZONE,
            "dow_sum": decimal_text(DOW_SUM),
            "quantity_precision": 6,
            "model": {
                "history_level": "bounded stable SKU-store level around 0.94-1.02",
                "growth_rate": "bounded stable SKU-store rate in [-1.5%, +1.5%]",
                "seasonality": "two low-amplitude smooth sinusoidal components",
                "noise": "smooth deterministic recurrence 0.72 previous / 0.28 new",
                "baseline": "average(ads * 7.45, exact source forecast_7d)",
                "future_anchor": "exact quantized source forecast_7d at forecast_w1",
            },
        }


@dataclass(frozen=True)
class SourceSkuStore:
    """One exact v8.5 source fact row plus the runtime modelling context."""

    sku_id: str
    store_id: str
    cat: str
    ads: Decimal
    forecast_7d: Decimal
    vertical_id: str
    size_index: Decimal | None = None
    health_index: Decimal | None = None
    footfall_index: Decimal | None = None
    cluster: str | None = None
    channel: str | None = None
    seasonality_index: Decimal | None = None
    growth_index: Decimal | None = None
    is_promo_eligible: bool = False
    cannibalisation_pct: Decimal | None = None
    is_viral: bool = False

    @property
    def key(self) -> tuple[str, str]:
        return self.sku_id, self.store_id

    def fingerprint_dict(self) -> dict[str, Any]:
        return {
            "sku_id": self.sku_id,
            "store_id": self.store_id,
            "cat": self.cat,
            "ads": decimal_text(self.ads),
            "forecast_7d": decimal_text(self.forecast_7d),
            "vertical_id": self.vertical_id,
            "size_index": decimal_or_none(self.size_index),
            "health_index": decimal_or_none(self.health_index),
            "footfall_index": decimal_or_none(self.footfall_index),
            "cluster": self.cluster,
            "channel": self.channel,
            "seasonality_index": decimal_or_none(self.seasonality_index),
            "growth_index": decimal_or_none(self.growth_index),
            "is_promo_eligible": self.is_promo_eligible,
            "cannibalisation_pct": decimal_or_none(self.cannibalisation_pct),
            "is_viral": self.is_viral,
        }


@dataclass(frozen=True)
class SourceSnapshot:
    """Validated source rows and read-only preflight evidence."""

    source_rows: tuple[SourceSkuStore, ...]
    batch_id: int = SOURCE_IMPORT_BATCH_ID
    agent_name: str = "retail_facts_seed"
    workbook_version: str = "v8.5"
    import_status: str = "COMPLETED"
    source_revision: str = SOURCE_REVISION
    source_sha256: str = SOURCE_REVISION_SHA
    source_snapshot_date: date = SOURCE_SNAPSHOT_DATE
    batch_total_rows: int = 36_440
    fact_rows: int = EXPECTED_ROW_COUNT
    fact_dates: int = 1
    fact_min_date: date = SOURCE_SNAPSHOT_DATE
    fact_max_date: date = SOURCE_SNAPSHOT_DATE
    fact_distinct_stores: int = EXPECTED_STORE_COUNT
    fact_distinct_skus: int = EXPECTED_SKU_COUNT
    duplicate_source_keys: int = 0
    bad_ads: int = 0
    bad_forecast_7d: int = 0
    rows_not_batch: int = 0
    dim_store_rows: int = EXPECTED_STORE_COUNT
    dim_store_distinct_ids: int = EXPECTED_STORE_COUNT
    dim_item_rows: int = EXPECTED_SKU_COUNT
    dim_item_distinct_ids: int = EXPECTED_SKU_COUNT
    dim_item_duplicate_groups: int = 0
    dim_item_null_categories: int = 0
    category_join_rows: int = EXPECTED_ROW_COUNT
    category_join_skus: int = EXPECTED_SKU_COUNT
    category_join_stores: int = EXPECTED_STORE_COUNT
    missing_category_items: int = 0
    null_join_categories: int = 0
    missing_join_stores: int = 0

    @property
    def source_key_set(self) -> set[tuple[str, str]]:
        return {row.key for row in self.source_rows}

    @property
    def store_ids(self) -> tuple[str, ...]:
        return tuple(sorted({row.store_id for row in self.source_rows}))

    @property
    def sku_ids(self) -> tuple[str, ...]:
        return tuple(sorted({row.sku_id for row in self.source_rows}))

    @property
    def total_forecast_7d(self) -> Decimal:
        return sum((row.forecast_7d for row in self.source_rows), Decimal("0"))

    @property
    def store_verticals(self) -> dict[str, str]:
        result: dict[str, str] = {}
        for row in self.source_rows:
            previous = result.setdefault(row.store_id, row.vertical_id)
            if previous != row.vertical_id:
                raise SourcePreflightError(
                    f"Store {row.store_id} resolves to multiple verticals"
                )
        return result

    def fingerprint_dict(self) -> dict[str, Any]:
        return {
            "source_revision": self.source_revision,
            "source_sha256": self.source_sha256,
            "source_import_batch": self.batch_id,
            "source_snapshot_date": self.source_snapshot_date.isoformat(),
            "fact_rows": self.fact_rows,
            "fact_dates": self.fact_dates,
            "fact_distinct_stores": self.fact_distinct_stores,
            "fact_distinct_skus": self.fact_distinct_skus,
            "source_rows": [
                row.fingerprint_dict()
                for row in sorted(self.source_rows, key=lambda item: item.key)
            ],
        }


@dataclass(frozen=True)
class DemandSkuStoreRow:
    """One canonical wide output row at SKU x Store grain."""

    sku_id: str
    store_id: str
    cat: str
    actuals: tuple[Decimal, ...]  # actual_w16 ... actual_w1
    forecasts: tuple[Decimal, ...]  # forecast_w1 ... forecast_w16

    @property
    def key(self) -> tuple[str, str]:
        return self.sku_id, self.store_id

    def actual_for_week(self, week: int) -> Decimal:
        if week < 1 or week > 16:
            raise ValueError(f"Actual week must be 1..16, got {week}")
        return self.actuals[16 - week]

    def forecast_for_week(self, week: int) -> Decimal:
        if week < 1 or week > 16:
            raise ValueError(f"Forecast week must be 1..16, got {week}")
        return self.forecasts[week - 1]

    def business_dict(self) -> dict[str, str]:
        values: dict[str, str] = {
            "sku_id": self.sku_id,
            "store_id": self.store_id,
            "cat": self.cat,
        }
        values.update(
            {
                column: quantity_text(value)
                for column, value in zip(ACTUAL_COLUMNS, self.actuals)
            }
        )
        values.update(
            {
                column: quantity_text(value)
                for column, value in zip(FORECAST_COLUMNS, self.forecasts)
            }
        )
        return values


def decimal_value(value: Any) -> Decimal:
    """Convert a database/test value without introducing binary float math."""

    if isinstance(value, Decimal):
        result = value
    else:
        result = Decimal(str(value))
    if not result.is_finite():
        raise GenerationError(f"Non-finite numeric value: {value!r}")
    return result


def quantize_qty(value: Decimal | float | int) -> Decimal:
    """Apply one deterministic six-decimal final quantisation."""

    try:
        result = decimal_value(value).quantize(
            QUANTITY_QUANTUM, rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, ValueError) as exc:
        raise GenerationError(f"Unable to quantize quantity {value!r}") from exc
    if result < 0:
        raise GenerationError(f"Negative generated quantity: {result}")
    return result


def decimal_text(value: Decimal | float | int) -> str:
    return format(decimal_value(value), "f")


def quantity_text(value: Decimal | float | int) -> str:
    return format(quantize_qty(value), "f")


def decimal_or_none(value: Decimal | None) -> str | None:
    return decimal_text(value) if value is not None else None


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _row_mapping(row: Any) -> Mapping[str, Any]:
    mapping = getattr(row, "_mapping", None)
    if mapping is not None:
        return mapping
    if isinstance(row, Mapping):
        return row
    raise TypeError(f"Unsupported database row type: {type(row)!r}")


def _parse_metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if value is None:
        return {}
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise SourcePreflightError("Batch metadata is not valid JSON") from exc
    if not isinstance(parsed, dict):
        raise SourcePreflightError("Batch metadata is not a JSON object")
    return parsed


def _as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    if isinstance(value, (int, float, Decimal)):
        return bool(value)
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on"}


def _optional_decimal(value: Any) -> Decimal | None:
    return decimal_value(value) if value is not None else None


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise SourcePreflightError(message)


def validate_source_snapshot(snapshot: SourceSnapshot) -> None:
    """Validate source lineage, population, uniqueness, and category mapping."""

    _require(snapshot.batch_id == SOURCE_IMPORT_BATCH_ID, "Expected import batch 23")
    _require(snapshot.workbook_version == "v8.5", "Expected source workbook v8.5")
    _require(snapshot.import_status == "COMPLETED", "Expected a completed source batch")
    _require(snapshot.agent_name == "retail_facts_seed", "Unexpected source batch agent")
    _require(snapshot.source_revision == SOURCE_REVISION, "Unexpected source revision")
    _require(
        snapshot.source_snapshot_date == SOURCE_SNAPSHOT_DATE,
        "Expected source snapshot date 2026-07-01",
    )
    _require(snapshot.fact_rows == EXPECTED_ROW_COUNT, "Expected 16,000 fact rows")
    _require(
        snapshot.fact_distinct_stores == EXPECTED_STORE_COUNT,
        "Expected 160 fact stores",
    )
    _require(
        snapshot.fact_distinct_skus == EXPECTED_SKU_COUNT,
        "Expected 800 fact SKUs",
    )
    _require(
        snapshot.fact_dates == 1
        and snapshot.fact_min_date == SOURCE_SNAPSHOT_DATE
        and snapshot.fact_max_date == SOURCE_SNAPSHOT_DATE,
        "Expected only the 2026-07-01 fact snapshot",
    )
    _require(snapshot.duplicate_source_keys == 0, "Duplicate source SKU-store keys found")
    _require(snapshot.bad_ads == 0, "Source contains null or negative ADS values")
    _require(
        snapshot.bad_forecast_7d == 0,
        "Source contains null or negative Forecast 7d values",
    )
    _require(snapshot.rows_not_batch == 0, "Fact rows are not all tied to batch 23")
    _require(
        snapshot.dim_store_rows == EXPECTED_STORE_COUNT
        and snapshot.dim_store_distinct_ids == EXPECTED_STORE_COUNT,
        "Expected 160 runtime stores",
    )
    _require(
        snapshot.dim_item_rows == EXPECTED_SKU_COUNT
        and snapshot.dim_item_distinct_ids == EXPECTED_SKU_COUNT,
        "Expected 800 runtime SKUs",
    )
    _require(snapshot.dim_item_duplicate_groups == 0, "Duplicate dim_item IDs found")
    _require(snapshot.dim_item_null_categories == 0, "Null dim_item categories found")
    _require(
        snapshot.category_join_rows == EXPECTED_ROW_COUNT
        and snapshot.category_join_skus == EXPECTED_SKU_COUNT
        and snapshot.category_join_stores == EXPECTED_STORE_COUNT
        and snapshot.missing_category_items == 0
        and snapshot.null_join_categories == 0
        and snapshot.missing_join_stores == 0,
        "Not every fact row resolves to exactly one store and category",
    )

    _require(
        len(snapshot.source_rows) == EXPECTED_ROW_COUNT,
        f"Expected 16,000 detailed source rows, found {len(snapshot.source_rows)}",
    )
    _require(
        len(snapshot.source_key_set) == EXPECTED_ROW_COUNT,
        "Expected 16,000 unique source SKU-store pairs",
    )
    store_rows: dict[str, list[SourceSkuStore]] = {}
    for source in snapshot.source_rows:
        _require(bool(source.sku_id.strip()), "Blank source SKU ID")
        _require(bool(source.store_id.strip()), "Blank source store ID")
        _require(bool(source.cat.strip()), f"Null/blank category for {source.key}")
        _require(source.ads.is_finite() and source.ads >= 0, f"Invalid ADS for {source.key}")
        _require(
            source.forecast_7d.is_finite() and source.forecast_7d >= 0,
            f"Invalid Forecast 7d for {source.key}",
        )
        store_rows.setdefault(source.store_id, []).append(source)

    _require(
        len(store_rows) == EXPECTED_STORE_COUNT,
        f"Expected 160 stores in source rows, found {len(store_rows)}",
    )
    _require(
        all(len(rows) == EXPECTED_ROWS_PER_STORE for rows in store_rows.values()),
        "Every store must have exactly 100 source SKU rows",
    )
    _require(
        tuple(sorted(store_rows)) == tuple(f"S{i:03d}" for i in range(1, 161)),
        "Runtime store population is not S001 through S160",
    )
    _require(
        all(len({row.sku_id for row in rows}) == EXPECTED_ROWS_PER_STORE
            for rows in store_rows.values()),
        "A store has duplicate source SKUs",
    )
    verticals = snapshot.store_verticals
    _require(len(verticals) == EXPECTED_STORE_COUNT, "Every store must resolve one vertical")


def load_source_snapshot() -> SourceSnapshot:
    """Run the mandatory SELECT-only Azure SQL preflight and fetch source rows."""

    from sqlalchemy import text

    from src.db.db import get_engine

    batch_sql = text(
        """
        SELECT id, agent_name, workbook_version, import_status, total_rows,
               CAST(metadata AS nvarchar(max)) AS metadata
        FROM audit.import_batches
        WHERE id = :batch_id
        """
    )
    fact_sql = text(
        """
        SELECT
            COUNT(*) AS fact_rows,
            COUNT(DISTINCT store_key) AS fact_distinct_stores,
            COUNT(DISTINCT item_key) AS fact_distinct_skus,
            COUNT(DISTINCT cal_date) AS fact_dates,
            MIN(cal_date) AS fact_min_date,
            MAX(cal_date) AS fact_max_date,
            SUM(CASE WHEN ads IS NULL OR ads < 0 THEN 1 ELSE 0 END) AS bad_ads,
            SUM(CASE WHEN forecast_7d IS NULL OR forecast_7d < 0 THEN 1 ELSE 0 END)
                AS bad_forecast_7d,
            COUNT(DISTINCT import_batch_id) AS fact_batch_ids,
            MIN(import_batch_id) AS min_batch_id,
            MAX(import_batch_id) AS max_batch_id,
            SUM(CASE WHEN import_batch_id IS NULL OR import_batch_id <> :batch_id
                     THEN 1 ELSE 0 END) AS rows_not_batch
        FROM retail.fact_inventory_daily
        """
    )
    duplicate_sql = text(
        """
        SELECT COUNT(*) AS duplicate_source_keys
        FROM (
            SELECT item_key, store_key, cal_date
            FROM retail.fact_inventory_daily
            GROUP BY item_key, store_key, cal_date
            HAVING COUNT(*) > 1
        ) AS duplicate_keys
        """
    )
    store_sql = text(
        """
        SELECT store_key, COUNT(*) AS source_rows,
               COUNT(DISTINCT item_key) AS distinct_skus
        FROM retail.fact_inventory_daily
        WHERE cal_date = :snapshot_date
        GROUP BY store_key
        """
    )
    dimensions_sql = text(
        """
        SELECT
            (SELECT COUNT(*) FROM retail.dim_store) AS dim_store_rows,
            (SELECT COUNT(DISTINCT store_id) FROM retail.dim_store)
                AS dim_store_distinct_ids,
            (SELECT COUNT(*) FROM retail.dim_item) AS dim_item_rows,
            (SELECT COUNT(DISTINCT item_id) FROM retail.dim_item)
                AS dim_item_distinct_ids,
            (SELECT SUM(CASE WHEN category_id IS NULL THEN 1 ELSE 0 END)
             FROM retail.dim_item) AS dim_item_null_categories
        """
    )
    item_duplicate_sql = text(
        """
        SELECT COUNT(*) AS dim_item_duplicate_groups
        FROM (
            SELECT item_id
            FROM retail.dim_item
            GROUP BY item_id
            HAVING COUNT(*) <> 1
        ) AS duplicate_items
        """
    )
    category_join_sql = text(
        """
        SELECT
            COUNT(*) AS category_join_rows,
            COUNT(DISTINCT f.item_key) AS category_join_skus,
            COUNT(DISTINCT f.store_key) AS category_join_stores,
            SUM(CASE WHEN i.item_id IS NULL THEN 1 ELSE 0 END)
                AS missing_category_items,
            SUM(CASE WHEN i.category_id IS NULL THEN 1 ELSE 0 END)
                AS null_join_categories,
            SUM(CASE WHEN s.store_id IS NULL THEN 1 ELSE 0 END)
                AS missing_join_stores
        FROM retail.fact_inventory_daily AS f
        LEFT JOIN retail.dim_item AS i ON i.item_id = f.item_key
        LEFT JOIN retail.dim_store AS s ON s.store_id = f.store_key
        WHERE f.cal_date = :snapshot_date
        """
    )
    detail_sql = text(
        """
        SELECT
            f.item_key AS sku_id,
            f.store_key AS store_id,
            i.category_id AS cat,
            f.ads,
            f.forecast_7d,
            s.vertical_id,
            s.size_index,
            s.health_index,
            s.footfall_index,
            s.cluster,
            s.channel,
            i.seasonality_index,
            i.growth_index,
            i.is_promo_eligible,
            i.cannibalisation_pct,
            i.is_viral
        FROM retail.fact_inventory_daily AS f
        JOIN retail.dim_item AS i ON i.item_id = f.item_key
        JOIN retail.dim_store AS s ON s.store_id = f.store_key
        WHERE f.cal_date = :snapshot_date
          AND f.import_batch_id = :batch_id
        ORDER BY f.item_key, f.store_key
        """
    )

    with get_engine().connect() as connection:
        batch_row = connection.execute(
            batch_sql, {"batch_id": SOURCE_IMPORT_BATCH_ID}
        ).mappings().first()
        fact_row = connection.execute(
            fact_sql, {"batch_id": SOURCE_IMPORT_BATCH_ID}
        ).mappings().one()
        duplicate_row = connection.execute(duplicate_sql).mappings().one()
        store_rows = connection.execute(
            store_sql, {"snapshot_date": SOURCE_SNAPSHOT_DATE}
        ).mappings().all()
        dimensions_row = connection.execute(dimensions_sql).mappings().one()
        item_duplicate_row = connection.execute(item_duplicate_sql).mappings().one()
        category_join_row = connection.execute(
            category_join_sql, {"snapshot_date": SOURCE_SNAPSHOT_DATE}
        ).mappings().one()
        detail_rows = connection.execute(
            detail_sql,
            {
                "snapshot_date": SOURCE_SNAPSHOT_DATE,
                "batch_id": SOURCE_IMPORT_BATCH_ID,
            },
        ).mappings().all()

    _require(batch_row is not None, "Expected audit.import_batches.id=23 was not found")
    batch = _row_mapping(batch_row)
    metadata = _parse_metadata(batch["metadata"])
    source_sha256 = str(metadata.get("workbook_sha256", ""))
    _require(
        str(batch["workbook_version"]) == "v8.5",
        f"Source drift: batch 23 workbook_version={batch['workbook_version']!r}",
    )
    _require(
        str(batch["import_status"]) == "COMPLETED",
        f"Source drift: batch 23 import_status={batch['import_status']!r}",
    )
    _require(
        str(batch["agent_name"]) == "retail_facts_seed",
        f"Source drift: batch 23 agent_name={batch['agent_name']!r}",
    )
    _require(
        source_sha256 == SOURCE_REVISION_SHA,
        "Source drift: batch 23 workbook SHA is not the approved v8.5 revision",
    )

    fact_rows = int(fact_row["fact_rows"] or 0)
    fact_min_date = (
        _parse_date(fact_row["fact_min_date"])
        if fact_row["fact_min_date"] is not None
        else date.min
    )
    fact_max_date = (
        _parse_date(fact_row["fact_max_date"])
        if fact_row["fact_max_date"] is not None
        else date.min
    )
    _require(fact_rows == EXPECTED_ROW_COUNT, f"Expected 16,000 fact rows, found {fact_rows}")
    _require(
        int(fact_row["fact_batch_ids"] or 0) == 1
        and int(fact_row["min_batch_id"] or 0) == SOURCE_IMPORT_BATCH_ID
        and int(fact_row["max_batch_id"] or 0) == SOURCE_IMPORT_BATCH_ID
        and int(fact_row["rows_not_batch"] or 0) == 0,
        "Fact rows are not all tied to import batch 23",
    )
    _require(
        len(store_rows) == EXPECTED_STORE_COUNT
        and all(
            int(_row_mapping(row)["source_rows"] or 0) == EXPECTED_ROWS_PER_STORE
            and int(_row_mapping(row)["distinct_skus"] or 0)
            == EXPECTED_ROWS_PER_STORE
            for row in store_rows
        ),
        "Every source store must have 100 unique SKU rows",
    )

    sources: list[SourceSkuStore] = []
    for row in detail_rows:
        mapping = _row_mapping(row)
        category = mapping["cat"]
        _require(category is not None and str(category).strip(), "Null source category")
        sources.append(
            SourceSkuStore(
                sku_id=str(mapping["sku_id"]),
                store_id=str(mapping["store_id"]),
                cat=str(category),
                ads=decimal_value(mapping["ads"]),
                forecast_7d=decimal_value(mapping["forecast_7d"]),
                vertical_id=str(mapping["vertical_id"]),
                size_index=_optional_decimal(mapping["size_index"]),
                health_index=_optional_decimal(mapping["health_index"]),
                footfall_index=_optional_decimal(mapping["footfall_index"]),
                cluster=(str(mapping["cluster"]) if mapping["cluster"] is not None else None),
                channel=(str(mapping["channel"]) if mapping["channel"] is not None else None),
                seasonality_index=_optional_decimal(mapping["seasonality_index"]),
                growth_index=_optional_decimal(mapping["growth_index"]),
                is_promo_eligible=_as_bool(mapping["is_promo_eligible"]),
                cannibalisation_pct=_optional_decimal(mapping["cannibalisation_pct"]),
                is_viral=_as_bool(mapping["is_viral"]),
            )
        )

    snapshot = SourceSnapshot(
        source_rows=tuple(sources),
        batch_id=SOURCE_IMPORT_BATCH_ID,
        agent_name=str(batch["agent_name"]),
        workbook_version=str(batch["workbook_version"]),
        import_status=str(batch["import_status"]),
        source_revision=SOURCE_REVISION,
        source_sha256=source_sha256,
        source_snapshot_date=SOURCE_SNAPSHOT_DATE,
        batch_total_rows=int(batch["total_rows"] or 0),
        fact_rows=fact_rows,
        fact_dates=int(fact_row["fact_dates"] or 0),
        fact_min_date=fact_min_date,
        fact_max_date=fact_max_date,
        fact_distinct_stores=int(fact_row["fact_distinct_stores"] or 0),
        fact_distinct_skus=int(fact_row["fact_distinct_skus"] or 0),
        duplicate_source_keys=int(duplicate_row["duplicate_source_keys"] or 0),
        bad_ads=int(fact_row["bad_ads"] or 0),
        bad_forecast_7d=int(fact_row["bad_forecast_7d"] or 0),
        rows_not_batch=int(fact_row["rows_not_batch"] or 0),
        dim_store_rows=int(dimensions_row["dim_store_rows"] or 0),
        dim_store_distinct_ids=int(dimensions_row["dim_store_distinct_ids"] or 0),
        dim_item_rows=int(dimensions_row["dim_item_rows"] or 0),
        dim_item_distinct_ids=int(dimensions_row["dim_item_distinct_ids"] or 0),
        dim_item_duplicate_groups=int(
            item_duplicate_row["dim_item_duplicate_groups"] or 0
        ),
        dim_item_null_categories=int(
            dimensions_row["dim_item_null_categories"] or 0
        ),
        category_join_rows=int(category_join_row["category_join_rows"] or 0),
        category_join_skus=int(category_join_row["category_join_skus"] or 0),
        category_join_stores=int(category_join_row["category_join_stores"] or 0),
        missing_category_items=int(
            category_join_row["missing_category_items"] or 0
        ),
        null_join_categories=int(category_join_row["null_join_categories"] or 0),
        missing_join_stores=int(category_join_row["missing_join_stores"] or 0),
    )
    validate_source_snapshot(snapshot)
    return snapshot


def _clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def _stable_unit(seed: int, *parts: object) -> float:
    token = "|".join([str(seed), *(str(part) for part in parts)])
    digest = hashlib.sha256(token.encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") / float(2**64 - 1)


def _stable_signed(seed: int, *parts: object) -> float:
    return 2.0 * _stable_unit(seed, *parts) - 1.0


def _normalised_feature(value: Decimal | None) -> float:
    if value is None:
        return 1.0
    numeric = float(value)
    if not math.isfinite(numeric):
        return 1.0
    return _clamp(numeric, 0.70, 1.30)


def _model_parameters(
    source: SourceSkuStore,
    params: GeneratorParameters,
) -> dict[str, float | Decimal]:
    """Build row-specific, bounded model parameters from approved inputs."""

    size = _normalised_feature(source.size_index)
    health = _normalised_feature(source.health_index)
    footfall = _normalised_feature(source.footfall_index)
    store_context = 0.45 * size + 0.25 * health + 0.30 * footfall
    seasonality = _normalised_feature(source.seasonality_index)
    growth_index = _normalised_feature(source.growth_index)
    category_unit = _stable_unit(params.seed, "category", source.cat)
    store_unit = _stable_unit(params.seed, "store", source.store_id)
    row_unit = _stable_unit(params.seed, "row", source.sku_id, source.store_id)

    growth_rate = _clamp(
        0.006 * (store_context - 1.0)
        + 0.025 * (growth_index - 1.0)
        + 0.003 * (seasonality - 1.0)
        + 0.010 * _stable_signed(params.seed, "growth", source.sku_id, source.store_id)
        + (0.0015 if source.is_viral else 0.0)
        - (0.0005 if source.is_promo_eligible else 0.0),
        -0.015,
        0.015,
    )
    history_level = _clamp(
        0.94
        + 0.035 * row_unit
        + 0.015 * (store_context - 1.0)
        + 0.005 * (category_unit - 0.5),
        0.90,
        1.02,
    )
    season_amplitude = 0.010 + 0.014 * _stable_unit(
        params.seed, "season-amplitude", source.sku_id, source.store_id
    )
    secondary_amplitude = 0.004 + 0.006 * _stable_unit(
        params.seed, "secondary-amplitude", source.sku_id, source.store_id
    )
    noise_amplitude = 0.004 + 0.006 * _stable_unit(
        params.seed, "noise-amplitude", source.sku_id, source.store_id
    )
    phase = 2.0 * math.pi * _stable_unit(
        params.seed, "phase", source.sku_id, source.store_id, source.cat
    )
    secondary_phase = 2.0 * math.pi * _stable_unit(
        params.seed, "secondary-phase", source.store_id, source.cat
    )

    ads_weekly = source.ads * DOW_SUM
    if source.forecast_7d > 0 and ads_weekly > 0:
        baseline = (source.forecast_7d + ads_weekly) / Decimal("2")
    else:
        baseline = max(source.forecast_7d, ads_weekly)
    if baseline < 0:
        raise GenerationError(f"Negative source baseline for {source.key}")

    return {
        "baseline": baseline,
        "growth_rate": growth_rate,
        "history_level": history_level,
        "season_amplitude": season_amplitude,
        "secondary_amplitude": secondary_amplitude,
        "noise_amplitude": noise_amplitude,
        "phase": phase,
        "secondary_phase": secondary_phase,
    }


def _seasonal_signal(period: int, model: Mapping[str, float | Decimal]) -> float:
    first = float(model["season_amplitude"]) * math.sin(
        2.0 * math.pi * (period + float(model["phase"])) / 13.0
    )
    second = float(model["secondary_amplitude"]) * math.cos(
        2.0 * math.pi * (period + float(model["secondary_phase"])) / 26.0
    )
    return first + second


def _smooth_noise(
    source: SourceSkuStore,
    params: GeneratorParameters,
    block: str,
    periods: Iterable[int],
) -> dict[int, float]:
    previous = 0.0
    result: dict[int, float] = {}
    for period in periods:
        raw = _stable_signed(
            params.seed, "noise", block, source.sku_id, source.store_id, period
        )
        previous = 0.72 * previous + 0.28 * raw
        result[period] = previous
    return result


def _relative_factor(
    period: int,
    reference_period: int,
    model: Mapping[str, float | Decimal],
    noise: Mapping[int, float],
    growth_exponent: int,
) -> float:
    growth_rate = float(model["growth_rate"])
    growth_factor = (1.0 + growth_rate) ** growth_exponent
    season_ratio = (1.0 + _seasonal_signal(period, model)) / (
        1.0 + _seasonal_signal(reference_period, model)
    )
    noise_ratio = 1.0 + float(model["noise_amplitude"]) * (
        noise[period] - noise[reference_period]
    )
    factor = growth_factor * season_ratio * noise_ratio
    return max(0.0, factor)


def generate_dataset(
    snapshot: SourceSnapshot,
    params: GeneratorParameters | None = None,
) -> list[DemandSkuStoreRow]:
    """Generate exactly one wide row per source SKU x Store pair."""

    params = params or GeneratorParameters()
    validate_source_snapshot(snapshot)
    if params.source_snapshot_date != snapshot.source_snapshot_date:
        raise GenerationError("Generator/source snapshot dates do not match")
    if params.generation_name == "":
        raise GenerationError("Generation name must not be blank")

    rows: list[DemandSkuStoreRow] = []
    for source in sorted(snapshot.source_rows, key=lambda item: item.key):
        model = _model_parameters(source, params)
        historical_noise = _smooth_noise(
            source, params, "historical", range(-16, 0)
        )
        future_noise = _smooth_noise(source, params, "future", range(1, 17))
        historical_reference_noise = historical_noise[-1]
        future_reference_noise = future_noise[1]
        historical_actuals: list[Decimal] = []
        future_forecasts: list[Decimal] = [quantize_qty(source.forecast_7d)]

        for week in range(-16, 0):
            factor = _relative_factor(
                week,
                -1,
                model,
                historical_noise,
                week + 1,
            )
            # The source ADS/Forecast scale is only a modelling anchor.  It is
            # never labelled as genuine historical sales.
            quantity = float(model["baseline"]) * float(model["history_level"]) * factor
            historical_actuals.append(quantize_qty(quantity))

        for week in range(2, 17):
            factor = _relative_factor(
                week,
                1,
                model,
                future_noise,
                week - 1,
            )
            # W+1 is deliberately absent from this model path: it remains the
            # exact source Forecast 7d for this SKU and Store pair.
            quantity = float(source.forecast_7d) * factor
            future_forecasts.append(quantize_qty(quantity))

        rows.append(
            DemandSkuStoreRow(
                sku_id=source.sku_id,
                store_id=source.store_id,
                cat=source.cat,
                actuals=tuple(historical_actuals),
                forecasts=tuple(future_forecasts),
            )
        )

    rows.sort(key=lambda row: row.key)
    return rows


def canonical_row_text(row: DemandSkuStoreRow) -> str:
    values = row.business_dict()
    return "|".join(values[column] for column in BUSINESS_COLUMNS)


def output_fingerprint(rows: Iterable[DemandSkuStoreRow]) -> str:
    digest = hashlib.sha256()
    for row in sorted(rows, key=lambda item: item.key):
        digest.update(canonical_row_text(row).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def input_fingerprint(
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
) -> str:
    payload = {
        "source": snapshot.fingerprint_dict(),
        "parameters": params.manifest_dict(),
        "output_columns": list(BUSINESS_COLUMNS),
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _rows_by_key(rows: Sequence[DemandSkuStoreRow]) -> dict[tuple[str, str], DemandSkuStoreRow]:
    result: dict[tuple[str, str], DemandSkuStoreRow] = {}
    for row in rows:
        if row.key in result:
            raise GenerationError(f"Duplicate output SKU-store key {row.key}")
        result[row.key] = row
    return result


def trend_totals(
    rows: Sequence[DemandSkuStoreRow],
    selected_keys: Iterable[tuple[str, str]] | None = None,
) -> tuple[Decimal, Decimal]:
    """Return aggregate actual W-4..W-1 and forecast W+1..W+4 totals."""

    lookup = _rows_by_key(rows)
    keys = set(lookup) if selected_keys is None else set(selected_keys)
    if not keys:
        raise GenerationError("Demand Trend scope contains no SKU-store rows")
    missing = keys - set(lookup)
    if missing:
        raise GenerationError(f"Demand Trend scope contains missing rows: {sorted(missing)[:3]}")
    actual_total = sum(
        (
            row.actual_for_week(week)
            for key in keys
            for row in [lookup[key]]
            for week in range(1, 5)
        ),
        Decimal("0"),
    )
    forecast_total = sum(
        (
            row.forecast_for_week(week)
            for key in keys
            for row in [lookup[key]]
            for week in range(1, 5)
        ),
        Decimal("0"),
    )
    return actual_total, forecast_total


def calculate_trend(
    rows: Sequence[DemandSkuStoreRow],
    selected_keys: Iterable[tuple[str, str]] | None = None,
) -> float:
    """Calculate Demand Trend after aggregate quantities, never row averages."""

    actual_total, forecast_total = trend_totals(rows, selected_keys)
    if actual_total <= 0:
        raise GenerationError("Demand Trend denominator is zero or negative")
    return float(forecast_total / actual_total - Decimal("1"))


def _calculate_row_trend(row: DemandSkuStoreRow) -> float:
    actual_total = sum(
        (row.actual_for_week(week) for week in range(1, 5)), Decimal("0")
    )
    forecast_total = sum(
        (row.forecast_for_week(week) for week in range(1, 5)), Decimal("0")
    )
    if actual_total <= 0:
        raise GenerationError(f"Demand Trend denominator is zero for {row.key}")
    return float(forecast_total / actual_total - Decimal("1"))


def _summary(values: Sequence[Decimal | float]) -> dict[str, float]:
    numbers = [float(value) for value in values]
    if not numbers:
        raise GenerationError("Cannot summarize an empty value sequence")
    return {
        "min": min(numbers),
        "median": statistics.median(numbers),
        "max": max(numbers),
    }


def _scope_keys(
    snapshot: SourceSnapshot,
    *,
    store_id: str | None = None,
    sku_id: str | None = None,
    cat: str | None = None,
    vertical_id: str | None = None,
) -> set[tuple[str, str]]:
    return {
        source.key
        for source in snapshot.source_rows
        if (store_id is None or source.store_id == store_id)
        and (sku_id is None or source.sku_id == sku_id)
        and (cat is None or source.cat == cat)
        and (vertical_id is None or source.vertical_id == vertical_id)
    }


def _pick_category(snapshot: SourceSnapshot) -> str:
    categories = sorted({source.cat for source in snapshot.source_rows})
    preferred = "GRC-C01"
    if preferred in categories:
        return preferred
    grc_categories = sorted(
        {source.cat for source in snapshot.source_rows if source.vertical_id == "GRC"}
    )
    return grc_categories[0] if grc_categories else categories[0]


def _pick_sku(snapshot: SourceSnapshot) -> str:
    sku_ids = sorted({source.sku_id for source in snapshot.source_rows})
    preferred = "GRC-001"
    if preferred in sku_ids:
        return preferred
    grc_skus = sorted(
        {source.sku_id for source in snapshot.source_rows if source.vertical_id == "GRC"}
    )
    return grc_skus[0] if grc_skus else sku_ids[0]


def filter_scope_validation(
    rows: Sequence[DemandSkuStoreRow],
    snapshot: SourceSnapshot,
) -> dict[str, Any]:
    """Calculate Trend and four-week totals for the required filter examples."""

    category = _pick_category(snapshot)
    sku = _pick_sku(snapshot)
    store = "S001" if "S001" in snapshot.store_ids else snapshot.store_ids[0]
    scopes: list[tuple[str, set[tuple[str, str]]]] = [
        ("All rows", _scope_keys(snapshot)),
        ("GRC", _scope_keys(snapshot, vertical_id="GRC")),
        (store, _scope_keys(snapshot, store_id=store)),
        (f"category {category}", _scope_keys(snapshot, cat=category)),
        (f"SKU {sku}", _scope_keys(snapshot, sku_id=sku)),
        (
            f"{store} + {category}",
            _scope_keys(snapshot, store_id=store, cat=category),
        ),
        (f"{store} + {sku}", _scope_keys(snapshot, store_id=store, sku_id=sku)),
    ]
    result: dict[str, Any] = {
        "category_example": category,
        "sku_example": sku,
        "store_example": store,
        "scopes": {},
    }
    for label, keys in scopes:
        if not keys:
            raise GenerationError(f"Required filter scope is empty: {label}")
        actual_total, forecast_total = trend_totals(rows, keys)
        result["scopes"][label] = {
            "row_count": len(keys),
            "actual_w4_w1_total": decimal_text(actual_total),
            "forecast_w1_w4_total": decimal_text(forecast_total),
            "demand_trend": calculate_trend(rows, keys),
        }
    return result


def _adjacent_changes(
    row: DemandSkuStoreRow,
) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    historical_labels = [f"actual_w{week}" for week in range(16, 0, -1)]
    historical_values = list(row.actuals)
    future_labels = [f"forecast_w{week}" for week in range(1, 17)]
    future_values = list(row.forecasts)
    for block, labels, values in (
        ("historical", historical_labels, historical_values),
        ("future", future_labels, future_values),
    ):
        for index in range(len(values) - 1):
            previous = values[index]
            following = values[index + 1]
            if previous <= 0:
                relative = 0.0 if following == 0 else float("inf")
            else:
                relative = float(following / previous - Decimal("1"))
            changes.append(
                {
                    "block": block,
                    "from": labels[index],
                    "to": labels[index + 1],
                    "relative_change": relative,
                    "from_value": float(previous),
                    "to_value": float(following),
                }
            )
    return changes


def plausibility_summary(
    rows: Sequence[DemandSkuStoreRow],
) -> dict[str, Any]:
    # Calculate row-level Trends directly.  Calling calculate_trend() for
    # every row would rebuild the full 16,000-row lookup 16,000 times.
    row_trends = {row.key: _calculate_row_trend(row) for row in rows}
    all_changes: list[dict[str, Any]] = []
    flat_series_count = 0
    extreme_series_count = 0
    for row in rows:
        historical_numbers = [float(value) for value in row.actuals]
        mean = statistics.mean(historical_numbers)
        stdev = statistics.pstdev(historical_numbers)
        if mean > 0 and stdev / mean < FLAT_HISTORY_CV_THRESHOLD:
            flat_series_count += 1
        changes = _adjacent_changes(row)
        all_changes.extend(
            [
                {
                    **change,
                    "sku_id": row.sku_id,
                    "store_id": row.store_id,
                }
                for change in changes
            ]
        )
        if (
            abs(row_trends[row.key]) > MAX_ROW_TREND_ABS
            or any(
                not math.isfinite(change["relative_change"])
                or abs(change["relative_change"]) > MAX_ADJACENT_CHANGE
                for change in changes
            )
        ):
            extreme_series_count += 1

    finite_changes = [
        change for change in all_changes if math.isfinite(change["relative_change"])
    ]
    if not finite_changes:
        raise GenerationError("No finite adjacent weekly changes to summarize")
    largest_increase = max(finite_changes, key=lambda change: change["relative_change"])
    largest_decrease = min(finite_changes, key=lambda change: change["relative_change"])
    return {
        "actual_w1": _summary([row.actual_for_week(1) for row in rows]),
        "forecast_w1": _summary([row.forecast_for_week(1) for row in rows]),
        "forecast_w16": _summary([row.forecast_for_week(16) for row in rows]),
        "row_trend": _summary(list(row_trends.values())),
        "largest_adjacent_weekly_increase": largest_increase,
        "largest_adjacent_weekly_decrease": largest_decrease,
        "suspiciously_flat_series_count": flat_series_count,
        "invalid_extreme_series_count": extreme_series_count,
        "thresholds": {
            "max_adjacent_change": MAX_ADJACENT_CHANGE,
            "max_row_trend_abs": MAX_ROW_TREND_ABS,
            "flat_history_cv_lt": FLAT_HISTORY_CV_THRESHOLD,
        },
        "row_trends": {
            f"{sku_id}|{store_id}": trend
            for (sku_id, store_id), trend in row_trends.items()
        },
    }


def _require_validation(condition: bool, message: str) -> None:
    if not condition:
        raise GenerationError(message)


def validate_dataset(
    rows: Sequence[DemandSkuStoreRow],
    snapshot: SourceSnapshot,
) -> dict[str, Any]:
    """Run all shape, source reconciliation, filter, and plausibility checks."""

    validate_source_snapshot(snapshot)
    source_by_key = {source.key: source for source in snapshot.source_rows}
    output_by_key = _rows_by_key(rows)
    _require_validation(
        len(rows) == EXPECTED_ROW_COUNT,
        f"Expected 16,000 output rows, found {len(rows)}",
    )
    _require_validation(
        set(output_by_key) == set(source_by_key),
        "Output SKU-store keys do not exactly match the source keys",
    )
    _require_validation(
        all(tuple(row.business_dict()) == BUSINESS_COLUMNS for row in rows),
        "Output row schema differs from the 35-column contract",
    )
    _require_validation(
        len({row.store_id for row in rows}) == EXPECTED_STORE_COUNT,
        "Expected 160 output stores",
    )
    _require_validation(
        len({row.sku_id for row in rows}) == EXPECTED_SKU_COUNT,
        "Expected 800 output SKUs",
    )
    store_counts = {
        store_id: sum(row.store_id == store_id for row in rows)
        for store_id in {row.store_id for row in rows}
    }
    _require_validation(
        set(store_counts.values()) == {EXPECTED_ROWS_PER_STORE},
        "Every output store must have exactly 100 rows",
    )
    _require_validation(
        len(output_by_key) == EXPECTED_ROW_COUNT,
        "Duplicate SKU-store output keys found",
    )

    for row in rows:
        source = source_by_key[row.key]
        _require_validation(row.cat == source.cat, f"Category mismatch for {row.key}")
        for value in (*row.actuals, *row.forecasts):
            _require_validation(
                value.is_finite() and value >= 0,
                f"Invalid negative/non-finite quantity for {row.key}",
            )
        _require_validation(
            row.forecast_for_week(1) == quantize_qty(source.forecast_7d),
            f"forecast_w1 does not preserve source Forecast 7d for {row.key}",
        )

    differences = {
        key: abs(output_by_key[key].forecast_for_week(1) - source.forecast_7d)
        for key, source in source_by_key.items()
    }
    passed = sum(value <= W1_RECONCILIATION_TOLERANCE for value in differences.values())
    _require_validation(
        passed == EXPECTED_ROW_COUNT,
        "Not all 16,000 forecast_w1 values reconcile to source Forecast 7d",
    )
    source_total = sum(
        (source.forecast_7d for source in snapshot.source_rows), Decimal("0")
    )
    generated_total = sum(
        (row.forecast_for_week(1) for row in rows), Decimal("0")
    )
    total_difference = abs(generated_total - source_total)
    _require_validation(
        total_difference <= W1_RECONCILIATION_TOLERANCE * EXPECTED_ROW_COUNT,
        f"Total forecast_w1 difference is too large: {total_difference}",
    )

    filter_scopes = filter_scope_validation(rows, snapshot)
    plausibility = plausibility_summary(rows)
    _require_validation(
        plausibility["invalid_extreme_series_count"] == 0,
        "One or more output series breached the plausibility thresholds",
    )

    sample_keys = [
        ("GRC-001", "S001"),
        ("GRC-001", "S002"),
        ("GRC-002", "S001"),
        ("GRC-002", "S002"),
        (rows[0].sku_id, rows[0].store_id),
    ]
    samples: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for key in sample_keys:
        if key in output_by_key and key not in seen:
            seen.add(key)
            samples.append(
                {
                    "sku_id": key[0],
                    "store_id": key[1],
                    "source_forecast_7d": decimal_text(source_by_key[key].forecast_7d),
                    "generated_forecast_w1": quantity_text(
                        output_by_key[key].forecast_for_week(1)
                    ),
                    "difference": decimal_text(differences[key]),
                }
            )

    return {
        "shape": {
            "row_count": len(rows),
            "column_count": len(BUSINESS_COLUMNS),
            "store_count": len({row.store_id for row in rows}),
            "sku_count": len({row.sku_id for row in rows}),
            "rows_per_store": store_counts,
            "unique_sku_store_pairs": len(output_by_key),
        },
        "value_counts": {
            "historical_synthetic": HISTORICAL_VALUE_COUNT,
            "source_w1": SOURCE_W1_VALUE_COUNT,
            "synthetic_future": SYNTHETIC_FUTURE_VALUE_COUNT,
            "total_period_values": TOTAL_PERIOD_VALUE_COUNT,
        },
        "category_mapping": {
            "non_null_rows": sum(bool(row.cat.strip()) for row in rows),
            "rows": len(rows),
        },
        "w1_reconciliation": {
            "source_row_count": EXPECTED_ROW_COUNT,
            "passed_count": passed,
            "failed_count": EXPECTED_ROW_COUNT - passed,
            "max_numeric_difference": decimal_text(max(differences.values())),
            "total_source_forecast_7d": decimal_text(source_total),
            "total_generated_forecast_w1": decimal_text(generated_total),
            "total_difference": decimal_text(total_difference),
            "tolerance": decimal_text(W1_RECONCILIATION_TOLERANCE),
            "samples": samples,
        },
        "filter_scope_validation": filter_scopes,
        "plausibility": plausibility,
        "business_columns": list(BUSINESS_COLUMNS),
    }


def _xlsx_value(column: str, value: str) -> Any:
    if column in {"sku_id", "store_id", "cat"}:
        return value
    return float(Decimal(value))


def write_csv(rows: Sequence[DemandSkuStoreRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=BUSINESS_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in sorted(rows, key=lambda item: item.key):
            writer.writerow(row.business_dict())


def write_xlsx(
    rows: Sequence[DemandSkuStoreRow],
    validation: Mapping[str, Any],
    path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Demand Store SKU 32W"
    sheet.append(list(BUSINESS_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sorted(rows, key=lambda item: item.key):
        values = row.business_dict()
        sheet.append([_xlsx_value(column, values[column]) for column in BUSINESS_COLUMNS])
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(BUSINESS_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(13, len(column) + 2)

    trend_sheet = workbook.create_sheet("Trend Summary")
    trend_sheet.append(
        ["Scope", "Rows", "Actual W-4...W-1 total", "Forecast W+1...W+4 total", "Demand Trend"]
    )
    for cell in trend_sheet[1]:
        cell.font = Font(bold=True)
    for label, item in validation["filter_scope_validation"]["scopes"].items():
        trend_sheet.append(
            [
                label,
                item["row_count"],
                float(Decimal(item["actual_w4_w1_total"])),
                float(Decimal(item["forecast_w1_w4_total"])),
                item["demand_trend"],
            ]
        )
    trend_sheet.freeze_panes = "A2"
    trend_sheet.auto_filter.ref = trend_sheet.dimensions

    validation_sheet = workbook.create_sheet("Validation Summary")
    validation_sheet.append(["Validation", "Result", "Detail"])
    for cell in validation_sheet[1]:
        cell.font = Font(bold=True)
    validation_rows = [
        ("shape", "PASS", f"{validation['shape']['row_count']} rows x {validation['shape']['column_count']} columns"),
        ("SKU-store uniqueness", "PASS", str(validation["shape"]["unique_sku_store_pairs"])),
        ("category mapping", "PASS", f"{validation['category_mapping']['non_null_rows']} non-null categories"),
        ("forecast_w1 reconciliation", "PASS", f"{validation['w1_reconciliation']['passed_count']}/{validation['w1_reconciliation']['source_row_count']} passed"),
        ("non-negative finite values", "PASS", str(validation["value_counts"]["total_period_values"])),
        ("plausibility", "PASS", f"{validation['plausibility']['invalid_extreme_series_count']} extreme series"),
    ]
    for row in validation_rows:
        validation_sheet.append(list(row))
    validation_sheet.freeze_panes = "A2"
    validation_sheet.auto_filter.ref = validation_sheet.dimensions
    for index, column in enumerate(("Validation", "Result", "Detail"), start=1):
        validation_sheet.column_dimensions[get_column_letter(index)].width = max(14, len(column) + 2)

    workbook.save(path)


def _normalise_export_value(column: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    if column in {"sku_id", "store_id", "cat"}:
        return str(value)
    return quantity_text(value)


def validate_export_parity(csv_path: Path, xlsx_path: Path) -> dict[str, Any]:
    """Verify that the XLSX canonical sheet matches the CSV business rows."""

    from openpyxl import load_workbook

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    _require_validation(
        len(csv_rows) == EXPECTED_ROW_COUNT,
        f"CSV export has {len(csv_rows)} rows instead of 16,000",
    )
    csv_canonical = [
        tuple(_normalise_export_value(column, row[column]) for column in BUSINESS_COLUMNS)
        for row in csv_rows
    ]

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        _require_validation(
            "Demand Store SKU 32W" in workbook.sheetnames,
            "XLSX is missing the canonical Demand Store SKU 32W sheet",
        )
        sheet = workbook["Demand Store SKU 32W"]
        values = sheet.iter_rows(values_only=True)
        header = tuple(next(values))
        _require_validation(header == BUSINESS_COLUMNS, "XLSX schema differs from CSV")
        xlsx_canonical = [
            tuple(
                _normalise_export_value(column, value)
                for column, value in zip(BUSINESS_COLUMNS, row)
            )
            for row in values
        ]
    finally:
        workbook.close()
    _require_validation(
        len(xlsx_canonical) == EXPECTED_ROW_COUNT,
        f"XLSX canonical sheet has {len(xlsx_canonical)} rows instead of 16,000",
    )
    _require_validation(csv_canonical == xlsx_canonical, "CSV and XLSX rows differ")
    return {
        "csv_rows": len(csv_canonical),
        "xlsx_rows": len(xlsx_canonical),
        "same_logical_rows": True,
    }


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return decimal_text(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    return value


def write_manifest(manifest: Mapping[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_json_safe(dict(manifest)), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _reproducibility_check(
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    rows: Sequence[DemandSkuStoreRow],
    fingerprint: str,
) -> dict[str, Any]:
    rerun = generate_dataset(snapshot, params)
    rerun_fingerprint = output_fingerprint(rerun)
    same_rows = [canonical_row_text(row) for row in rows] == [
        canonical_row_text(row) for row in rerun
    ]
    negative_params = replace(params, seed=params.seed + 1)
    negative_rows = generate_dataset(snapshot, negative_params)
    original_by_key = _rows_by_key(rows)
    negative_by_key = _rows_by_key(negative_rows)
    historical_changed = any(
        original_by_key[key].actuals != negative_by_key[key].actuals
        for key in original_by_key
    )
    future_changed = any(
        original_by_key[key].forecasts[1:] != negative_by_key[key].forecasts[1:]
        for key in original_by_key
    )
    w1_same = all(
        original_by_key[key].forecasts[0] == negative_by_key[key].forecasts[0]
        for key in original_by_key
    )
    identifiers_same = all(
        (
            original_by_key[key].sku_id,
            original_by_key[key].store_id,
            original_by_key[key].cat,
        )
        == (
            negative_by_key[key].sku_id,
            negative_by_key[key].store_id,
            negative_by_key[key].cat,
        )
        for key in original_by_key
    )
    negative_fingerprint = output_fingerprint(negative_rows)
    result = {
        "same_fingerprint": fingerprint == rerun_fingerprint,
        "same_rows": same_rows,
        "changed_seed_historical_values_differ": historical_changed,
        "changed_seed_future_w2_w16_differ": future_changed,
        "changed_seed_w1_same": w1_same,
        "changed_seed_identifiers_categories_same": identifiers_same,
        "rerun_output_fingerprint": rerun_fingerprint,
        "negative_control_seed": negative_params.seed,
        "negative_control_output_fingerprint": negative_fingerprint,
    }
    _require_validation(result["same_fingerprint"], "Same-seed fingerprint changed")
    _require_validation(result["same_rows"], "Same-seed rows changed")
    _require_validation(
        result["changed_seed_historical_values_differ"],
        "Changed seed did not change synthetic historical values",
    )
    _require_validation(
        result["changed_seed_future_w2_w16_differ"],
        "Changed seed did not change synthetic W+2...W+16 values",
    )
    _require_validation(result["changed_seed_w1_same"], "Changed seed changed source W+1")
    _require_validation(
        result["changed_seed_identifiers_categories_same"],
        "Changed seed changed identifiers or categories",
    )
    return result


def build_manifest(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    generated_at: str,
    reproducibility: Mapping[str, Any],
) -> dict[str, Any]:
    counts = validation["value_counts"]
    return {
        "generation_name": params.generation_name,
        "generator_version": params.generator_version,
        "seed": params.seed,
        "source_revision": snapshot.source_revision,
        "source_import_batch": snapshot.batch_id,
        "source_import_batch_id": snapshot.batch_id,
        "source_snapshot_date": snapshot.source_snapshot_date.isoformat(),
        "source_as_of_date": snapshot.source_snapshot_date.isoformat(),
        "source_row_count": snapshot.fact_rows,
        "output_row_count": validation["shape"]["row_count"],
        "sku_count": validation["shape"]["sku_count"],
        "store_count": validation["shape"]["store_count"],
        "historical_value_count": counts["historical_synthetic"],
        "source_w1_value_count": counts["source_w1"],
        "synthetic_future_value_count": counts["synthetic_future"],
        "total_period_value_count": counts["total_period_values"],
        "input_fingerprint": input_fp,
        "output_fingerprint": output_fp,
        "generated_at": generated_at,
        "business_timezone": BUSINESS_TIMEZONE,
        "quantity_precision": 6,
        "column_contract": {
            "columns": list(BUSINESS_COLUMNS),
            "identifier_columns": ["sku_id", "store_id", "cat"],
            "historical_columns": list(ACTUAL_COLUMNS),
            "forecast_columns": list(FORECAST_COLUMNS),
            "numeric_precision": 6,
            "canonical_sort": ["sku_id", "store_id"],
        },
        "provenance": {
            "actual_w16_to_actual_w1": "synthetic",
            "forecast_w1": "v8.5 source retail.fact_inventory_daily.forecast_7d for the exact SKU-store row, quantized to six decimals",
            "forecast_w2_to_forecast_w16": "synthetic",
        },
        "source_preflight": {
            "agent_name": snapshot.agent_name,
            "workbook_version": snapshot.workbook_version,
            "import_status": snapshot.import_status,
            "batch_total_rows": snapshot.batch_total_rows,
            "fact_rows": snapshot.fact_rows,
            "fact_distinct_stores": snapshot.fact_distinct_stores,
            "fact_distinct_skus": snapshot.fact_distinct_skus,
            "fact_dates": snapshot.fact_dates,
            "fact_min_date": snapshot.fact_min_date.isoformat(),
            "fact_max_date": snapshot.fact_max_date.isoformat(),
            "duplicate_source_keys": snapshot.duplicate_source_keys,
            "bad_ads": snapshot.bad_ads,
            "bad_forecast_7d": snapshot.bad_forecast_7d,
            "rows_not_batch": snapshot.rows_not_batch,
            "dim_store_rows": snapshot.dim_store_rows,
            "dim_store_distinct_ids": snapshot.dim_store_distinct_ids,
            "dim_item_rows": snapshot.dim_item_rows,
            "dim_item_distinct_ids": snapshot.dim_item_distinct_ids,
            "dim_item_duplicate_groups": snapshot.dim_item_duplicate_groups,
            "dim_item_null_categories": snapshot.dim_item_null_categories,
            "category_join_rows": snapshot.category_join_rows,
            "category_join_skus": snapshot.category_join_skus,
            "category_join_stores": snapshot.category_join_stores,
            "missing_category_items": snapshot.missing_category_items,
            "null_join_categories": snapshot.null_join_categories,
            "missing_join_stores": snapshot.missing_join_stores,
        },
        "w1_reconciliation": validation["w1_reconciliation"],
        "filter_scope_validation": validation["filter_scope_validation"],
        "plausibility": validation["plausibility"],
        "reproducibility": dict(reproducibility),
        "supersedes": {
            "generation_name": "demand_store_week_poc_v1",
            "status": "superseded for SQL loading; do not load the old 5,120-row store-week candidate",
        },
        "sql_changes_performed": False,
    }


def _fmt_pct(value: float) -> str:
    return f"{value:.2%}"


def _fmt_number(value: float) -> str:
    return f"{value:,.6f}"


def _path_for_report(path: Path) -> str:
    try:
        return str(path.relative_to(REPO_ROOT))
    except ValueError:
        return str(path)


def render_generation_report(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    csv_path: Path,
    xlsx_path: Path,
    manifest_path: Path,
    test_result: str,
    reproducibility: Mapping[str, Any],
) -> str:
    counts = validation["value_counts"]
    shape = validation["shape"]
    w1 = validation["w1_reconciliation"]
    scopes = validation["filter_scope_validation"]["scopes"]
    plausibility = validation["plausibility"]
    category = validation["filter_scope_validation"]["category_example"]
    sku = validation["filter_scope_validation"]["sku_example"]
    store = validation["filter_scope_validation"]["store_example"]
    old_path = "artifacts/demand_store_week_poc_v1.csv"
    lines = [
        "# Demand Store SKU 32W Generation Report",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Generation name | `{params.generation_name}` |",
        f"| Generator version | `{params.generator_version}` |",
        f"| Seed | `{params.seed}` |",
        f"| Source revision | `{snapshot.source_revision}` |",
        f"| Source batch | `{snapshot.batch_id}` |",
        f"| Rows | `{shape['row_count']}` |",
        f"| Columns | `{len(BUSINESS_COLUMNS)}` |",
        f"| Historical synthetic values | `{counts['historical_synthetic']}` |",
        f"| Source-derived W+1 values | `{counts['source_w1']}` |",
        f"| Future synthetic values | `{counts['synthetic_future']}` |",
        f"| Output fingerprint | `{output_fp}` |",
        "| Overall verdict | **READY FOR SQL LOAD WITH CAVEATS** |",
        "",
        "The canonical CSV is the only proposed SQL-load artifact. Azure SQL was queried read-only for source preflight and was not modified.",
        "",
        "## 1. Why the Previous Store-Level Candidate Was Superseded",
        "",
        f"The previous `{old_path}` candidate had 5,120 rows at Store × period grain. It aggregated the 16,000 SKU-store Forecast 7d values to one W+1 value per store and had no truthful SKU or category allocation. It is **superseded for SQL loading** and must not be loaded into Azure SQL. This generation keeps each source SKU-store pair and therefore supports category/SKU filter scopes independently.",
        "",
        "## 2. Source Preflight",
        "",
        "The generator used SELECT-only queries against the current v8.5 runtime family: `retail.fact_inventory_daily`, `retail.dim_store`, and `retail.dim_item`. It did not query or use `retail.StoreSkuSnapshot`, v8.2 `retail.Sku`, v8.2 workbook measures, or old workbook Trend constants.",
        "",
        f"- Batch `{snapshot.batch_id}`: `{snapshot.import_status}`, agent `{snapshot.agent_name}`, workbook version `{snapshot.workbook_version}`, source SHA `{snapshot.source_sha256}`.",
        f"- Snapshot date: `{snapshot.source_snapshot_date.isoformat()}`; fact rows: `{snapshot.fact_rows}`; stores: `{snapshot.fact_distinct_stores}`; SKUs: `{snapshot.fact_distinct_skus}`.",
        f"- Per-store coverage: `{EXPECTED_ROWS_PER_STORE}` unique SKU rows for every `{EXPECTED_STORE_COUNT}` store; duplicate source keys: `{snapshot.duplicate_source_keys}`.",
        f"- ADS quality failures: `{snapshot.bad_ads}`; Forecast 7d quality failures: `{snapshot.bad_forecast_7d}`; rows outside batch 23: `{snapshot.rows_not_batch}`.",
        f"- Category join: `{snapshot.category_join_rows}` rows, `{snapshot.category_join_skus}` SKUs, `{snapshot.category_join_stores}` stores, missing items `{snapshot.missing_category_items}`, null categories `{snapshot.null_join_categories}`, missing stores `{snapshot.missing_join_stores}`.",
        "",
        "All source preflight checks passed. If batch, revision, snapshot date, population, or key/category resolution drifts, this generator is designed to stop before writing artifacts.",
        "",
        "## 3. New SKU × Store Grain",
        "",
        f"The output has exactly one row per `sku_id + store_id`: `{shape['row_count']}` rows, `{shape['sku_count']}` SKUs, `{shape['store_count']}` stores, `{EXPECTED_ROWS_PER_STORE}` rows per store, and `{shape['unique_sku_store_pairs']}` unique pairs. Rows are sorted by `sku_id, store_id`.",
        "",
        "Identifier mapping:",
        "",
        "- `sku_id = retail.fact_inventory_daily.item_key`.",
        "- `store_id = retail.fact_inventory_daily.store_key`.",
        "- `cat = retail.dim_item.category_id` through `fact_inventory_daily.item_key = dim_item.item_id`.",
        "",
        "## 4. Simplified 35-Column Schema",
        "",
        "The canonical dataset sheet contains exactly these business columns and no row-level dates, provenance, seed, version, or timestamp fields:",
        "",
        "```text",
        " | ".join(BUSINESS_COLUMNS),
        "```",
        "",
        "The 16 `actual_*` columns are synthetic historical quantities, `forecast_w1` is the source-derived value, and `forecast_w2` through `forecast_w16` are synthetic future quantities. All numeric cells use deterministic six-decimal precision.",
        "",
        "## 5. Generator Method",
        "",
        "The model is deterministic and row-specific. For each SKU-store pair it uses the source ADS, exact Forecast 7d scale, category, store context, item seasonality/growth flags, and stable identifier hashes. It creates a bounded history level, a small bounded growth rate, two smooth low-amplitude seasonal waves, and smoothed deterministic noise. The model is deliberately transparent and does not target any workbook/reference Trend value.",
        "",
        "```text",
        "ads_weekly = ADS × 7.45",
        "baseline = average(ads_weekly, source Forecast 7d)",
        "actual_w16...actual_w1 = baseline × row-specific level × smooth growth/seasonality/noise",
        "forecast_w1 = exact source Forecast 7d",
        "forecast_w2...forecast_w16 = forecast_w1 × smooth row-specific growth/seasonality/noise",
        "```",
        "",
        "No SQL-side randomness, workbook Demand Trend, Forecast Accuracy, old Trend constants, or target KPI back-solving is used.",
        "",
        "## 6. W+1 Source Preservation",
        "",
        "For every one of the 16,000 source rows, `forecast_w1` is the exact corresponding `retail.fact_inventory_daily.forecast_7d` value after the agreed six-decimal output quantization. It is not aggregated to store level, smoothed, scaled, trended, noised, or reconciled to a target KPI.",
        "",
        f"- Reconciliation: `{w1['passed_count']}` passed, `{w1['failed_count']}` failed out of `{w1['source_row_count']}`.",
        f"- Maximum raw source-to-output difference: `{w1['max_numeric_difference']}`.",
        f"- Total source Forecast 7d: `{w1['total_source_forecast_7d']}`; total generated forecast_w1: `{w1['total_generated_forecast_w1']}`; difference: `{w1['total_difference']}`.",
        "",
        "Sample source/output W+1 rows:",
        "",
        "| SKU | Store | Source Forecast 7d | Generated forecast_w1 | Difference |",
        "|---|---|---:|---:|---:|",
        *[
            f"| {sample['sku_id']} | {sample['store_id']} | {sample['source_forecast_7d']} | {sample['generated_forecast_w1']} | {sample['difference']} |"
            for sample in w1["samples"]
        ],
        "",
        "## 7. Historical Generation",
        "",
        f"The generator creates `{counts['historical_synthetic']}` synthetic historical values (`{shape['row_count']} × 16`) in `actual_w16` through `actual_w1`. They are non-negative, finite, row-specific quantities anchored to each SKU-store source scale. They are not genuine sales or POS history.",
        "",
        "The stable hash includes the global seed, SKU ID, Store ID, category, and model block. This preserves independent, reproducible differences between stores for one SKU and between SKUs in one store, while the smooth recurrence avoids flat or erratic series.",
        "",
        "## 8. Future Generation",
        "",
        f"The generator creates `{counts['synthetic_future']}` synthetic future values (`{shape['row_count']} × 15`) in `forecast_w2` through `forecast_w16`. Each trajectory starts from the exact source-derived `forecast_w1`, preserves relative SKU-store scale, and uses bounded smooth variation. No workbook Demand Trend or Forecast Accuracy constant is used.",
        "",
        "## 9. Output Artifacts",
        "",
        f"- CSV: `{_path_for_report(csv_path)}` (canonical future SQL-load artifact).",
        f"- XLSX: `{_path_for_report(xlsx_path)}`; main sheet `Demand Store SKU 32W` contains exactly the 16,000 × 35 canonical dataset. `Trend Summary` and `Validation Summary` are review-only.",
        f"- Manifest: `{_path_for_report(manifest_path)}`; it carries provenance, source preflight, counts, fingerprints, validation, and reproducibility metadata.",
        "",
        "## 10. Row / Value Counts",
        "",
        "| Block | Database rows | Period values | Provenance |",
        "|---|---:|---:|---|",
        f"| SKU × Store output rows | `{shape['row_count']}` | — | One row per SKU × Store |",
        f"| Historical actual columns | `{shape['row_count']}` | `{counts['historical_synthetic']}` | Synthetic |",
        f"| Forecast W+1 | `{shape['row_count']}` | `{counts['source_w1']}` | Exact v8.5 source Forecast 7d per SKU × Store |",
        f"| Forecast W+2…W+16 | `{shape['row_count']}` | `{counts['synthetic_future']}` | Synthetic |",
        f"| **Total period values** | **`{shape['row_count']}`** | **`{counts['total_period_values']}`** | **256,000 synthetic history + 16,000 source W+1 + 240,000 synthetic future** |",
        "",
        "This is not a 512,000-row table: it is a 16,000-row table containing 512,000 weekly values across its wide columns.",
        "",
        "## 11. Filter-Scope Validation",
        "",
        "The new grain independently calculates four-week totals and Demand Trend for all required scopes:",
        "",
        "| Scope | Rows | Actual W-4...W-1 total | Forecast W+1...W+4 total | Demand Trend |",
        "|---|---:|---:|---:|---:|",
        *[
            f"| {label} | {item['row_count']} | {item['actual_w4_w1_total']} | {item['forecast_w1_w4_total']} | {_fmt_pct(item['demand_trend'])} |"
            for label, item in scopes.items()
        ],
        "",
        f"Examples use available IDs: category `{category}`, SKU `{sku}`, and Store `{store}`. The GRC scope is the source store vertical `GRC`, not a workbook target.",
        "",
        "## 12. Demand Trend Results",
        "",
        "For every selected set of SKU-store rows, the formula is applied after summing quantities:",
        "",
        "```text",
        "Demand Trend = SUM(forecast_w1 + forecast_w2 + forecast_w3 + forecast_w4)",
        "                / SUM(actual_w4 + actual_w3 + actual_w2 + actual_w1) - 1",
        "```",
        "",
        "Individual row percentages are never averaged. The formula remains fixed when the forecast horizon changes.",
        "",
        "Horizon contract:",
        "",
        "- 4w: `forecast_w1` through `forecast_w4`.",
        "- 8w: `forecast_w1` through `forecast_w8`.",
        "- 12w: `forecast_w1` through `forecast_w12`.",
        "- 16w: `forecast_w1` through `forecast_w16`.",
        "- Demand Trend always remains the W+1…W+4 versus W-4…W-1 formula above.",
        "",
        "## 13. Plausibility Results",
        "",
        f"- `actual_w1` min/median/max: `{_fmt_number(plausibility['actual_w1']['min'])}` / `{_fmt_number(plausibility['actual_w1']['median'])}` / `{_fmt_number(plausibility['actual_w1']['max'])}`.",
        f"- `forecast_w1` min/median/max: `{_fmt_number(plausibility['forecast_w1']['min'])}` / `{_fmt_number(plausibility['forecast_w1']['median'])}` / `{_fmt_number(plausibility['forecast_w1']['max'])}`.",
        f"- `forecast_w16` min/median/max: `{_fmt_number(plausibility['forecast_w16']['min'])}` / `{_fmt_number(plausibility['forecast_w16']['median'])}` / `{_fmt_number(plausibility['forecast_w16']['max'])}`.",
        f"- Row-level Trend min/median/max: `{_fmt_pct(plausibility['row_trend']['min'])}` / `{_fmt_pct(plausibility['row_trend']['median'])}` / `{_fmt_pct(plausibility['row_trend']['max'])}`.",
        f"- GRC aggregate Trend: `{_fmt_pct(scopes['GRC']['demand_trend'])}`; all-row aggregate Trend: `{_fmt_pct(scopes['All rows']['demand_trend'])}`.",
        f"- Largest adjacent weekly increase: `{plausibility['largest_adjacent_weekly_increase']['sku_id']} + {plausibility['largest_adjacent_weekly_increase']['store_id']}` `{plausibility['largest_adjacent_weekly_increase']['from']}→{plausibility['largest_adjacent_weekly_increase']['to']}` `{_fmt_pct(plausibility['largest_adjacent_weekly_increase']['relative_change'])}`.",
        f"- Largest adjacent weekly decrease: `{plausibility['largest_adjacent_weekly_decrease']['sku_id']} + {plausibility['largest_adjacent_weekly_decrease']['store_id']}` `{plausibility['largest_adjacent_weekly_decrease']['from']}→{plausibility['largest_adjacent_weekly_decrease']['to']}` `{_fmt_pct(plausibility['largest_adjacent_weekly_decrease']['relative_change'])}`.",
        f"- Suspiciously flat historical series count (CV < {FLAT_HISTORY_CV_THRESHOLD:.3f}): `{plausibility['suspiciously_flat_series_count']}`.",
        f"- Invalid/extreme series count: `{plausibility['invalid_extreme_series_count']}`; threshold was adjacent movement > {MAX_ADJACENT_CHANGE:.0%} or row Trend beyond ±{MAX_ROW_TREND_ABS:.0%}.",
        "",
        "Adjacent changes are measured within the contiguous historical block and within the contiguous future block; the intentional W0 gap is not treated as a weekly period.",
        "",
        "## 14. Reproducibility",
        "",
        f"- Same seed rerun fingerprint identical: `{reproducibility['same_fingerprint']}`; identical canonical rows: `{reproducibility['same_rows']}`.",
        f"- Changed-seed historical values changed: `{reproducibility['changed_seed_historical_values_differ']}`; changed-seed W+2…W+16 values changed: `{reproducibility['changed_seed_future_w2_w16_differ']}`.",
        f"- Changed-seed W+1 values remained identical: `{reproducibility['changed_seed_w1_same']}`; identifiers/categories remained identical: `{reproducibility['changed_seed_identifiers_categories_same']}`.",
        f"- Input fingerprint: `{input_fp}`; output fingerprint: `{output_fp}`.",
        "",
        "The output fingerprint hashes canonical rows sorted by `sku_id, store_id`, with all 35 business columns and six-decimal quantity formatting. It excludes XLSX formatting, row order outside the canonical sort, and volatile timestamps.",
        "",
        "## 15. Tests",
        "",
        test_result,
        "",
        "The generator also runs same-seed and changed-seed negative-control validation during artifact generation.",
        "",
        "## 16. Known Limitations",
        "",
        "- `actual_w16` through `actual_w1` are synthetic and are not genuine transaction/POS history; ADS is only a modelling baseline.",
        "- `forecast_w2` through `forecast_w16` are synthetic future forecasts.",
        "- `forecast_w1` comes directly from the existing v8.5 Forecast 7d source row and is quantized to the output precision; it has no persisted target-week date.",
        "- Fixed W-16...W+16 labels intentionally carry no explicit dates in this simplified POC table.",
        "- The table is a temporary synthetic POC intended to be removed or replaced when governed historical demand and forecast runs exist.",
        "- Values follow current sales-unit SUM semantics; no pack-factor, returns, cancellations, or stockout-censored demand policy is claimed.",
        "- This task did not wire backend/frontend runtime behavior and did not modify the Demand Forecasting application.",
        "",
        "## 17. SQL Load Readiness",
        "",
        "**READY FOR SQL LOAD WITH CAVEATS.** All required offline shape, source mapping, W+1 reconciliation, non-negative value, filter-scope, plausibility, fingerprint, and reproducibility checks passed. The CSV is ready for a later explicitly authorized SQL-load task after human review. The previous store-week candidate remains superseded and must not be loaded. Azure SQL was not modified in this task.",
        "",
    ]
    return "\n".join(lines)


def write_generation_report(content: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def run_generation(
    *,
    output_dir: Path,
    params: GeneratorParameters,
    test_result: str,
) -> dict[str, Any]:
    snapshot = load_source_snapshot()
    rows = generate_dataset(snapshot, params)
    validation = validate_dataset(rows, snapshot)
    output_fp = output_fingerprint(rows)
    input_fp = input_fingerprint(snapshot, params)
    reproducibility = _reproducibility_check(snapshot, params, rows, output_fp)

    csv_path = output_dir / "demand_store_sku_32w_poc_v1.csv"
    xlsx_path = output_dir / "demand_store_sku_32w_poc_v1.xlsx"
    manifest_path = output_dir / "demand_store_sku_32w_poc_v1_manifest.json"
    report_path = REPO_ROOT / "plans" / "demand-store-sku-32w-generation-report.md"
    write_csv(rows, csv_path)
    write_xlsx(rows, validation, xlsx_path)
    validation["export_parity"] = validate_export_parity(csv_path, xlsx_path)
    generated_at = datetime.now(timezone.utc).isoformat()
    manifest = build_manifest(
        snapshot=snapshot,
        params=params,
        validation=validation,
        input_fp=input_fp,
        output_fp=output_fp,
        generated_at=generated_at,
        reproducibility=reproducibility,
    )
    write_manifest(manifest, manifest_path)
    write_generation_report(
        render_generation_report(
            snapshot=snapshot,
            params=params,
            validation=validation,
            input_fp=input_fp,
            output_fp=output_fp,
            csv_path=csv_path,
            xlsx_path=xlsx_path,
            manifest_path=manifest_path,
            test_result=test_result,
            reproducibility=reproducibility,
        ),
        report_path,
    )
    return {
        "snapshot": snapshot,
        "rows": rows,
        "validation": validation,
        "input_fingerprint": input_fp,
        "output_fingerprint": output_fp,
        "reproducibility": reproducibility,
        "csv_path": csv_path,
        "xlsx_path": xlsx_path,
        "manifest_path": manifest_path,
        "report_path": report_path,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "artifacts",
        help="Local output directory (default: artifacts)",
    )
    parser.add_argument("--seed", type=int, default=FIXED_SEED)
    parser.add_argument("--generation-name", default=GENERATION_NAME)
    parser.add_argument("--generator-version", default=GENERATOR_VERSION)
    parser.add_argument(
        "--test-result",
        default="Focused generator tests: run separately; see final handoff.",
        help="Text recorded in the generation report Test section",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    params = GeneratorParameters(
        generation_name=args.generation_name,
        generator_version=args.generator_version,
        seed=args.seed,
    )
    try:
        result = run_generation(
            output_dir=args.output_dir,
            params=params,
            test_result=args.test_result,
        )
    except SourcePreflightError as exc:
        print(f"SOURCE PREFLIGHT FAILED: {exc}", file=sys.stderr)
        return 2
    except GenerationError as exc:
        print(f"GENERATION/VALIDATION FAILED: {exc}", file=sys.stderr)
        return 3
    print(
        json.dumps(
            {
                "report_path": str(result["report_path"]),
                "generator_path": str(Path(__file__).resolve()),
                "csv_path": str(result["csv_path"]),
                "xlsx_path": str(result["xlsx_path"]),
                "manifest_path": str(result["manifest_path"]),
                "output_fingerprint": result["output_fingerprint"],
                "shape": result["validation"]["shape"],
                "value_counts": result["validation"]["value_counts"],
                "w1_reconciliation": result["validation"]["w1_reconciliation"],
                "trend": {
                    label: item["demand_trend"]
                    for label, item in result["validation"]["filter_scope_validation"]["scopes"].items()
                },
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
