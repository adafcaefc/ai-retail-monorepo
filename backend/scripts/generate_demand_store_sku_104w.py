"""Generate the local 104-week SKU x Store demand extension.

This generator is deliberately an extension of the approved 32W artifact.  It
reads the existing CSV as the source of truth for the first 32 periods, copies
that block without regenerating it, and creates only W-52..W-17 and
W+17..W+52.  The live v8.5 source is queried read-only for preflight and W+1
reconciliation.  No Azure SQL object or row is created or changed.

Run from the repository root, for example::

    .venv/bin/python backend/scripts/generate_demand_store_sku_104w.py \
        --test-result 'Focused 104W generator tests: PASS'
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import statistics
import sys
from dataclasses import dataclass, replace
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence


REPO_ROOT = Path(__file__).resolve().parents[2]
SCRIPT_DIR = Path(__file__).resolve().parent
if str(REPO_ROOT / "backend") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "backend"))
if str(SCRIPT_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPT_DIR))

import generate_demand_store_sku_32w as previous_generator


GENERATION_NAME = "demand_store_sku_104w_poc_v1"
GENERATOR_VERSION = "demand-store-sku-104w-generator-v1.0.0"
FIXED_SEED = previous_generator.FIXED_SEED
SOURCE_IMPORT_BATCH_ID = previous_generator.SOURCE_IMPORT_BATCH_ID
SOURCE_SNAPSHOT_DATE = previous_generator.SOURCE_SNAPSHOT_DATE
SOURCE_REVISION_SHA = previous_generator.SOURCE_REVISION_SHA
SOURCE_REVISION = previous_generator.SOURCE_REVISION
BUSINESS_TIMEZONE = previous_generator.BUSINESS_TIMEZONE
QUANTITY_QUANTUM = previous_generator.QUANTITY_QUANTUM
EXPECTED_ROW_COUNT = previous_generator.EXPECTED_ROW_COUNT
EXPECTED_STORE_COUNT = previous_generator.EXPECTED_STORE_COUNT
EXPECTED_SKU_COUNT = previous_generator.EXPECTED_SKU_COUNT
EXPECTED_ROWS_PER_STORE = previous_generator.EXPECTED_ROWS_PER_STORE
W1_RECONCILIATION_TOLERANCE = previous_generator.W1_RECONCILIATION_TOLERANCE
MAX_ADJACENT_CHANGE = previous_generator.MAX_ADJACENT_CHANGE
MAX_ROW_TREND_ABS = previous_generator.MAX_ROW_TREND_ABS
FLAT_HISTORY_CV_THRESHOLD = previous_generator.FLAT_HISTORY_CV_THRESHOLD

PREVIOUS_CSV_PATH = REPO_ROOT / "artifacts" / "demand_store_sku_32w_poc_v1.csv"
PREVIOUS_MANIFEST_PATH = (
    REPO_ROOT / "artifacts" / "demand_store_sku_32w_poc_v1_manifest.json"
)
REPORT_PATH = REPO_ROOT / "plans" / "demand-store-sku-104w-generation-report.md"
APPROVED_PREVIOUS_FINGERPRINT = (
    "0e3df661a941440d0e43fa93e62fe166d69c5d12caa1b6ed65333729c78f550d"
)

ACTUAL_COLUMNS = tuple(f"actual_w{week}" for week in range(52, 0, -1))
FORECAST_COLUMNS = tuple(f"forecast_w{week}" for week in range(1, 53))
PRESERVED_ACTUAL_COLUMNS = tuple(
    f"actual_w{week}" for week in range(16, 0, -1)
)
PRESERVED_FORECAST_COLUMNS = tuple(f"forecast_w{week}" for week in range(1, 17))
NEW_HISTORICAL_COLUMNS = tuple(f"actual_w{week}" for week in range(52, 16, -1))
NEW_FORECAST_COLUMNS = tuple(f"forecast_w{week}" for week in range(17, 53))
PRESERVED_COLUMNS = (
    "sku_id",
    "store_id",
    "cat",
    *PRESERVED_ACTUAL_COLUMNS,
    *PRESERVED_FORECAST_COLUMNS,
)
BUSINESS_COLUMNS = ("sku_id", "store_id", "cat", *ACTUAL_COLUMNS, *FORECAST_COLUMNS)

EXISTING_PRESERVED_VALUE_COUNT = EXPECTED_ROW_COUNT * 32
NEW_HISTORICAL_VALUE_COUNT = EXPECTED_ROW_COUNT * 36
NEW_FUTURE_VALUE_COUNT = EXPECTED_ROW_COUNT * 36
HISTORICAL_VALUE_COUNT = EXPECTED_ROW_COUNT * 52
SOURCE_W1_VALUE_COUNT = EXPECTED_ROW_COUNT
SYNTHETIC_FUTURE_VALUE_COUNT = EXPECTED_ROW_COUNT * 51
TOTAL_PERIOD_VALUE_COUNT = EXPECTED_ROW_COUNT * 104


GenerationError = previous_generator.GenerationError
SourcePreflightError = previous_generator.SourcePreflightError
SourceSkuStore = previous_generator.SourceSkuStore
SourceSnapshot = previous_generator.SourceSnapshot


@dataclass(frozen=True)
class GeneratorParameters:
    generation_name: str = GENERATION_NAME
    generator_version: str = GENERATOR_VERSION
    seed: int = FIXED_SEED
    source_snapshot_date: date = SOURCE_SNAPSHOT_DATE

    def manifest_dict(self) -> dict[str, Any]:
        return {
            "generation_name": self.generation_name,
            "generator_version": self.generator_version,
            "seed": self.seed,
            "source_snapshot_date": self.source_snapshot_date.isoformat(),
            "business_timezone": BUSINESS_TIMEZONE,
            "quantity_precision": 6,
            "extension_method": {
                "historical": "recursive deterministic continuation backward from fixed actual_w16",
                "future": "recursive deterministic continuation forward from fixed forecast_w16",
                "growth": "existing bounded row-specific model growth rate",
                "seasonality": "existing two smooth bounded sinusoidal components",
                "noise": "stable identifier-hashed smooth recurrence, 0.72 previous / 0.28 new",
                "preserved_block": "approved 32W CSV copied period-for-period",
            },
        }


@dataclass(frozen=True)
class DemandSkuStore104WRow:
    """One canonical 107-column row at SKU x Store grain."""

    sku_id: str
    store_id: str
    cat: str
    actuals: tuple[Decimal, ...]  # actual_w52 ... actual_w1
    forecasts: tuple[Decimal, ...]  # forecast_w1 ... forecast_w52

    @property
    def key(self) -> tuple[str, str]:
        return self.sku_id, self.store_id

    def actual_for_week(self, week: int) -> Decimal:
        if week < 1 or week > 52:
            raise ValueError(f"Actual week must be 1..52, got {week}")
        return self.actuals[52 - week]

    def forecast_for_week(self, week: int) -> Decimal:
        if week < 1 or week > 52:
            raise ValueError(f"Forecast week must be 1..52, got {week}")
        return self.forecasts[week - 1]

    def business_dict(self) -> dict[str, str]:
        values: dict[str, str] = {
            "sku_id": self.sku_id,
            "store_id": self.store_id,
            "cat": self.cat,
        }
        values.update(
            {
                column: quantity_text(value)
                for column, value in zip(ACTUAL_COLUMNS, self.actuals)
            }
        )
        values.update(
            {
                column: quantity_text(value)
                for column, value in zip(FORECAST_COLUMNS, self.forecasts)
            }
        )
        return values


def decimal_value(value: Any) -> Decimal:
    if isinstance(value, Decimal):
        result = value
    else:
        result = Decimal(str(value))
    if not result.is_finite():
        raise GenerationError(f"Non-finite numeric value: {value!r}")
    return result


def quantize_qty(value: Decimal | float | int) -> Decimal:
    try:
        result = decimal_value(value).quantize(
            QUANTITY_QUANTUM, rounding=ROUND_HALF_UP
        )
    except (InvalidOperation, ValueError) as exc:
        raise GenerationError(f"Unable to quantize quantity {value!r}") from exc
    if result < 0:
        raise GenerationError(f"Negative generated quantity: {result}")
    return result


def decimal_text(value: Decimal | float | int) -> str:
    return format(decimal_value(value), "f")


def quantity_text(value: Decimal | float | int) -> str:
    return format(quantize_qty(value), "f")


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise GenerationError(message)


def _read_previous_manifest(path: Path) -> dict[str, Any]:
    try:
        manifest = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise GenerationError(f"Unable to read approved 32W manifest: {path}") from exc
    _require(
        manifest.get("generation_name") == "demand_store_sku_32w_poc_v1",
        "The previous manifest is not demand_store_sku_32w_poc_v1",
    )
    _require(
        manifest.get("output_fingerprint") == APPROVED_PREVIOUS_FINGERPRINT,
        "The previous manifest fingerprint is not the approved 32W fingerprint",
    )
    _require(
        manifest.get("output_row_count") == EXPECTED_ROW_COUNT
        and len(manifest.get("column_contract", {}).get("columns", [])) == 35,
        "The previous manifest does not describe the approved 16,000 x 35 artifact",
    )
    return manifest


def load_previous_rows(
    csv_path: Path = PREVIOUS_CSV_PATH,
    manifest_path: Path = PREVIOUS_MANIFEST_PATH,
) -> tuple[list[previous_generator.DemandSkuStoreRow], dict[str, Any], str]:
    """Load and fingerprint the approved 32W CSV without regenerating it."""

    manifest = _read_previous_manifest(manifest_path)
    try:
        with csv_path.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle)
            _require(
                tuple(reader.fieldnames or ()) == previous_generator.BUSINESS_COLUMNS,
                "Approved 32W CSV header differs from its 35-column contract",
            )
            rows: list[previous_generator.DemandSkuStoreRow] = []
            for raw in reader:
                _require(
                    all(raw.get(column) not in (None, "") for column in previous_generator.BUSINESS_COLUMNS),
                    "Approved 32W CSV contains a blank business value",
                )
                rows.append(
                    previous_generator.DemandSkuStoreRow(
                        sku_id=str(raw["sku_id"]),
                        store_id=str(raw["store_id"]),
                        cat=str(raw["cat"]),
                        actuals=tuple(
                            decimal_value(raw[column])
                            for column in previous_generator.ACTUAL_COLUMNS
                        ),
                        forecasts=tuple(
                            decimal_value(raw[column])
                            for column in previous_generator.FORECAST_COLUMNS
                        ),
                    )
                )
    except OSError as exc:
        raise GenerationError(f"Unable to read approved 32W CSV: {csv_path}") from exc

    rows.sort(key=lambda row: row.key)
    _require(len(rows) == EXPECTED_ROW_COUNT, "Approved 32W CSV must contain 16,000 rows")
    _require(
        len({row.key for row in rows}) == EXPECTED_ROW_COUNT,
        "Approved 32W CSV contains duplicate SKU-store keys",
    )
    fingerprint = previous_generator.output_fingerprint(rows)
    _require(
        fingerprint == APPROVED_PREVIOUS_FINGERPRINT,
        f"Approved 32W CSV fingerprint differs: {fingerprint}",
    )
    _require(
        manifest.get("output_fingerprint") == fingerprint,
        "Approved 32W CSV and manifest fingerprints differ",
    )
    return rows, manifest, fingerprint


def validate_previous_against_source(
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
    snapshot: SourceSnapshot,
) -> None:
    """Ensure the approved block still maps to the current v8.5 source."""

    source_by_key = {source.key: source for source in snapshot.source_rows}
    previous_by_key = {row.key: row for row in previous_rows}
    _require(set(previous_by_key) == set(source_by_key), "32W keys drifted from v8.5 source")
    for key, previous in previous_by_key.items():
        source = source_by_key[key]
        _require(previous.cat == source.cat, f"32W category drift for {key}")
        _require(
            previous.forecast_for_week(1) == quantize_qty(source.forecast_7d),
            f"32W forecast_w1 no longer matches source Forecast 7d for {key}",
        )


def _model_parameters(
    source: SourceSkuStore,
    params: GeneratorParameters,
) -> dict[str, float | Decimal]:
    # The approved 32W model is reused verbatim for the row-specific inputs.
    return previous_generator._model_parameters(source, params)


def _extension_noise(
    source: SourceSkuStore,
    params: GeneratorParameters,
    block: str,
    periods: Iterable[int],
    boundary_period: int,
    direction: str,
) -> dict[int, float]:
    """Create a smooth deterministic noise path anchored at the fixed boundary."""

    result: dict[int, float] = {boundary_period: 0.0}
    previous = 0.0
    for period in periods:
        raw = previous_generator._stable_signed(
            params.seed,
            "104w-extension-noise",
            params.generator_version,
            block,
            direction,
            source.sku_id,
            source.store_id,
            period,
        )
        previous = 0.72 * previous + 0.28 * raw
        result[period] = previous
    return result


def _extension_factor(
    period: int,
    previous_period: int,
    model: Mapping[str, float | Decimal],
    noise: Mapping[int, float],
    direction: str,
) -> float:
    growth_rate = float(model["growth_rate"])
    growth_factor = 1.0 + growth_rate if direction == "forward" else 1.0 / (1.0 + growth_rate)
    season_ratio = (1.0 + previous_generator._seasonal_signal(period, model)) / (
        1.0 + previous_generator._seasonal_signal(previous_period, model)
    )
    noise_ratio = 1.0 + float(model["noise_amplitude"]) * (
        noise[period] - noise[previous_period]
    )
    factor = growth_factor * season_ratio * noise_ratio
    if not math.isfinite(factor) or factor < 0:
        raise GenerationError(f"Invalid extension factor at period {period}: {factor}")
    return factor


def _extend_history(
    previous: previous_generator.DemandSkuStoreRow,
    source: SourceSkuStore,
    params: GeneratorParameters,
) -> tuple[Decimal, ...]:
    model = _model_parameters(source, params)
    noise = _extension_noise(
        source,
        params,
        "historical",
        range(-17, -53, -1),
        -16,
        "backward",
    )
    values_by_week: dict[int, Decimal] = {}
    previous_value = previous.forecasts[0] * Decimal("0") + previous.actuals[0]
    previous_period = -16
    for period in range(-17, -53, -1):
        factor = _extension_factor(
            period, previous_period, model, noise, "backward"
        )
        previous_value = quantize_qty(previous_value * Decimal(str(factor)))
        values_by_week[-period] = previous_value
        previous_period = period
    return tuple(values_by_week[week] for week in range(52, 16, -1)) + previous.actuals


def _extend_future(
    previous: previous_generator.DemandSkuStoreRow,
    source: SourceSkuStore,
    params: GeneratorParameters,
) -> tuple[Decimal, ...]:
    model = _model_parameters(source, params)
    noise = _extension_noise(
        source,
        params,
        "future",
        range(17, 53),
        16,
        "forward",
    )
    values = list(previous.forecasts)
    previous_value = values[-1]
    previous_period = 16
    for period in range(17, 53):
        factor = _extension_factor(
            period, previous_period, model, noise, "forward"
        )
        previous_value = quantize_qty(previous_value * Decimal(str(factor)))
        values.append(previous_value)
        previous_period = period
    return tuple(values)


def generate_dataset(
    snapshot: SourceSnapshot,
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
    params: GeneratorParameters | None = None,
) -> list[DemandSkuStore104WRow]:
    """Copy the approved 32W block and generate only the 72 extension periods."""

    params = params or GeneratorParameters()
    previous_generator.validate_source_snapshot(snapshot)
    validate_previous_against_source(previous_rows, snapshot)
    _require(
        params.source_snapshot_date == snapshot.source_snapshot_date,
        "Generator/source snapshot dates do not match",
    )
    _require(bool(params.generation_name.strip()), "Generation name must not be blank")

    source_by_key = {source.key: source for source in snapshot.source_rows}
    previous_by_key = {row.key: row for row in previous_rows}
    rows: list[DemandSkuStore104WRow] = []
    for key in sorted(previous_by_key):
        previous = previous_by_key[key]
        source = source_by_key[key]
        rows.append(
            DemandSkuStore104WRow(
                sku_id=previous.sku_id,
                store_id=previous.store_id,
                cat=previous.cat,
                actuals=_extend_history(previous, source, params),
                forecasts=_extend_future(previous, source, params),
            )
        )
    return rows


def canonical_row_text(row: DemandSkuStore104WRow) -> str:
    values = row.business_dict()
    return "|".join(values[column] for column in BUSINESS_COLUMNS)


def output_fingerprint(rows: Iterable[DemandSkuStore104WRow]) -> str:
    digest = hashlib.sha256()
    for row in sorted(rows, key=lambda item: item.key):
        digest.update(canonical_row_text(row).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def input_fingerprint(
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    previous_fingerprint: str,
) -> str:
    payload = {
        "source": snapshot.fingerprint_dict(),
        "parameters": params.manifest_dict(),
        "previous_32w_fingerprint": previous_fingerprint,
        "output_columns": list(BUSINESS_COLUMNS),
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _rows_by_key(rows: Sequence[DemandSkuStore104WRow]) -> dict[tuple[str, str], DemandSkuStore104WRow]:
    result: dict[tuple[str, str], DemandSkuStore104WRow] = {}
    for row in rows:
        if row.key in result:
            raise GenerationError(f"Duplicate output SKU-store key {row.key}")
        result[row.key] = row
    return result


def trend_totals(
    rows: Sequence[DemandSkuStore104WRow],
    selected_keys: Iterable[tuple[str, str]] | None = None,
) -> tuple[Decimal, Decimal]:
    lookup = _rows_by_key(rows)
    keys = set(lookup) if selected_keys is None else set(selected_keys)
    _require(bool(keys), "Demand Trend scope contains no SKU-store rows")
    _require(keys <= set(lookup), "Demand Trend scope contains missing rows")
    actual_total = sum(
        (lookup[key].actual_for_week(week) for key in keys for week in range(1, 5)),
        Decimal("0"),
    )
    forecast_total = sum(
        (lookup[key].forecast_for_week(week) for key in keys for week in range(1, 5)),
        Decimal("0"),
    )
    return actual_total, forecast_total


def calculate_trend(
    rows: Sequence[DemandSkuStore104WRow],
    selected_keys: Iterable[tuple[str, str]] | None = None,
) -> float:
    actual_total, forecast_total = trend_totals(rows, selected_keys)
    _require(actual_total > 0, "Demand Trend denominator is zero or negative")
    return float(forecast_total / actual_total - Decimal("1"))


def _scope_keys(
    snapshot: SourceSnapshot,
    *,
    store_id: str | None = None,
    sku_id: str | None = None,
    cat: str | None = None,
    vertical_id: str | None = None,
) -> set[tuple[str, str]]:
    return {
        source.key
        for source in snapshot.source_rows
        if (store_id is None or source.store_id == store_id)
        and (sku_id is None or source.sku_id == sku_id)
        and (cat is None or source.cat == cat)
        and (vertical_id is None or source.vertical_id == vertical_id)
    }


def _pick_category(snapshot: SourceSnapshot) -> str:
    categories = sorted({source.cat for source in snapshot.source_rows})
    if "GRC-C01" in categories:
        return "GRC-C01"
    return categories[0]


def _pick_sku(snapshot: SourceSnapshot) -> str:
    skus = sorted({source.sku_id for source in snapshot.source_rows})
    if "GRC-001" in skus:
        return "GRC-001"
    return skus[0]


def _regression_scope_specs(snapshot: SourceSnapshot) -> list[tuple[str, set[tuple[str, str]]]]:
    category = _pick_category(snapshot)
    sku = _pick_sku(snapshot)
    store = "S001" if "S001" in snapshot.store_ids else snapshot.store_ids[0]
    return [
        ("ALL", _scope_keys(snapshot)),
        ("GRC", _scope_keys(snapshot, vertical_id="GRC")),
        (store, _scope_keys(snapshot, store_id=store)),
        (category, _scope_keys(snapshot, cat=category)),
        (sku, _scope_keys(snapshot, sku_id=sku)),
        (f"{store} + {category}", _scope_keys(snapshot, store_id=store, cat=category)),
        (f"{store} + {sku}", _scope_keys(snapshot, store_id=store, sku_id=sku)),
    ]


def demand_trend_regression(
    rows: Sequence[DemandSkuStore104WRow],
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
    snapshot: SourceSnapshot,
) -> dict[str, Any]:
    previous_lookup = {row.key: row for row in previous_rows}
    result: dict[str, Any] = {"scopes": {}, "changed_scope_count": 0}
    for label, keys in _regression_scope_specs(snapshot):
        before = previous_generator.calculate_trend(
            [previous_lookup[key] for key in keys]
        )
        after = calculate_trend(rows, keys)
        difference = after - before
        result["scopes"][label] = {
            "row_count": len(keys),
            "before_trend_pct": before * 100.0,
            "after_trend_pct": after * 100.0,
            "difference_pct_points": difference * 100.0,
            "unchanged": difference == 0.0,
        }
        if difference != 0.0:
            result["changed_scope_count"] += 1
    _require(result["changed_scope_count"] == 0, "Demand Trend changed in the preserved block")
    return result


def _relative_change(previous: Decimal, following: Decimal) -> float:
    if previous <= 0:
        return 0.0 if following == 0 else float("inf")
    return float(following / previous - Decimal("1"))


def _adjacent_changes(row: DemandSkuStore104WRow) -> list[dict[str, Any]]:
    changes: list[dict[str, Any]] = []
    blocks = (
        ("historical", [f"actual_w{week}" for week in range(52, 0, -1)], row.actuals),
        ("future", [f"forecast_w{week}" for week in range(1, 53)], row.forecasts),
    )
    for block, labels, values in blocks:
        for index in range(len(values) - 1):
            changes.append(
                {
                    "block": block,
                    "from": labels[index],
                    "to": labels[index + 1],
                    "relative_change": _relative_change(values[index], values[index + 1]),
                    "from_value": float(values[index]),
                    "to_value": float(values[index + 1]),
                    "sku_id": row.sku_id,
                    "store_id": row.store_id,
                }
            )
    return changes


def _summary(values: Sequence[Decimal | float]) -> dict[str, float]:
    numbers = [float(value) for value in values]
    _require(bool(numbers), "Cannot summarize empty values")
    return {"min": min(numbers), "median": statistics.median(numbers), "max": max(numbers)}


def _boundary_summary(
    rows: Sequence[DemandSkuStore104WRow],
    *,
    block: str,
    from_week: int,
    to_week: int,
) -> dict[str, Any]:
    changes: list[dict[str, Any]] = []
    for row in rows:
        if block == "historical":
            previous = row.actual_for_week(from_week)
            following = row.actual_for_week(to_week)
            from_label = f"actual_w{from_week}"
            to_label = f"actual_w{to_week}"
        else:
            previous = row.forecast_for_week(from_week)
            following = row.forecast_for_week(to_week)
            from_label = f"forecast_w{from_week}"
            to_label = f"forecast_w{to_week}"
        changes.append(
            {
                "sku_id": row.sku_id,
                "store_id": row.store_id,
                "block": block,
                "from": from_label,
                "to": to_label,
                "relative_change": _relative_change(previous, following),
                "from_value": float(previous),
                "to_value": float(following),
            }
        )
    finite = [item for item in changes if math.isfinite(item["relative_change"])]
    _require(len(finite) == EXPECTED_ROW_COUNT, f"Non-finite {block} boundary change")
    absolute = [abs(item["relative_change"]) for item in finite]
    result = {
        "from": from_label,
        "to": to_label,
        "row_count": len(changes),
        "max_abs_percentage_jump": max(absolute) * 100.0,
        "median_abs_percentage_jump": statistics.median(absolute) * 100.0,
        "max_positive_percentage_change": max(item["relative_change"] for item in finite) * 100.0,
        "max_negative_percentage_change": min(item["relative_change"] for item in finite) * 100.0,
        "rows_exceeding_threshold": sum(
            abs(item["relative_change"]) > MAX_ADJACENT_CHANGE for item in finite
        ),
        "top_10_positive": sorted(
            finite, key=lambda item: item["relative_change"], reverse=True
        )[:10],
        "top_10_negative": sorted(
            finite, key=lambda item: item["relative_change"]
        )[:10],
    }
    return result


def plausibility_summary(rows: Sequence[DemandSkuStore104WRow]) -> dict[str, Any]:
    row_trends = {row.key: calculate_trend([row]) for row in rows}
    extrema_by_block: dict[str, dict[str, dict[str, Any] | None]] = {
        "historical": {"increase": None, "decrease": None},
        "future": {"increase": None, "decrease": None},
    }
    flat_count = 0
    extreme_count = 0
    for row in rows:
        history = [float(value) for value in row.actuals]
        mean = statistics.mean(history)
        if mean > 0 and statistics.pstdev(history) / mean < FLAT_HISTORY_CV_THRESHOLD:
            flat_count += 1
        row_extreme = False
        for change in _adjacent_changes(row):
            relative = change["relative_change"]
            if not math.isfinite(relative):
                row_extreme = True
                continue
            block_extrema = extrema_by_block[change["block"]]
            if (
                block_extrema["increase"] is None
                or relative > block_extrema["increase"]["relative_change"]
            ):
                block_extrema["increase"] = change
            if (
                block_extrema["decrease"] is None
                or relative < block_extrema["decrease"]["relative_change"]
            ):
                block_extrema["decrease"] = change
            if abs(relative) > MAX_ADJACENT_CHANGE:
                row_extreme = True
        if abs(row_trends[row.key]) > MAX_ROW_TREND_ABS or row_extreme:
            extreme_count += 1

    def extrema(block: str, direction: str) -> dict[str, Any]:
        chosen = extrema_by_block[block][direction]
        if chosen is None:
            raise GenerationError(f"No finite {block} adjacent changes")
        return chosen

    return {
        "actual_w52": _summary([row.actual_for_week(52) for row in rows]),
        "actual_w1": _summary([row.actual_for_week(1) for row in rows]),
        "forecast_w1": _summary([row.forecast_for_week(1) for row in rows]),
        "forecast_w52": _summary([row.forecast_for_week(52) for row in rows]),
        "largest_adjacent_historical_increase": extrema("historical", "increase"),
        "largest_adjacent_historical_decrease": extrema("historical", "decrease"),
        "largest_adjacent_forecast_increase": extrema("future", "increase"),
        "largest_adjacent_forecast_decrease": extrema("future", "decrease"),
        "suspiciously_flat_series_count": flat_count,
        "extreme_volatility_count": extreme_count,
        "thresholds": {
            "max_adjacent_change": MAX_ADJACENT_CHANGE,
            "max_row_trend_abs": MAX_ROW_TREND_ABS,
            "flat_history_cv_lt": FLAT_HISTORY_CV_THRESHOLD,
        },
    }


def preservation_validation(
    rows: Sequence[DemandSkuStore104WRow],
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
) -> dict[str, Any]:
    current = _rows_by_key(rows)
    previous = {row.key: row for row in previous_rows}
    _require(set(current) == set(previous), "New and approved 32W keys differ")
    changed_values = 0
    changed_rows: set[tuple[str, str]] = set()
    for key, old_row in previous.items():
        old_values = old_row.business_dict()
        new_values = current[key].business_dict()
        for column in PRESERVED_COLUMNS:
            if old_values[column] != new_values[column]:
                changed_values += 1
                changed_rows.add(key)
    result = {
        "matching_rows": len(previous),
        "preserved_rows": len(previous) - len(changed_rows),
        "preserved_period_value_count": EXISTING_PRESERVED_VALUE_COUNT,
        "changed_existing_values": changed_values,
        "changed_existing_rows": len(changed_rows),
        "result": changed_values == 0,
        "columns_checked": list(PRESERVED_COLUMNS),
    }
    _require(result["matching_rows"] == EXPECTED_ROW_COUNT, "Expected 16,000 preserved rows")
    _require(result["changed_existing_values"] == 0, "An approved 32W value changed")
    return result


def validate_dataset(
    rows: Sequence[DemandSkuStore104WRow],
    snapshot: SourceSnapshot,
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
) -> dict[str, Any]:
    previous_generator.validate_source_snapshot(snapshot)
    source_by_key = {source.key: source for source in snapshot.source_rows}
    output_by_key = _rows_by_key(rows)
    _require(len(rows) == EXPECTED_ROW_COUNT, f"Expected 16,000 rows, found {len(rows)}")
    _require(set(output_by_key) == set(source_by_key), "Output keys differ from source")
    _require(
        all(tuple(row.business_dict()) == BUSINESS_COLUMNS for row in rows),
        "Output row schema differs from the 107-column contract",
    )
    _require(len({row.sku_id for row in rows}) == EXPECTED_SKU_COUNT, "Expected 800 SKUs")
    _require(len({row.store_id for row in rows}) == EXPECTED_STORE_COUNT, "Expected 160 stores")
    _require(
        {sum(row.store_id == store for row in rows) for store in {row.store_id for row in rows}}
        == {EXPECTED_ROWS_PER_STORE},
        "Every store must have 100 rows",
    )
    for row in rows:
        source = source_by_key[row.key]
        _require(row.cat == source.cat, f"Category mismatch for {row.key}")
        _require(len(row.actuals) == len(row.forecasts) == 52, f"Period count mismatch for {row.key}")
        _require(
            all(value.is_finite() and value >= 0 for value in (*row.actuals, *row.forecasts)),
            f"Invalid period value for {row.key}",
        )
        _require(
            row.forecast_for_week(1) == quantize_qty(source.forecast_7d),
            f"forecast_w1 source mismatch for {row.key}",
        )

    previous_by_key = {row.key: row for row in previous_rows}
    differences = {
        key: abs(row.forecast_for_week(1) - source_by_key[key].forecast_7d)
        for key, row in output_by_key.items()
    }
    passed = sum(value <= W1_RECONCILIATION_TOLERANCE for value in differences.values())
    _require(passed == EXPECTED_ROW_COUNT, "W+1 source reconciliation failed")
    preservation = preservation_validation(rows, previous_rows)
    trend_regression = demand_trend_regression(rows, previous_rows, snapshot)
    continuity = {
        "historical_w17_to_w16": _boundary_summary(
            rows, block="historical", from_week=17, to_week=16
        ),
        "future_w16_to_w17": _boundary_summary(
            rows, block="future", from_week=16, to_week=17
        ),
    }
    plausibility = plausibility_summary(rows)
    _require(
        plausibility["extreme_volatility_count"] == 0,
        "One or more extended series breached plausibility thresholds",
    )
    source_total = sum((source.forecast_7d for source in snapshot.source_rows), Decimal("0"))
    generated_total = sum((row.forecast_for_week(1) for row in rows), Decimal("0"))
    return {
        "shape": {
            "row_count": len(rows),
            "column_count": len(BUSINESS_COLUMNS),
            "sku_count": len({row.sku_id for row in rows}),
            "store_count": len({row.store_id for row in rows}),
            "rows_per_store": EXPECTED_ROWS_PER_STORE,
            "unique_sku_store_pairs": len(output_by_key),
        },
        "value_counts": {
            "historical_synthetic": HISTORICAL_VALUE_COUNT,
            "source_w1": SOURCE_W1_VALUE_COUNT,
            "synthetic_future": SYNTHETIC_FUTURE_VALUE_COUNT,
            "total_period_values": TOTAL_PERIOD_VALUE_COUNT,
            "existing_preserved": EXISTING_PRESERVED_VALUE_COUNT,
            "new_historical": NEW_HISTORICAL_VALUE_COUNT,
            "new_future": NEW_FUTURE_VALUE_COUNT,
        },
        "category_mapping": {
            "non_null_rows": sum(bool(row.cat.strip()) for row in rows),
            "null_rows": sum(not bool(row.cat.strip()) for row in rows),
        },
        "w1_reconciliation": {
            "source_row_count": EXPECTED_ROW_COUNT,
            "passed_count": passed,
            "failed_count": EXPECTED_ROW_COUNT - passed,
            "max_numeric_difference": decimal_text(max(differences.values())),
            "total_source_forecast_7d": decimal_text(source_total),
            "total_generated_forecast_w1": decimal_text(generated_total),
            "total_difference": decimal_text(abs(generated_total - source_total)),
        },
        "preservation": preservation,
        "demand_trend_regression": trend_regression,
        "continuity": continuity,
        "plausibility": plausibility,
        "business_columns": list(BUSINESS_COLUMNS),
        "previous_32w_fingerprint": APPROVED_PREVIOUS_FINGERPRINT,
        "source_row_count": len(source_by_key),
        "previous_by_key_count": len(previous_by_key),
    }


def _xlsx_value(column: str, value: str) -> Any:
    if column in {"sku_id", "store_id", "cat"}:
        return value
    return float(Decimal(value))


def write_csv(rows: Sequence[DemandSkuStore104WRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=BUSINESS_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in sorted(rows, key=lambda item: item.key):
            writer.writerow(row.business_dict())


def _append_bold_header(sheet: Any, values: Sequence[Any], font: Any) -> None:
    sheet.append(list(values))
    for cell in sheet[1]:
        cell.font = font


def write_xlsx(
    rows: Sequence[DemandSkuStore104WRow],
    validation: Mapping[str, Any],
    path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Demand Store SKU 104W"
    header_font = Font(bold=True)
    _append_bold_header(sheet, BUSINESS_COLUMNS, header_font)
    for row in sorted(rows, key=lambda item: item.key):
        values = row.business_dict()
        sheet.append([_xlsx_value(column, values[column]) for column in BUSINESS_COLUMNS])
    sheet.freeze_panes = "D2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(BUSINESS_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(13, len(column) + 2)

    trend_sheet = workbook.create_sheet("Trend Summary")
    _append_bold_header(
        trend_sheet,
        ["Scope", "Rows", "Before Trend %", "After Trend %", "Difference % points", "Unchanged"],
        header_font,
    )
    for label, item in validation["demand_trend_regression"]["scopes"].items():
        trend_sheet.append(
            [
                label,
                item["row_count"],
                item["before_trend_pct"],
                item["after_trend_pct"],
                item["difference_pct_points"],
                item["unchanged"],
            ]
        )
    trend_sheet.freeze_panes = "A2"
    trend_sheet.auto_filter.ref = trend_sheet.dimensions

    validation_sheet = workbook.create_sheet("Validation Summary")
    _append_bold_header(validation_sheet, ["Validation", "Result", "Detail"], header_font)
    rows_for_validation = [
        ("shape", "PASS", f"{validation['shape']['row_count']} rows x {validation['shape']['column_count']} columns"),
        ("SKU-store uniqueness", "PASS", str(validation["shape"]["unique_sku_store_pairs"])),
        ("preserved 32W block", "PASS", f"{validation['preservation']['preserved_rows']}/{validation['preservation']['matching_rows']} rows; 0 changed values"),
        ("forecast_w1 reconciliation", "PASS", f"{validation['w1_reconciliation']['passed_count']}/{validation['w1_reconciliation']['source_row_count']} passed"),
        ("Demand Trend regression", "PASS", f"{validation['demand_trend_regression']['changed_scope_count']} changed scopes"),
        ("boundary continuity", "PASS", "W17->W16 and W16->W17 within thresholds"),
        ("non-negative finite values", "PASS", str(validation["value_counts"]["total_period_values"])),
        ("CSV/XLSX parity", "PASS", "canonical sheet checked after save"),
    ]
    for item in rows_for_validation:
        validation_sheet.append(list(item))
    validation_sheet.freeze_panes = "A2"
    validation_sheet.auto_filter.ref = validation_sheet.dimensions

    continuity_sheet = workbook.create_sheet("Extension Continuity")
    _append_bold_header(
        continuity_sheet,
        ["Boundary", "Rows", "Max abs jump %", "Median abs jump %", "Rows over threshold", "Top change"],
        header_font,
    )
    for label, item in validation["continuity"].items():
        top = item["top_10_positive"][0]
        continuity_sheet.append(
            [
                label,
                item["row_count"],
                item["max_abs_percentage_jump"],
                item["median_abs_percentage_jump"],
                item["rows_exceeding_threshold"],
                f"{top['sku_id']} / {top['store_id']} {top['relative_change']:.4%}",
            ]
        )
    continuity_sheet.freeze_panes = "A2"
    continuity_sheet.auto_filter.ref = continuity_sheet.dimensions

    workbook.save(path)


def _normalise_export_value(column: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    if column in {"sku_id", "store_id", "cat"}:
        return str(value)
    return quantity_text(value)


def validate_export_parity(csv_path: Path, xlsx_path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    _require(len(csv_rows) == EXPECTED_ROW_COUNT, "CSV does not contain 16,000 rows")
    csv_canonical = [
        tuple(_normalise_export_value(column, row[column]) for column in BUSINESS_COLUMNS)
        for row in csv_rows
    ]
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        _require("Demand Store SKU 104W" in workbook.sheetnames, "Missing 104W canonical sheet")
        sheet = workbook["Demand Store SKU 104W"]
        values = sheet.iter_rows(values_only=True)
        header = tuple(next(values))
        _require(header == BUSINESS_COLUMNS, "XLSX schema differs from CSV")
        xlsx_canonical = [
            tuple(_normalise_export_value(column, value) for column, value in zip(BUSINESS_COLUMNS, row))
            for row in values
        ]
    finally:
        workbook.close()
    _require(len(xlsx_canonical) == EXPECTED_ROW_COUNT, "XLSX does not contain 16,000 rows")
    _require(csv_canonical == xlsx_canonical, "CSV and XLSX canonical rows differ")
    return {"csv_rows": len(csv_canonical), "xlsx_rows": len(xlsx_canonical), "same_logical_rows": True}


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return decimal_text(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple, set)):
        return [_json_safe(item) for item in value]
    return value


def write_manifest(manifest: Mapping[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_json_safe(dict(manifest)), indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _reproducibility_check(
    snapshot: SourceSnapshot,
    previous_rows: Sequence[previous_generator.DemandSkuStoreRow],
    params: GeneratorParameters,
    rows: Sequence[DemandSkuStore104WRow],
    fingerprint: str,
) -> dict[str, Any]:
    rerun = generate_dataset(snapshot, previous_rows, params)
    rerun_fingerprint = output_fingerprint(rerun)
    same_rows = [canonical_row_text(row) for row in rows] == [canonical_row_text(row) for row in rerun]
    negative_params = replace(params, seed=params.seed + 1)
    negative_rows = generate_dataset(snapshot, previous_rows, negative_params)
    original = _rows_by_key(rows)
    negative = _rows_by_key(negative_rows)
    preserved_same = all(
        original[key].actuals[36:] == negative[key].actuals[36:]
        and original[key].forecasts[:16] == negative[key].forecasts[:16]
        for key in original
    )
    new_history_changed = any(
        original[key].actuals[:36] != negative[key].actuals[:36] for key in original
    )
    new_future_changed = any(
        original[key].forecasts[16:] != negative[key].forecasts[16:] for key in original
    )
    identifiers_same = all(
        (original[key].sku_id, original[key].store_id, original[key].cat)
        == (negative[key].sku_id, negative[key].store_id, negative[key].cat)
        for key in original
    )
    result = {
        "same_fingerprint": fingerprint == rerun_fingerprint,
        "same_rows": same_rows,
        "rerun_output_fingerprint": rerun_fingerprint,
        "negative_control_seed": negative_params.seed,
        "negative_control_new_historical_values_differ": new_history_changed,
        "negative_control_new_future_values_differ": new_future_changed,
        "negative_control_preserved_32w_values_same": preserved_same,
        "negative_control_identifiers_categories_same": identifiers_same,
        "negative_control_output_fingerprint": output_fingerprint(negative_rows),
    }
    _require(result["same_fingerprint"] and result["same_rows"], "Same-seed reproducibility failed")
    _require(result["negative_control_new_historical_values_differ"], "Changed seed did not change new history")
    _require(result["negative_control_new_future_values_differ"], "Changed seed did not change new future")
    _require(result["negative_control_preserved_32w_values_same"], "Changed seed changed preserved 32W")
    _require(result["negative_control_identifiers_categories_same"], "Changed seed changed identifiers/categories")
    return result


def build_manifest(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    previous_fingerprint: str,
    generated_at: str,
    reproducibility: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "generation_name": params.generation_name,
        "generator_version": params.generator_version,
        "seed": params.seed,
        "source_revision": snapshot.source_revision,
        "source_import_batch": snapshot.batch_id,
        "source_snapshot_date": snapshot.source_snapshot_date.isoformat(),
        "source_row_count": snapshot.fact_rows,
        "output_row_count": validation["shape"]["row_count"],
        "output_column_count": validation["shape"]["column_count"],
        "sku_count": validation["shape"]["sku_count"],
        "store_count": validation["shape"]["store_count"],
        "total_weekly_value_count": validation["value_counts"]["total_period_values"],
        "historical_value_count": validation["value_counts"]["historical_synthetic"],
        "source_w1_value_count": validation["value_counts"]["source_w1"],
        "synthetic_future_value_count": validation["value_counts"]["synthetic_future"],
        "existing_preserved_period_count": validation["value_counts"]["existing_preserved"],
        "newly_generated_historical_value_count": validation["value_counts"]["new_historical"],
        "newly_generated_forecast_value_count": validation["value_counts"]["new_future"],
        "input_fingerprint": input_fp,
        "output_fingerprint": output_fp,
        "previous_32w_fingerprint": previous_fingerprint,
        "generated_at": generated_at,
        "business_timezone": BUSINESS_TIMEZONE,
        "quantity_precision": 6,
        "column_contract": {
            "columns": list(BUSINESS_COLUMNS),
            "identifier_columns": ["sku_id", "store_id", "cat"],
            "historical_columns": list(ACTUAL_COLUMNS),
            "forecast_columns": list(FORECAST_COLUMNS),
            "canonical_sort": ["sku_id", "store_id"],
            "numeric_precision": 6,
        },
        "provenance": {
            "actual_w52_to_actual_w17": "newly generated synthetic history extending backward from fixed approved actual_w16",
            "actual_w16_to_actual_w1": "preserved synthetic history from approved 32W CSV",
            "forecast_w1": "preserved v8.5 source-derived Forecast 7d from approved 32W CSV",
            "forecast_w2_to_forecast_w16": "preserved synthetic forecast from approved 32W CSV",
            "forecast_w17_to_forecast_w52": "newly generated synthetic forecast extending forward from fixed approved forecast_w16",
        },
        "source_preflight": {
            "agent_name": snapshot.agent_name,
            "workbook_version": snapshot.workbook_version,
            "import_status": snapshot.import_status,
            "batch_total_rows": snapshot.batch_total_rows,
            "fact_rows": snapshot.fact_rows,
            "fact_distinct_stores": snapshot.fact_distinct_stores,
            "fact_distinct_skus": snapshot.fact_distinct_skus,
            "fact_dates": snapshot.fact_dates,
            "fact_min_date": snapshot.fact_min_date.isoformat(),
            "fact_max_date": snapshot.fact_max_date.isoformat(),
            "duplicate_source_keys": snapshot.duplicate_source_keys,
            "bad_ads": snapshot.bad_ads,
            "bad_forecast_7d": snapshot.bad_forecast_7d,
            "rows_not_batch": snapshot.rows_not_batch,
            "dim_store_rows": snapshot.dim_store_rows,
            "dim_store_distinct_ids": snapshot.dim_store_distinct_ids,
            "dim_item_rows": snapshot.dim_item_rows,
            "dim_item_distinct_ids": snapshot.dim_item_distinct_ids,
            "dim_item_duplicate_groups": snapshot.dim_item_duplicate_groups,
            "dim_item_null_categories": snapshot.dim_item_null_categories,
            "category_join_rows": snapshot.category_join_rows,
            "category_join_skus": snapshot.category_join_skus,
            "category_join_stores": snapshot.category_join_stores,
            "missing_category_items": snapshot.missing_category_items,
            "null_join_categories": snapshot.null_join_categories,
            "missing_join_stores": snapshot.missing_join_stores,
        },
        "preservation_validation": validation["preservation"],
        "continuity_validation": validation["continuity"],
        "demand_trend_regression": validation["demand_trend_regression"],
        "w1_reconciliation": validation["w1_reconciliation"],
        "plausibility": validation["plausibility"],
        "reproducibility": dict(reproducibility),
        "previous_artifact": {
            "csv": str(PREVIOUS_CSV_PATH.relative_to(REPO_ROOT)),
            "manifest": str(PREVIOUS_MANIFEST_PATH.relative_to(REPO_ROOT)),
            "status": "preserved; not overwritten",
        },
        "sql_changes_performed": False,
    }


def _fmt_pct(value: float) -> str:
    return f"{value:.2%}"


def _fmt_number(value: float) -> str:
    return f"{value:,.6f}"


def _format_change(item: Mapping[str, Any]) -> str:
    return f"{item['sku_id']} + {item['store_id']} {item['from']}→{item['to']} {_fmt_pct(item['relative_change'])}"


def _boundary_lines(boundary: Mapping[str, Any]) -> list[str]:
    return [
        f"- `{boundary['from']} -> {boundary['to']}` maximum absolute jump: `{boundary['max_abs_percentage_jump']:.4f}%`; median absolute jump: `{boundary['median_abs_percentage_jump']:.4f}%`.",
        f"- Positive boundary change maximum: `{boundary['max_positive_percentage_change']:.4f}%`; negative boundary change minimum: `{boundary['max_negative_percentage_change']:.4f}%`.",
        f"- Rows exceeding the existing `{MAX_ADJACENT_CHANGE:.0%}` adjacent-change threshold: `{boundary['rows_exceeding_threshold']}` of `{boundary['row_count']}`.",
        "- Top 10 positive boundary changes:",
        *[f"  - `{_format_change(item)}`" for item in boundary["top_10_positive"]],
        "- Top 10 negative boundary changes:",
        *[f"  - `{_format_change(item)}`" for item in boundary["top_10_negative"]],
    ]


def render_generation_report(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    previous_fingerprint: str,
    csv_path: Path,
    xlsx_path: Path,
    manifest_path: Path,
    reproducibility: Mapping[str, Any],
    test_result: str,
) -> str:
    shape = validation["shape"]
    counts = validation["value_counts"]
    trend = validation["demand_trend_regression"]["scopes"]
    continuity = validation["continuity"]
    plausibility = validation["plausibility"]
    lines = [
        "# Demand Store SKU 104W Generation Report",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Generation version | `{params.generation_name}` |",
        f"| Generator version | `{params.generator_version}` |",
        f"| Seed | `{params.seed}` |",
        f"| Source revision | `{snapshot.source_revision}` |",
        f"| Source batch | `{snapshot.batch_id}` |",
        f"| Rows | `{shape['row_count']}` |",
        f"| Columns | `{shape['column_count']}` |",
        f"| SKUs | `{shape['sku_count']}` |",
        f"| Stores | `{shape['store_count']}` |",
        f"| Total weekly values | `{counts['total_period_values']}` |",
        f"| Newly generated historical values | `{counts['new_historical']}` |",
        f"| Newly generated future values | `{counts['new_future']}` |",
        f"| Existing preserved values | `{counts['existing_preserved']}` |",
        f"| Output fingerprint | `{output_fp}` |",
        f"| Existing-block preservation | **{validation['preservation']['preserved_rows']}/{validation['preservation']['matching_rows']} rows; {validation['preservation']['changed_existing_values']} changed values** |",
        "| Overall verdict | **READY FOR REVIEW / SQL LOAD** |",
        "",
        "This is a local generation and validation artifact. Azure SQL was queried read-only for source preflight only and was not modified.",
        "",
        "## 1. Executive Summary",
        "",
        "The approved 32W SKU × Store CSV was treated as immutable input. The new output retains its 16,000 rows, identifiers, categories, actual_w16…actual_w1, and forecast_w1…forecast_w16 values exactly at six-decimal canonical precision, then adds 36 historical and 36 future periods.",
        "",
        "The canonical output remains one wide row per SKU × Store. It is not a SQL load in this task; it is ready for manual review before a separate SQL implementation task.",
        "",
        "## 2. Reason for Expansion",
        "",
        "The 32W artifact is being extended to support a 52-week historical and 52-week forecast review horizon while keeping the existing live Demand Trend calculation unchanged. The extension does not regenerate any approved period.",
        "",
        "## 3. Existing 32W Dataset Preservation",
        "",
        f"Previous artifact: `{PREVIOUS_CSV_PATH.relative_to(REPO_ROOT)}`. Previous fingerprint: `{previous_fingerprint}`.",
        f"The output comparison checked all `{len(PRESERVED_COLUMNS)}` preserved business columns for all `{EXPECTED_ROW_COUNT}` matching SKU × Store rows. Result: `{validation['preservation']['preserved_rows']}/{validation['preservation']['matching_rows']} rows preserved`; changed existing values: `{validation['preservation']['changed_existing_values']}`; changed existing rows: `{validation['preservation']['changed_existing_rows']}`.",
        "",
        "The existing `forecast_w1` remains the exact source-derived v8.5 Forecast 7d value already present in the approved CSV. No old CSV, XLSX, or manifest was overwritten.",
        "",
        "## 4. New 52+52 Schema",
        "",
        f"The canonical sheet has exactly `{shape['column_count']}` columns: 3 identifiers, 52 historical columns (`actual_w52`…`actual_w1`), and 52 forecast columns (`forecast_w1`…`forecast_w52`). No date, metadata, or provenance columns were added to the canonical sheet.",
        "",
        "```text",
        " | ".join(BUSINESS_COLUMNS),
        "```",
        "",
        "## 5. Generator Extension Method",
        "",
        "The generator imports the tested 32W source-preflight and row-model utilities but does not call the old generator to recreate the approved block. It parses the approved CSV, verifies its manifest and fingerprint, copies the fixed block, and computes only extension values from stable identifier-hashed continuations.",
        "",
        "Historical extension steps backward from fixed `actual_w16` using the existing row-specific bounded growth, seasonality, and smoothed deterministic noise model. Forecast extension steps forward from fixed `forecast_w16` with the same model family. Each new quantity is quantized to six decimals and checked non-negative and finite.",
        "",
        "## 6. Historical W-52...W-17 Generation",
        "",
        f"Generated `{counts['new_historical']}` new historical values (`36 × {shape['row_count']}`) for `actual_w52` through `actual_w17`. The first extension boundary is generated outward from the immutable `actual_w16`; `actual_w16` itself is never modified.",
        "",
        "## 7. Forecast W+17...W+52 Generation",
        "",
        f"Generated `{counts['new_future']}` new future values (`36 × {shape['row_count']}`) for `forecast_w17` through `forecast_w52`. The first extension boundary is generated outward from immutable `forecast_w16`; `forecast_w1` through `forecast_w16` are copied unchanged.",
        "",
        "## 8. Boundary Continuity",
        "",
        "The existing adjacent-change threshold is 20%. Boundary statistics and the top 10 positive/negative changes are reported below; the boundary is not smoothed by changing an approved value.",
        "",
        "### Historical W-17 → W-16",
        "",
        *_boundary_lines(continuity["historical_w17_to_w16"]),
        "",
        "### Forecast W+16 → W+17",
        "",
        *_boundary_lines(continuity["future_w16_to_w17"]),
        "",
        "## 9. Existing-Block Regression",
        "",
        f"Hard gate result: `{validation['preservation']['preserved_rows']}/{validation['preservation']['matching_rows']} rows preserved`, with `{validation['preservation']['changed_existing_values']}` changed existing period values. Any non-zero result would make the output NOT READY.",
        "",
        "## 10. Demand Trend Regression",
        "",
        "Demand Trend remains the fixed aggregate-before-divide formula over W-4…W-1 versus W+1…W+4. Because all those columns are preserved, every required scope is unchanged.",
        "",
        "| Scope | Rows | Before | After | Difference |",
        "|---|---:|---:|---:|---:|",
        *[
            f"| {label} | {item['row_count']} | {_fmt_pct(item['before_trend_pct'] / 100)} | {_fmt_pct(item['after_trend_pct'] / 100)} | `{item['difference_pct_points']:.12f} pp` |"
            for label, item in trend.items()
        ],
        "",
        "All seven scopes passed exact regression comparison: ALL, GRC, S001, GRC-C01, GRC-001, S001 + GRC-C01, and S001 + GRC-001.",
        "",
        "## 11. Output Artifacts",
        "",
        f"- CSV: `{csv_path.relative_to(REPO_ROOT)}` — canonical future SQL-load artifact.",
        f"- XLSX: `{xlsx_path.relative_to(REPO_ROOT)}` — main sheet `Demand Store SKU 104W` contains exactly `{shape['row_count']} × {shape['column_count']}` canonical rows/columns; review sheets include Trend Summary, Validation Summary, and Extension Continuity.",
        f"- Manifest: `{manifest_path.relative_to(REPO_ROOT)}` — provenance, fingerprints, preservation, continuity, source preflight, and reproducibility metadata.",
        "",
        "## 12. Shape and Value Counts",
        "",
        "| Block | Values | Provenance |",
        "|---|---:|---|",
        f"| Historical actuals, W-52…W-1 | `{counts['historical_synthetic']}` | `{counts['new_historical']}` new synthetic + `{EXISTING_PRESERVED_VALUE_COUNT - SOURCE_W1_VALUE_COUNT}` preserved synthetic |",
        f"| Forecast W+1 | `{counts['source_w1']}` | Preserved v8.5 source-derived values |",
        f"| Forecast W+2…W+16 | `{EXPECTED_ROW_COUNT * 15}` | Preserved synthetic values |",
        f"| Forecast W+17…W+52 | `{counts['new_future']}` | New synthetic values |",
        f"| **Total period values** | **`{counts['total_period_values']}`** | **16,000 rows, not 1,664,000 rows** |",
        "",
        f"Shape checks passed: `{shape['row_count']}` rows, `{shape['column_count']}` columns, `{shape['sku_count']}` SKUs, `{shape['store_count']}` stores, `{shape['rows_per_store']}` rows/store, `{shape['unique_sku_store_pairs']}` unique SKU-store pairs, and `{validation['category_mapping']['null_rows']}` null categories.",
        "",
        "## 13. Plausibility Review",
        "",
        f"- `actual_w52` min/median/max: `{_fmt_number(plausibility['actual_w52']['min'])}` / `{_fmt_number(plausibility['actual_w52']['median'])}` / `{_fmt_number(plausibility['actual_w52']['max'])}`.",
        f"- `actual_w1` min/median/max: `{_fmt_number(plausibility['actual_w1']['min'])}` / `{_fmt_number(plausibility['actual_w1']['median'])}` / `{_fmt_number(plausibility['actual_w1']['max'])}`.",
        f"- `forecast_w1` min/median/max: `{_fmt_number(plausibility['forecast_w1']['min'])}` / `{_fmt_number(plausibility['forecast_w1']['median'])}` / `{_fmt_number(plausibility['forecast_w1']['max'])}`.",
        f"- `forecast_w52` min/median/max: `{_fmt_number(plausibility['forecast_w52']['min'])}` / `{_fmt_number(plausibility['forecast_w52']['median'])}` / `{_fmt_number(plausibility['forecast_w52']['max'])}`.",
        f"- Largest adjacent historical increase: `{_format_change(plausibility['largest_adjacent_historical_increase'])}`; decrease: `{_format_change(plausibility['largest_adjacent_historical_decrease'])}`.",
        f"- Largest adjacent forecast increase: `{_format_change(plausibility['largest_adjacent_forecast_increase'])}`; decrease: `{_format_change(plausibility['largest_adjacent_forecast_decrease'])}`.",
        f"- Suspiciously flat 52-week historical series: `{plausibility['suspiciously_flat_series_count']}`.",
        f"- Extreme volatility series: `{plausibility['extreme_volatility_count']}`; thresholds: adjacent change > `{MAX_ADJACENT_CHANGE:.0%}` or row Trend beyond ±`{MAX_ROW_TREND_ABS:.0%}`.",
        f"- GRC Trend regression: `{_fmt_pct(trend['GRC']['after_trend_pct'] / 100)}`; all-row Trend regression: `{_fmt_pct(trend['ALL']['after_trend_pct'] / 100)}`.",
        "",
        "## 14. Reproducibility",
        "",
        f"Same-seed output fingerprint: `{output_fp}`; rerun identical fingerprint: `{reproducibility['same_fingerprint']}`; identical canonical rows: `{reproducibility['same_rows']}`.",
        f"Changed-seed new history changed: `{reproducibility['negative_control_new_historical_values_differ']}`; changed-seed new future changed: `{reproducibility['negative_control_new_future_values_differ']}`; preserved 32W remained identical: `{reproducibility['negative_control_preserved_32w_values_same']}`; identifiers/categories remained identical: `{reproducibility['negative_control_identifiers_categories_same']}`.",
        f"Input fingerprint: `{input_fp}`; previous 32W fingerprint: `{previous_fingerprint}`; output fingerprint: `{output_fp}`.",
        "",
        "The output fingerprint hashes canonical rows sorted by `sku_id, store_id`, using all 107 business columns and six-decimal quantity formatting. XLSX formatting and volatile generated timestamps are excluded.",
        "",
        "## 15. Tests",
        "",
        test_result,
        "",
        "The generator also performs same-seed and changed-seed negative-control validation during artifact generation, plus CSV/XLSX canonical parity after writing.",
        "",
        "## 16. Known Limitations",
        "",
        "- All historical values remain synthetic; the new W-52…W-17 values are a deterministic continuation, not genuine sales history.",
        "- W+2…W+52 remain synthetic; only W+1 is source-derived, and the existing W+1 is preserved unchanged.",
        "- Fixed W-52…W+52 labels carry no explicit dates in this simplified wide POC table.",
        "- The extension is model-based and does not claim governed POS, returns, cancellations, stockout-censored demand, or forecast-run semantics.",
        "- This task did not modify Azure SQL, backend runtime behavior, frontend behavior, Demand Trend, or forecast charts.",
        "",
        "## 17. Azure SQL Load Readiness",
        "",
        "**READY FOR REVIEW / SQL LOAD.** Local shape, non-negative/finite, exact preservation, source W+1, continuity, Trend regression, CSV/XLSX parity, and reproducibility checks passed. No Azure SQL writes were performed. A separate authorized task must review and implement any SQL schema/load changes.",
        "",
    ]
    return "\n".join(lines)


def run_generation(
    *,
    output_dir: Path,
    params: GeneratorParameters,
    test_result: str,
    previous_csv: Path = PREVIOUS_CSV_PATH,
    previous_manifest: Path = PREVIOUS_MANIFEST_PATH,
) -> dict[str, Any]:
    previous_rows, previous_manifest_data, previous_fp = load_previous_rows(
        previous_csv, previous_manifest
    )
    # This is the only live operation. The imported 32W preflight is SELECT-only.
    snapshot = previous_generator.load_source_snapshot()
    validate_previous_against_source(previous_rows, snapshot)
    rows = generate_dataset(snapshot, previous_rows, params)
    validation = validate_dataset(rows, snapshot, previous_rows)
    output_fp = output_fingerprint(rows)
    input_fp = input_fingerprint(snapshot, params, previous_fp)
    reproducibility = _reproducibility_check(
        snapshot, previous_rows, params, rows, output_fp
    )

    csv_path = output_dir / "demand_store_sku_104w_poc_v1.csv"
    xlsx_path = output_dir / "demand_store_sku_104w_poc_v1.xlsx"
    manifest_path = output_dir / "demand_store_sku_104w_poc_v1_manifest.json"
    write_csv(rows, csv_path)
    write_xlsx(rows, validation, xlsx_path)
    export_parity = validate_export_parity(csv_path, xlsx_path)
    validation["export_parity"] = export_parity
    generated_at = datetime.now(timezone.utc).isoformat()
    manifest = build_manifest(
        snapshot=snapshot,
        params=params,
        validation=validation,
        input_fp=input_fp,
        output_fp=output_fp,
        previous_fingerprint=previous_fp,
        generated_at=generated_at,
        reproducibility=reproducibility,
    )
    write_manifest(manifest, manifest_path)
    report = render_generation_report(
        snapshot=snapshot,
        params=params,
        validation=validation,
        input_fp=input_fp,
        output_fp=output_fp,
        previous_fingerprint=previous_fp,
        csv_path=csv_path,
        xlsx_path=xlsx_path,
        manifest_path=manifest_path,
        reproducibility=reproducibility,
        test_result=test_result,
    )
    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(report, encoding="utf-8")
    return {
        "snapshot": snapshot,
        "previous_manifest": previous_manifest_data,
        "rows": rows,
        "validation": validation,
        "input_fingerprint": input_fp,
        "previous_fingerprint": previous_fp,
        "output_fingerprint": output_fp,
        "reproducibility": reproducibility,
        "csv_path": csv_path,
        "xlsx_path": xlsx_path,
        "manifest_path": manifest_path,
        "report_path": REPORT_PATH,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", type=Path, default=REPO_ROOT / "artifacts")
    parser.add_argument("--previous-csv", type=Path, default=PREVIOUS_CSV_PATH)
    parser.add_argument("--previous-manifest", type=Path, default=PREVIOUS_MANIFEST_PATH)
    parser.add_argument("--seed", type=int, default=FIXED_SEED)
    parser.add_argument("--generation-name", default=GENERATION_NAME)
    parser.add_argument("--generator-version", default=GENERATOR_VERSION)
    parser.add_argument(
        "--test-result",
        default="Focused 104W generator tests: run separately; see final handoff.",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    params = GeneratorParameters(
        generation_name=args.generation_name,
        generator_version=args.generator_version,
        seed=args.seed,
    )
    try:
        result = run_generation(
            output_dir=args.output_dir,
            params=params,
            test_result=args.test_result,
            previous_csv=args.previous_csv,
            previous_manifest=args.previous_manifest,
        )
    except SourcePreflightError as exc:
        print(f"SOURCE PREFLIGHT FAILED: {exc}", file=sys.stderr)
        return 2
    except GenerationError as exc:
        print(f"GENERATION/VALIDATION FAILED: {exc}", file=sys.stderr)
        return 3
    print(
        json.dumps(
            {
                "report_path": str(result["report_path"]),
                "generator_path": str(Path(__file__).resolve()),
                "csv_path": str(result["csv_path"]),
                "xlsx_path": str(result["xlsx_path"]),
                "manifest_path": str(result["manifest_path"]),
                "previous_fingerprint": result["previous_fingerprint"],
                "output_fingerprint": result["output_fingerprint"],
                "shape": result["validation"]["shape"],
                "value_counts": result["validation"]["value_counts"],
                "preservation": result["validation"]["preservation"],
                "continuity": result["validation"]["continuity"],
                "trend": result["validation"]["demand_trend_regression"],
                "reproducibility": result["reproducibility"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
