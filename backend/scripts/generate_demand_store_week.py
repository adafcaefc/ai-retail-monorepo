"""Generate and validate the local synthetic demand store-week candidate.

This script is deliberately an offline/export step.  It reads the current
v8.5 source through SELECT-only queries, materialises CSV/XLSX/JSON locally,
and never creates or changes Azure SQL objects or rows.

Run from the repository root with the repository virtualenv, for example:

    .venv/bin/python backend/scripts/generate_demand_store_week.py

The pure generation and validation functions accept an in-memory
``SourceSnapshot`` so they can be tested without a database connection.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import random
import statistics
import sys
from dataclasses import asdict, dataclass, replace
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT / "backend") not in sys.path:
    sys.path.insert(0, str(REPO_ROOT / "backend"))


GENERATION_VERSION = "demand_store_week_poc_v1"
GENERATOR_VERSION = "demand-store-week-generator-v1.0.0"
FIXED_SEED = 20260820
SOURCE_IMPORT_BATCH_ID = 23
SOURCE_AS_OF_DATE = date(2026, 7, 1)
SOURCE_REVISION_SHA = "a6f4c7fabae4c27af9c4c035adb8252fa27cc062ea6b4bad4b0f81c976ec9510"
SOURCE_REVISION = f"v8.5:{SOURCE_REVISION_SHA}"
BUSINESS_TIMEZONE = "Asia/Jakarta"
DOW_PROFILE = (0.85, 0.90, 0.95, 1.00, 1.15, 1.35, 1.25)
DOW_SUM = Decimal("7.45")
QUANTITY_QUANTUM = Decimal("0.000001")
W1_TOLERANCE = Decimal("0.0001")
PERIOD_OFFSETS = tuple(range(-16, 0)) + tuple(range(1, 17))
EXPECTED_STORE_IDS = tuple(f"S{i:03d}" for i in range(1, 161))
EXPECTED_PERIOD_OFFSETS = frozenset(PERIOD_OFFSETS)
NEAR_ZERO_TREND_THRESHOLD = 0.01
MAX_STORE_TREND_ABS = 0.35
MAX_WEEK_OVER_WEEK_ABS_CHANGE = 0.20
FLAT_HISTORY_COEFFICIENT_OF_VARIATION = 0.001

CSV_COLUMNS = (
    "generation_version",
    "store_id",
    "week_start",
    "week_end",
    "period_offset",
    "actual_qty",
    "forecast_qty",
    "source_as_of_date",
    "forecast_as_of_date",
    "data_source",
    "source_semantics",
    "is_synthetic",
    "source_revision",
    "source_import_batch_id",
    "generator_version",
    "generator_seed",
)


class GenerationError(RuntimeError):
    """Raised when source preflight, generation, or offline validation fails."""


class SourcePreflightError(GenerationError):
    """Raised when the current database is not the approved source revision."""


@dataclass(frozen=True)
class GeneratorParameters:
    """Versioned knobs for the intentionally small POC model."""

    generation_version: str = GENERATION_VERSION
    generator_version: str = GENERATOR_VERSION
    seed: int = FIXED_SEED
    source_as_of_date: date = SOURCE_AS_OF_DATE
    dow_sum: Decimal = DOW_SUM

    def manifest_dict(self) -> dict[str, Any]:
        return {
            "generation_version": self.generation_version,
            "generator_version": self.generator_version,
            "seed": self.seed,
            "source_as_of_date": self.source_as_of_date.isoformat(),
            "business_timezone": BUSINESS_TIMEZONE,
            "dow_profile": list(DOW_PROFILE),
            "dow_sum": decimal_text(self.dow_sum),
            "quantity_precision": 6,
            "near_zero_trend_threshold": NEAR_ZERO_TREND_THRESHOLD,
            "max_store_trend_abs": MAX_STORE_TREND_ABS,
            "max_week_over_week_abs_change": MAX_WEEK_OVER_WEEK_ABS_CHANGE,
            "flat_history_coefficient_of_variation": FLAT_HISTORY_COEFFICIENT_OF_VARIATION,
        }


@dataclass(frozen=True)
class StoreInput:
    """The audited, store-level inputs used by the generator."""

    store_id: str
    vertical_id: str
    size_index: Decimal | None
    health_index: Decimal | None
    footfall_index: Decimal | None
    cluster: str | None
    channel: str | None
    source_rows: int
    distinct_items: int
    ads_total: Decimal
    forecast_7d_total: Decimal

    def fingerprint_dict(self) -> dict[str, Any]:
        return {
            "store_id": self.store_id,
            "vertical_id": self.vertical_id,
            "size_index": decimal_or_none(self.size_index),
            "health_index": decimal_or_none(self.health_index),
            "footfall_index": decimal_or_none(self.footfall_index),
            "cluster": self.cluster,
            "channel": self.channel,
            "source_rows": self.source_rows,
            "distinct_items": self.distinct_items,
            "ads_total": decimal_text(self.ads_total),
            "forecast_7d_total": decimal_text(self.forecast_7d_total),
        }


@dataclass(frozen=True)
class SourceSnapshot:
    """Validated source snapshot and its read-only preflight evidence."""

    stores: tuple[StoreInput, ...]
    batch_id: int
    agent_name: str
    workbook_version: str
    import_status: str
    source_revision: str
    source_sha256: str
    source_as_of_date: date
    batch_total_rows: int
    fact_rows: int
    fact_dates: int
    fact_min_date: date
    fact_max_date: date
    fact_distinct_stores: int
    fact_distinct_items: int
    duplicate_groups: int
    bad_ads: int
    bad_forecast_7d: int
    dim_store_rows: int
    dim_store_distinct_ids: int
    dim_item_rows: int

    @property
    def store_map(self) -> dict[str, StoreInput]:
        return {store.store_id: store for store in self.stores}

    @property
    def total_forecast_7d(self) -> Decimal:
        return sum(
            (store.forecast_7d_total for store in self.stores),
            Decimal("0"),
        )

    @property
    def s001_forecast_7d(self) -> Decimal:
        return self.store_map["S001"].forecast_7d_total

    def fingerprint_dict(self) -> dict[str, Any]:
        return {
            "source_revision": self.source_revision,
            "source_sha256": self.source_sha256,
            "source_import_batch_id": self.batch_id,
            "source_as_of_date": self.source_as_of_date.isoformat(),
            "fact_rows": self.fact_rows,
            "fact_dates": self.fact_dates,
            "fact_distinct_stores": self.fact_distinct_stores,
            "fact_distinct_items": self.fact_distinct_items,
            "stores": [store.fingerprint_dict() for store in self.stores],
        }


@dataclass(frozen=True)
class DemandRow:
    """One canonical store-week row before CSV/XLSX serialisation."""

    generation_version: str
    store_id: str
    week_start: date
    week_end: date
    period_offset: int
    actual_qty: Decimal | None
    forecast_qty: Decimal | None
    source_as_of_date: date
    forecast_as_of_date: date | None
    data_source: str
    source_semantics: str
    is_synthetic: bool
    source_revision: str
    source_import_batch_id: int
    generator_version: str | None
    generator_seed: int | None

    def key(self) -> tuple[str, str, int]:
        return (self.generation_version, self.store_id, self.period_offset)

    def week_key(self) -> tuple[str, str, date]:
        return (self.generation_version, self.store_id, self.week_start)

    def csv_dict(self) -> dict[str, str | int]:
        return {
            "generation_version": self.generation_version,
            "store_id": self.store_id,
            "week_start": self.week_start.isoformat(),
            "week_end": self.week_end.isoformat(),
            "period_offset": self.period_offset,
            "actual_qty": decimal_or_blank(self.actual_qty),
            "forecast_qty": decimal_or_blank(self.forecast_qty),
            "source_as_of_date": self.source_as_of_date.isoformat(),
            "forecast_as_of_date": (
                self.forecast_as_of_date.isoformat()
                if self.forecast_as_of_date is not None
                else ""
            ),
            "data_source": self.data_source,
            "source_semantics": self.source_semantics,
            "is_synthetic": int(self.is_synthetic),
            "source_revision": self.source_revision,
            "source_import_batch_id": self.source_import_batch_id,
            "generator_version": self.generator_version or "",
            "generator_seed": (
                self.generator_seed if self.generator_seed is not None else ""
            ),
        }


def decimal_value(value: Any) -> Decimal:
    """Convert a database or test value without introducing float arithmetic."""

    if isinstance(value, Decimal):
        result = value
    else:
        result = Decimal(str(value))
    if not result.is_finite():
        raise GenerationError(f"Non-finite numeric value: {value!r}")
    return result


def quantize_qty(value: Decimal | float | int) -> Decimal:
    """Apply the contract's deterministic six-decimal final quantisation."""

    try:
        result = decimal_value(value).quantize(QUANTITY_QUANTUM, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise GenerationError(f"Unable to quantize quantity {value!r}") from exc
    if result < 0:
        raise GenerationError(f"Negative generated quantity: {result}")
    return result


def decimal_text(value: Decimal | float | int) -> str:
    """Return a stable non-exponent decimal representation."""

    return format(decimal_value(value), "f")


def decimal_or_none(value: Decimal | None) -> str | None:
    return decimal_text(value) if value is not None else None


def decimal_or_blank(value: Decimal | None) -> str:
    return decimal_text(value) if value is not None else ""


def _parse_date(value: Any) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value)[:10])


def _row_mapping(row: Any) -> Mapping[str, Any]:
    mapping = getattr(row, "_mapping", None)
    if mapping is not None:
        return mapping
    if isinstance(row, Mapping):
        return row
    raise TypeError(f"Unsupported database row type: {type(row)!r}")


def _parse_metadata(value: Any) -> dict[str, Any]:
    if isinstance(value, Mapping):
        return dict(value)
    if value is None:
        return {}
    try:
        parsed = json.loads(str(value))
    except json.JSONDecodeError as exc:
        raise SourcePreflightError("Batch metadata is not valid JSON") from exc
    if not isinstance(parsed, dict):
        raise SourcePreflightError("Batch metadata is not a JSON object")
    return parsed


def _require(condition: bool, message: str) -> None:
    if not condition:
        raise SourcePreflightError(message)


def validate_source_snapshot(snapshot: SourceSnapshot) -> None:
    """Validate the source evidence independently of the SQL connection.

    Keeping this check separate makes source-drift behavior testable without
    opening a database connection and gives callers one final guard before
    generation begins.
    """

    _require(snapshot.batch_id == SOURCE_IMPORT_BATCH_ID, "Expected import batch 23")
    _require(snapshot.workbook_version == "v8.5", "Expected source workbook v8.5")
    _require(snapshot.import_status == "COMPLETED", "Expected a completed source batch")
    _require(snapshot.source_revision == SOURCE_REVISION, "Unexpected source revision")
    _require(
        snapshot.source_as_of_date == SOURCE_AS_OF_DATE,
        "Expected source as-of date 2026-07-01",
    )
    _require(snapshot.fact_rows == 16_000, "Expected 16,000 fact rows")
    _require(snapshot.fact_distinct_stores == 160, "Expected 160 fact stores")
    _require(snapshot.fact_distinct_items == 800, "Expected 800 fact items")
    _require(
        snapshot.fact_dates == 1
        and snapshot.fact_min_date == SOURCE_AS_OF_DATE
        and snapshot.fact_max_date == SOURCE_AS_OF_DATE,
        "Expected only the 2026-07-01 fact snapshot",
    )
    _require(snapshot.duplicate_groups == 0, "Expected no source duplicate groups")
    _require(snapshot.bad_ads == 0, "Expected no bad ADS values")
    _require(snapshot.bad_forecast_7d == 0, "Expected no bad forecast_7d values")
    _require(
        snapshot.dim_store_rows == 160
        and snapshot.dim_store_distinct_ids == 160,
        "Expected 160 runtime stores",
    )
    _require(snapshot.dim_item_rows == 800, "Expected 800 runtime items")
    _require(
        tuple(store.store_id for store in snapshot.stores) == EXPECTED_STORE_IDS,
        "Expected runtime stores S001 through S160",
    )
    for store in snapshot.stores:
        _require(
            store.source_rows == 100 and store.distinct_items == 100,
            f"{store.store_id} does not have 100 unique store-SKU rows",
        )
        _require(
            store.ads_total >= 0 and store.forecast_7d_total >= 0,
            f"{store.store_id} has a negative aggregate source quantity",
        )


def load_source_snapshot() -> SourceSnapshot:
    """Run the mandatory SELECT-only Azure SQL preflight and load inputs.

    No transaction is opened for writes and no DDL/DML is present in any
    statement in this function.
    """

    from sqlalchemy import text

    from src.db.db import get_engine

    batch_sql = text(
        """
        SELECT id, agent_name, workbook_version, import_status, total_rows,
               CAST(metadata AS nvarchar(max)) AS metadata
        FROM audit.import_batches
        WHERE id = :batch_id
        """
    )
    fact_sql = text(
        """
        SELECT
            COUNT(*) AS fact_rows,
            COUNT(DISTINCT store_key) AS fact_distinct_stores,
            COUNT(DISTINCT item_key) AS fact_distinct_items,
            COUNT(DISTINCT cal_date) AS fact_dates,
            MIN(cal_date) AS fact_min_date,
            MAX(cal_date) AS fact_max_date,
            SUM(CASE WHEN ads IS NULL OR ads < 0 THEN 1 ELSE 0 END) AS bad_ads,
            SUM(CASE WHEN forecast_7d IS NULL OR forecast_7d < 0 THEN 1 ELSE 0 END)
                AS bad_forecast_7d,
            COUNT(DISTINCT import_batch_id) AS fact_batch_ids,
            MIN(import_batch_id) AS min_batch_id,
            MAX(import_batch_id) AS max_batch_id,
            SUM(CASE WHEN import_batch_id IS NULL OR import_batch_id <> :batch_id
                     THEN 1 ELSE 0 END) AS rows_not_batch,
            CAST(SUM(forecast_7d) AS decimal(30,12)) AS total_forecast_7d
        FROM retail.fact_inventory_daily
        """
    )
    store_sql = text(
        """
        SELECT
            s.store_id,
            s.vertical_id,
            s.size_index,
            s.health_index,
            s.footfall_index,
            s.cluster,
            s.channel,
            COUNT(f.item_key) AS source_rows,
            COUNT(DISTINCT f.item_key) AS distinct_items,
            CAST(SUM(f.ads) AS decimal(30,12)) AS ads_total,
            CAST(SUM(f.forecast_7d) AS decimal(30,12)) AS forecast_7d_total
        FROM retail.dim_store AS s
        LEFT JOIN retail.fact_inventory_daily AS f
          ON f.store_key = s.store_id
         AND f.cal_date = :as_of_date
        GROUP BY
            s.store_id, s.vertical_id, s.size_index, s.health_index,
            s.footfall_index, s.cluster, s.channel
        ORDER BY s.store_id
        """
    )
    dim_store_sql = text(
        """
        SELECT COUNT(*) AS dim_store_rows,
               COUNT(DISTINCT store_id) AS dim_store_distinct_ids
        FROM retail.dim_store
        """
    )
    dim_item_sql = text("SELECT COUNT(*) AS dim_item_rows FROM retail.dim_item")
    duplicate_sql = text(
        """
        SELECT COUNT(*) AS duplicate_groups
        FROM (
            SELECT store_key, item_key, cal_date
            FROM retail.fact_inventory_daily
            GROUP BY store_key, item_key, cal_date
            HAVING COUNT(*) > 1
        ) AS duplicate_keys
        """
    )

    with get_engine().connect() as connection:
        batch_row = connection.execute(
            batch_sql, {"batch_id": SOURCE_IMPORT_BATCH_ID}
        ).mappings().first()
        fact_row = connection.execute(
            fact_sql, {"batch_id": SOURCE_IMPORT_BATCH_ID}
        ).mappings().one()
        store_rows = connection.execute(
            store_sql, {"as_of_date": SOURCE_AS_OF_DATE}
        ).mappings().all()
        dim_store_row = connection.execute(dim_store_sql).mappings().one()
        dim_item_row = connection.execute(dim_item_sql).mappings().one()
        duplicate_row = connection.execute(duplicate_sql).mappings().one()

    _require(batch_row is not None, "Expected audit.import_batches.id=23 was not found")
    batch = _row_mapping(batch_row)
    metadata = _parse_metadata(batch["metadata"])
    source_sha256 = str(metadata.get("workbook_sha256", ""))

    _require(
        str(batch["workbook_version"]) == "v8.5",
        f"Source drift: batch 23 workbook_version={batch['workbook_version']!r}, expected v8.5",
    )
    _require(
        str(batch["import_status"]) == "COMPLETED",
        f"Source drift: batch 23 import_status={batch['import_status']!r}",
    )
    _require(
        str(batch["agent_name"]) == "retail_facts_seed",
        f"Source drift: batch 23 agent_name={batch['agent_name']!r}",
    )
    _require(
        source_sha256 == SOURCE_REVISION_SHA,
        "Source drift: batch 23 workbook SHA does not match the approved v8.5 revision",
    )

    fact_rows = int(fact_row["fact_rows"] or 0)
    fact_dates = int(fact_row["fact_dates"] or 0)
    fact_min_date = _parse_date(fact_row["fact_min_date"])
    fact_max_date = _parse_date(fact_row["fact_max_date"])
    _require(fact_rows == 16_000, f"Expected 16,000 fact rows, found {fact_rows}")
    _require(
        int(fact_row["fact_distinct_stores"] or 0) == 160,
        "Expected 160 fact stores",
    )
    _require(
        int(fact_row["fact_distinct_items"] or 0) == 800,
        "Expected 800 fact items",
    )
    _require(
        fact_dates == 1 and fact_min_date == SOURCE_AS_OF_DATE
        and fact_max_date == SOURCE_AS_OF_DATE,
        "Fact snapshot date drifted from 2026-07-01",
    )
    _require(
        int(fact_row["bad_ads"] or 0) == 0,
        "Source contains null or negative ADS values",
    )
    _require(
        int(fact_row["bad_forecast_7d"] or 0) == 0,
        "Source contains null or negative forecast_7d values",
    )
    _require(
        int(fact_row["fact_batch_ids"] or 0) == 1
        and int(fact_row["min_batch_id"] or 0) == SOURCE_IMPORT_BATCH_ID
        and int(fact_row["max_batch_id"] or 0) == SOURCE_IMPORT_BATCH_ID
        and int(fact_row["rows_not_batch"] or 0) == 0,
        "Fact rows are not all tied to import batch 23",
    )
    _require(
        int(duplicate_row["duplicate_groups"] or 0) == 0,
        "Duplicate (store_key, item_key, cal_date) groups found",
    )
    _require(
        int(dim_store_row["dim_store_rows"] or 0) == 160
        and int(dim_store_row["dim_store_distinct_ids"] or 0) == 160,
        "Expected 160 distinct runtime dim_store rows",
    )
    _require(
        int(dim_item_row["dim_item_rows"] or 0) == 800,
        "Expected 800 runtime dim_item rows",
    )
    _require(
        len(store_rows) == 160,
        f"Expected 160 grouped dim_store rows, found {len(store_rows)}",
    )

    stores: list[StoreInput] = []
    for row in store_rows:
        mapping = _row_mapping(row)
        store = StoreInput(
            store_id=str(mapping["store_id"]),
            vertical_id=str(mapping["vertical_id"]),
            size_index=(
                decimal_value(mapping["size_index"])
                if mapping["size_index"] is not None
                else None
            ),
            health_index=(
                decimal_value(mapping["health_index"])
                if mapping["health_index"] is not None
                else None
            ),
            footfall_index=(
                decimal_value(mapping["footfall_index"])
                if mapping["footfall_index"] is not None
                else None
            ),
            cluster=(str(mapping["cluster"]) if mapping["cluster"] is not None else None),
            channel=(str(mapping["channel"]) if mapping["channel"] is not None else None),
            source_rows=int(mapping["source_rows"] or 0),
            distinct_items=int(mapping["distinct_items"] or 0),
            ads_total=decimal_value(mapping["ads_total"]),
            forecast_7d_total=decimal_value(mapping["forecast_7d_total"]),
        )
        _require(
            store.source_rows == 100 and store.distinct_items == 100,
            f"{store.store_id} does not have exactly 100 unique source items",
        )
        _require(
            store.ads_total >= 0 and store.forecast_7d_total >= 0,
            f"{store.store_id} has negative ADS or forecast_7d aggregate",
        )
        stores.append(store)

    stores.sort(key=lambda store: store.store_id)
    _require(
        tuple(store.store_id for store in stores) == EXPECTED_STORE_IDS,
        "Runtime store population is not exactly S001 through S160",
    )

    snapshot = SourceSnapshot(
        stores=tuple(stores),
        batch_id=SOURCE_IMPORT_BATCH_ID,
        agent_name=str(batch["agent_name"]),
        workbook_version=str(batch["workbook_version"]),
        import_status=str(batch["import_status"]),
        source_revision=SOURCE_REVISION,
        source_sha256=source_sha256,
        source_as_of_date=SOURCE_AS_OF_DATE,
        batch_total_rows=int(batch["total_rows"] or 0),
        fact_rows=fact_rows,
        fact_dates=fact_dates,
        fact_min_date=fact_min_date,
        fact_max_date=fact_max_date,
        fact_distinct_stores=int(fact_row["fact_distinct_stores"] or 0),
        fact_distinct_items=int(fact_row["fact_distinct_items"] or 0),
        duplicate_groups=int(duplicate_row["duplicate_groups"] or 0),
        bad_ads=int(fact_row["bad_ads"] or 0),
        bad_forecast_7d=int(fact_row["bad_forecast_7d"] or 0),
        dim_store_rows=int(dim_store_row["dim_store_rows"] or 0),
        dim_store_distinct_ids=int(dim_store_row["dim_store_distinct_ids"] or 0),
        dim_item_rows=int(dim_item_row["dim_item_rows"] or 0),
    )
    validate_source_snapshot(snapshot)
    return snapshot


def week_start_for_offset(period_offset: int, as_of_date: date = SOURCE_AS_OF_DATE) -> date:
    """Return the Monday date for a contract period offset."""

    anchor = as_of_date - timedelta(days=as_of_date.weekday())
    return anchor + timedelta(days=7 * period_offset)


def week_end_for_offset(period_offset: int, as_of_date: date = SOURCE_AS_OF_DATE) -> date:
    return week_start_for_offset(period_offset, as_of_date) + timedelta(days=6)


def _stable_unit(label: str) -> float:
    digest = hashlib.sha256(label.encode("utf-8")).digest()
    return int.from_bytes(digest[:8], "big") / float(2**64)


def _store_rng(seed: int, store_id: str) -> random.Random:
    digest = hashlib.sha256(f"{seed}|{store_id}".encode("utf-8")).digest()
    return random.Random(int.from_bytes(digest[:8], "big"))


def _clamp(value: float, minimum: float, maximum: float) -> float:
    return max(minimum, min(maximum, value))


def _normalised_feature(value: Decimal | None) -> float:
    if value is None:
        return 1.0
    numeric = float(value)
    if not math.isfinite(numeric):
        return 1.0
    return _clamp(numeric, 0.65, 1.35)


def _safe_ratio(numerator: Decimal, denominator: Decimal) -> float:
    if denominator <= 0:
        return 1.0
    value = float(numerator / denominator)
    return value if math.isfinite(value) else 1.0


def _store_model(store: StoreInput, params: GeneratorParameters) -> dict[str, Any]:
    """Build small, explainable deterministic parameters for one store."""

    rng = _store_rng(params.seed, store.store_id)
    size = _normalised_feature(store.size_index)
    health = _normalised_feature(store.health_index)
    footfall = _normalised_feature(store.footfall_index)
    context_score = 0.45 * size + 0.25 * health + 0.30 * footfall
    vertical_unit = _stable_unit(f"vertical:{store.vertical_id}")
    group_unit = _stable_unit(
        f"group:{store.vertical_id}|{store.cluster or ''}|{store.channel or ''}"
    )

    # A small rate gives positive and negative store trajectories without
    # turning the 16-week demonstration into a trend target.
    growth_rate = _clamp(
        -0.001
        + 0.004 * (context_score - 1.0)
        + (vertical_unit - 0.5) * 0.006
        + rng.uniform(-0.006, 0.006),
        -0.012,
        0.012,
    )
    season_amplitude = 0.014 + 0.010 * rng.random()
    secondary_amplitude = 0.005 + 0.005 * rng.random()
    noise_amplitude = 0.006 + 0.004 * rng.random()
    phase = 2.0 * math.pi * rng.random() + 0.9 * group_unit
    secondary_phase = 2.0 * math.pi * rng.random() + 0.5 * vertical_unit

    # Average the two audited current scale signals.  In the observed v8.5
    # data they reconcile closely; keeping both makes the baseline explicit.
    ads_weekly = store.ads_total * params.dow_sum
    if store.forecast_7d_total > 0 and ads_weekly > 0:
        baseline = (store.forecast_7d_total + ads_weekly) / Decimal("2")
    else:
        baseline = max(store.forecast_7d_total, ads_weekly)
    if baseline <= 0:
        raise GenerationError(f"{store.store_id} has no positive source baseline")

    return {
        "rng": rng,
        "baseline": baseline,
        "growth_rate": growth_rate,
        "season_amplitude": season_amplitude,
        "secondary_amplitude": secondary_amplitude,
        "noise_amplitude": noise_amplitude,
        "phase": phase,
        "secondary_phase": secondary_phase,
    }


def _seasonal_factor(period_offset: int, model: Mapping[str, Any]) -> float:
    first = model["season_amplitude"] * math.sin(
        2.0 * math.pi * (period_offset + model["phase"]) / 13.0
    )
    second = model["secondary_amplitude"] * math.cos(
        2.0 * math.pi * (period_offset + model["secondary_phase"]) / 26.0
    )
    return 1.0 + first + second


def generate_dataset(
    snapshot: SourceSnapshot,
    params: GeneratorParameters | None = None,
) -> list[DemandRow]:
    """Generate all 5,120 rows from a preflighted source snapshot."""

    params = params or GeneratorParameters()
    if params.source_as_of_date != snapshot.source_as_of_date:
        raise GenerationError("Generator/source as-of dates do not match")
    if snapshot.batch_id != SOURCE_IMPORT_BATCH_ID:
        raise GenerationError("Generator/source import batches do not match")
    if snapshot.source_revision != SOURCE_REVISION:
        raise GenerationError("Generator/source revisions do not match")

    rows: list[DemandRow] = []
    for store in sorted(snapshot.stores, key=lambda item: item.store_id):
        model = _store_model(store, params)
        rng: random.Random = model["rng"]
        smooth_noise: dict[int, float] = {}
        previous_noise = 0.0
        for period_offset in PERIOD_OFFSETS:
            raw_noise = rng.uniform(-1.0, 1.0)
            previous_noise = 0.78 * previous_noise + 0.22 * raw_noise
            smooth_noise[period_offset] = previous_noise

        anchor = quantize_qty(store.forecast_7d_total)
        baseline = model["baseline"]
        reference_season = _seasonal_factor(1, model)
        reference_noise = smooth_noise[1]

        for period_offset in PERIOD_OFFSETS:
            week_start = week_start_for_offset(period_offset, params.source_as_of_date)
            week_end = week_start + timedelta(days=6)

            if period_offset == 1:
                # This is the only source-derived row.  No model factor is
                # applied, so the source W+1 anchor survives seed changes.
                rows.append(
                    DemandRow(
                        generation_version=params.generation_version,
                        store_id=store.store_id,
                        week_start=week_start,
                        week_end=week_end,
                        period_offset=period_offset,
                        actual_qty=None,
                        forecast_qty=anchor,
                        source_as_of_date=params.source_as_of_date,
                        forecast_as_of_date=params.source_as_of_date,
                        data_source="SOURCE_FORECAST_7D",
                        source_semantics=(
                            "ROLLING_7D_REANCHORED_TO_NEXT_FULL_WEEK"
                        ),
                        is_synthetic=False,
                        source_revision=snapshot.source_revision,
                        source_import_batch_id=snapshot.batch_id,
                        generator_version=None,
                        generator_seed=None,
                    )
                )
                continue

            growth_factor = (1.0 + model["growth_rate"]) ** (period_offset - 1)
            season_ratio = _seasonal_factor(period_offset, model) / reference_season
            noise_ratio = 1.0 + model["noise_amplitude"] * (
                smooth_noise[period_offset] - reference_noise
            )
            factor = growth_factor * season_ratio * noise_ratio

            # Historical quantities use the source-scale baseline but follow
            # their own trajectory; they are not W+1 copied backwards.
            if period_offset < 0:
                quantity = float(baseline) * factor
                actual_qty = quantize_qty(quantity)
                forecast_qty = None
                data_source = "SYNTHETIC_GENERATOR"
                semantics = "SYNTHETIC_HISTORICAL_ACTUAL"
                forecast_as_of_date = None
            else:
                quantity = float(store.forecast_7d_total) * factor
                actual_qty = None
                forecast_qty = quantize_qty(quantity)
                data_source = "SYNTHETIC_GENERATOR"
                semantics = "SYNTHETIC_FUTURE_FORECAST"
                forecast_as_of_date = params.source_as_of_date

            rows.append(
                DemandRow(
                    generation_version=params.generation_version,
                    store_id=store.store_id,
                    week_start=week_start,
                    week_end=week_end,
                    period_offset=period_offset,
                    actual_qty=actual_qty,
                    forecast_qty=forecast_qty,
                    source_as_of_date=params.source_as_of_date,
                    forecast_as_of_date=forecast_as_of_date,
                    data_source=data_source,
                    source_semantics=semantics,
                    is_synthetic=True,
                    source_revision=snapshot.source_revision,
                    source_import_batch_id=snapshot.batch_id,
                    generator_version=params.generator_version,
                    generator_seed=params.seed,
                )
            )

    rows.sort(key=lambda row: (row.store_id, row.period_offset))
    return rows


def canonical_row_text(row: DemandRow) -> str:
    """Canonical row representation used by the output fingerprint."""

    values = row.csv_dict()
    return "|".join(str(values[column]) for column in CSV_COLUMNS)


def output_fingerprint(rows: Iterable[DemandRow]) -> str:
    digest = hashlib.sha256()
    for row in sorted(rows, key=lambda item: (item.store_id, item.period_offset)):
        digest.update(canonical_row_text(row).encode("utf-8"))
        digest.update(b"\n")
    return digest.hexdigest()


def input_fingerprint(
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
) -> str:
    payload = {
        "source": snapshot.fingerprint_dict(),
        "parameters": params.manifest_dict(),
        "model": {
            "baseline": "average(source forecast_7d, sum(ads) * DOW_SUM)",
            "growth": "bounded store-specific seeded rate from store context and stable group hash",
            "seasonality": "13-week and 26-week bounded sinusoidal components",
            "noise": "seeded smoothed noise with fixed recurrence 0.78/0.22",
        },
    }
    encoded = json.dumps(payload, sort_keys=True, separators=(",", ":")).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def _rows_by_store(rows: Sequence[DemandRow]) -> dict[str, list[DemandRow]]:
    grouped: dict[str, list[DemandRow]] = {}
    for row in rows:
        grouped.setdefault(row.store_id, []).append(row)
    for store_rows in grouped.values():
        store_rows.sort(key=lambda row: row.period_offset)
    return grouped


def calculate_trend(
    rows: Sequence[DemandRow],
    store_ids: Iterable[str] | None = None,
) -> float:
    """Calculate aggregate-before-divide Demand Trend for a scope."""

    selected = set(store_ids) if store_ids is not None else None
    grouped = _rows_by_store(rows)
    if selected is None:
        selected = set(grouped)
    if not selected:
        raise GenerationError("Demand Trend scope contains no stores")

    actual_total = Decimal("0")
    forecast_total = Decimal("0")
    for store_id in selected:
        store_rows = {row.period_offset: row for row in grouped.get(store_id, [])}
        if set(store_rows) != EXPECTED_PERIOD_OFFSETS:
            raise GenerationError(f"Incomplete time series for Trend store {store_id}")
        actual_values = [
            store_rows[offset].actual_qty for offset in range(-4, 0)
        ]
        forecast_values = [
            store_rows[offset].forecast_qty for offset in range(1, 5)
        ]
        if any(value is None for value in actual_values + forecast_values):
            raise GenerationError(f"Null Trend quantity for store {store_id}")
        actual_total += sum(
            (value for value in actual_values if value is not None), Decimal("0")
        )
        forecast_total += sum(
            (value for value in forecast_values if value is not None), Decimal("0")
        )

    if actual_total <= 0:
        raise GenerationError("Demand Trend denominator is zero or negative")
    return float(forecast_total / actual_total - Decimal("1"))


def _range_summary(values: Sequence[Decimal | float]) -> dict[str, float]:
    numbers = [float(value) for value in values]
    return {
        "min": min(numbers),
        "median": statistics.median(numbers),
        "max": max(numbers),
    }


def _trend_counts(trends: Mapping[str, float]) -> dict[str, int]:
    positive = sum(value >= NEAR_ZERO_TREND_THRESHOLD for value in trends.values())
    negative = sum(value <= -NEAR_ZERO_TREND_THRESHOLD for value in trends.values())
    near_zero = len(trends) - positive - negative
    return {"positive": positive, "negative": negative, "near_zero": near_zero}


def plausibility_summary(
    rows: Sequence[DemandRow],
    store_ids: Sequence[str],
) -> dict[str, Any]:
    grouped = _rows_by_store(rows)
    historical_values = [
        row.actual_qty
        for row in rows
        if row.period_offset < 0 and row.actual_qty is not None
    ]
    future_values = [
        row.forecast_qty
        for row in rows
        if row.period_offset > 0 and row.forecast_qty is not None
    ]
    trends = {
        store_id: calculate_trend(rows, [store_id]) for store_id in store_ids
    }

    wow_changes: list[dict[str, Any]] = []
    flat_history_stores: list[str] = []
    extreme_volatility_stores: list[str] = []
    for store_id in store_ids:
        store_rows = grouped[store_id]
        by_offset = {row.period_offset: row for row in store_rows}
        historical = [
            row.actual_qty for row in store_rows if row.period_offset < 0
        ]
        historical_numbers = [float(value) for value in historical if value is not None]
        mean = statistics.mean(historical_numbers)
        stdev = statistics.pstdev(historical_numbers)
        if mean > 0 and stdev / mean < FLAT_HISTORY_COEFFICIENT_OF_VARIATION:
            flat_history_stores.append(store_id)

        store_wow: list[float] = []
        for offset in PERIOD_OFFSETS:
            if offset + 1 not in by_offset:
                continue  # The omitted W0 is not an adjacent stored period.
            current = by_offset[offset].actual_qty or by_offset[offset].forecast_qty
            following = by_offset[offset + 1].actual_qty or by_offset[offset + 1].forecast_qty
            if current is None or following is None or current <= 0:
                continue
            change = float(following / current - Decimal("1"))
            store_wow.append(abs(change))
            wow_changes.append(
                {
                    "store_id": store_id,
                    "from_period": offset,
                    "to_period": offset + 1,
                    "relative_change": change,
                    "from_qty": float(current),
                    "to_qty": float(following),
                }
            )
        if store_wow and max(store_wow) > MAX_WEEK_OVER_WEEK_ABS_CHANGE:
            extreme_volatility_stores.append(store_id)

    largest_increase = max(wow_changes, key=lambda item: item["relative_change"])
    largest_decrease = min(wow_changes, key=lambda item: item["relative_change"])
    high_trends = sorted(trends.items(), key=lambda item: item[1], reverse=True)[:5]
    low_trends = sorted(trends.items(), key=lambda item: item[1])[:5]

    result = {
        "historical_weekly_store_quantity": _range_summary(historical_values),
        "future_weekly_store_quantity": _range_summary(future_values),
        "store_trend": _range_summary(list(trends.values())),
        "trend_counts": _trend_counts(trends),
        "largest_week_over_week_increase": largest_increase,
        "largest_week_over_week_decrease": largest_decrease,
        "highest_store_trends": [
            {"store_id": store_id, "trend": trend}
            for store_id, trend in high_trends
        ],
        "lowest_store_trends": [
            {"store_id": store_id, "trend": trend}
            for store_id, trend in low_trends
        ],
        "flat_history_stores": flat_history_stores,
        "extreme_volatility_stores": extreme_volatility_stores,
        "thresholds": {
            "near_zero_trend_abs_lt": NEAR_ZERO_TREND_THRESHOLD,
            "max_store_trend_abs": MAX_STORE_TREND_ABS,
            "max_week_over_week_abs_change": MAX_WEEK_OVER_WEEK_ABS_CHANGE,
            "flat_history_cv_lt": FLAT_HISTORY_COEFFICIENT_OF_VARIATION,
        },
        "trends": trends,
    }

    if any(abs(value) > MAX_STORE_TREND_ABS for value in trends.values()):
        raise GenerationError("Plausibility failure: store Trend exceeds ±35%")
    if extreme_volatility_stores:
        raise GenerationError(
            "Plausibility failure: week-over-week change exceeds 20% for "
            + ", ".join(extreme_volatility_stores)
        )
    return result


def validate_dataset(
    rows: Sequence[DemandRow],
    snapshot: SourceSnapshot,
) -> dict[str, Any]:
    """Run all required offline data and contract validations."""

    grouped = _rows_by_store(rows)
    store_ids = tuple(sorted(snapshot.store_map))
    _require_validation(
        len(rows) == 5_120,
        f"Expected 5,120 rows, found {len(rows)}",
    )
    _require_validation(
        len(grouped) == 160 and tuple(sorted(grouped)) == EXPECTED_STORE_IDS,
        "Expected exactly 160 stores S001 through S160",
    )
    _require_validation(
        all(len(store_rows) == 32 for store_rows in grouped.values()),
        "Every store must have exactly 32 rows",
    )
    _require_validation(
        len({row.key() for row in rows}) == len(rows),
        "Duplicate generation_version + store_id + period_offset key",
    )
    _require_validation(
        len({row.week_key() for row in rows}) == len(rows),
        "Duplicate generation_version + store_id + week_start key",
    )

    for row in rows:
        _require_validation(
            row.period_offset in EXPECTED_PERIOD_OFFSETS,
            f"Invalid period offset {row.period_offset}",
        )
        _require_validation(
            row.week_start == week_start_for_offset(row.period_offset),
            f"Invalid week_start for {row.store_id} offset {row.period_offset}",
        )
        _require_validation(
            row.week_end == row.week_start + timedelta(days=6),
            f"Invalid week_end for {row.store_id} offset {row.period_offset}",
        )
        _require_validation(
            row.week_start.weekday() == 0 and row.week_end.weekday() == 6,
            f"Non-Monday/Sunday week for {row.store_id} offset {row.period_offset}",
        )
        for quantity in (row.actual_qty, row.forecast_qty):
            if quantity is not None:
                _require_validation(
                    quantity.is_finite() and quantity >= 0,
                    f"Invalid quantity in {row.store_id} offset {row.period_offset}",
                )

        if row.period_offset < 0:
            _require_validation(
                row.actual_qty is not None
                and row.forecast_qty is None
                and row.data_source == "SYNTHETIC_GENERATOR"
                and row.source_semantics == "SYNTHETIC_HISTORICAL_ACTUAL"
                and row.is_synthetic
                and row.forecast_as_of_date is None,
                f"Invalid historical provenance for {row.store_id} offset {row.period_offset}",
            )
        elif row.period_offset == 1:
            _require_validation(
                row.actual_qty is None
                and row.forecast_qty is not None
                and row.data_source == "SOURCE_FORECAST_7D"
                and row.source_semantics
                == "ROLLING_7D_REANCHORED_TO_NEXT_FULL_WEEK"
                and not row.is_synthetic
                and row.forecast_as_of_date == SOURCE_AS_OF_DATE,
                f"Invalid W+1 provenance for {row.store_id}",
            )
        else:
            _require_validation(
                row.actual_qty is None
                and row.forecast_qty is not None
                and row.data_source == "SYNTHETIC_GENERATOR"
                and row.source_semantics == "SYNTHETIC_FUTURE_FORECAST"
                and row.is_synthetic
                and row.forecast_as_of_date == SOURCE_AS_OF_DATE,
                f"Invalid future provenance for {row.store_id} offset {row.period_offset}",
            )

    expected_counts = {
        "historical_synthetic": 2_560,
        "w1_source_derived": 160,
        "later_forecast_synthetic": 2_400,
        "synthetic_total": 4_960,
        "total": 5_120,
    }
    actual_counts = {
        "historical_synthetic": sum(row.period_offset < 0 for row in rows),
        "w1_source_derived": sum(row.period_offset == 1 for row in rows),
        "later_forecast_synthetic": sum(row.period_offset > 1 for row in rows),
        "synthetic_total": sum(row.is_synthetic for row in rows),
        "total": len(rows),
    }
    _require_validation(actual_counts == expected_counts, f"Row counts differ: {actual_counts}")
    _require_validation(
        all(
            {row.period_offset for row in store_rows} == EXPECTED_PERIOD_OFFSETS
            for store_rows in grouped.values()
        ),
        "Every store must have the complete offset set without W0",
    )

    source_w1 = {
        store.store_id: store.forecast_7d_total for store in snapshot.stores
    }
    w1_rows = {
        row.store_id: row.forecast_qty
        for row in rows
        if row.period_offset == 1
    }
    differences = {
        store_id: abs(decimal_value(w1_rows[store_id]) - source_value)
        for store_id, source_value in source_w1.items()
    }
    max_difference = max(differences.values())
    _require_validation(
        max_difference <= W1_TOLERANCE,
        f"W+1 reconciliation exceeds tolerance: {max_difference}",
    )
    source_total = sum(source_w1.values(), Decimal("0"))
    generated_total = sum(
        (value for value in w1_rows.values() if value is not None), Decimal("0")
    )
    total_difference = abs(generated_total - source_total)
    _require_validation(
        total_difference <= W1_TOLERANCE,
        f"Total W+1 reconciliation exceeds tolerance: {total_difference}",
    )

    trends = {
        "S001": calculate_trend(rows, ["S001"]),
        "S002": calculate_trend(rows, ["S002"]),
        "S003": calculate_trend(rows, ["S003"]),
        "GRC": calculate_trend(
            rows,
            [
                store.store_id
                for store in snapshot.stores
                if store.vertical_id == "GRC"
            ],
        ),
        "ALL": calculate_trend(rows),
    }
    plausibility = plausibility_summary(rows, store_ids)

    return {
        "row_counts": actual_counts,
        "expected_row_counts": expected_counts,
        "period_offsets": list(sorted(EXPECTED_PERIOD_OFFSETS)),
        "calendar": {
            "w_minus_16": {
                "week_start": week_start_for_offset(-16).isoformat(),
                "week_end": week_end_for_offset(-16).isoformat(),
            },
            "w_minus_1": {
                "week_start": week_start_for_offset(-1).isoformat(),
                "week_end": week_end_for_offset(-1).isoformat(),
            },
            "w_plus_1": {
                "week_start": week_start_for_offset(1).isoformat(),
                "week_end": week_end_for_offset(1).isoformat(),
            },
            "w_plus_16": {
                "week_start": week_start_for_offset(16).isoformat(),
                "week_end": week_end_for_offset(16).isoformat(),
            },
            "w0_materialized": False,
        },
        "w1_reconciliation": {
            "max_per_store_difference": decimal_text(max_difference),
            "total_source_w1": decimal_text(source_total),
            "total_generated_w1": decimal_text(generated_total),
            "total_difference": decimal_text(total_difference),
            "s001_source": decimal_text(source_w1["S001"]),
            "s001_generated": decimal_text(w1_rows["S001"]),
            "tolerance": decimal_text(W1_TOLERANCE),
        },
        "trend": trends,
        "plausibility": plausibility,
    }


def _require_validation(condition: bool, message: str) -> None:
    if not condition:
        raise GenerationError(message)


def write_csv(rows: Sequence[DemandRow], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for row in sorted(rows, key=lambda item: (item.store_id, item.period_offset)):
            writer.writerow(row.csv_dict())


def _excel_value(column: str, row: DemandRow) -> Any:
    values = row.csv_dict()
    if column in {"week_start", "week_end", "source_as_of_date", "forecast_as_of_date"}:
        raw = values[column]
        return date.fromisoformat(str(raw)) if raw else None
    if column in {"actual_qty", "forecast_qty"}:
        raw = values[column]
        return float(raw) if raw else None
    return values[column]


def write_xlsx(
    rows: Sequence[DemandRow],
    validation: Mapping[str, Any],
    path: Path,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Demand Store Week"
    sheet.append(list(CSV_COLUMNS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in sorted(rows, key=lambda item: (item.store_id, item.period_offset)):
        sheet.append([_excel_value(column, row) for column in CSV_COLUMNS])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(CSV_COLUMNS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(14, len(column) + 2)

    trend_sheet = workbook.create_sheet("Trend Summary")
    trend_sheet.append(["Scope", "Demand Trend"])
    for cell in trend_sheet[1]:
        cell.font = Font(bold=True)
    for scope, value in validation["trend"].items():
        trend_sheet.append([scope, value])

    validation_sheet = workbook.create_sheet("Validation Summary")
    validation_sheet.append(["Validation", "Result"])
    for cell in validation_sheet[1]:
        cell.font = Font(bold=True)
    summary_rows = [
        ("total_rows", validation["row_counts"]["total"]),
        ("synthetic_rows", validation["row_counts"]["synthetic_total"]),
        ("source_derived_rows", validation["row_counts"]["w1_source_derived"]),
        ("max_w1_difference", validation["w1_reconciliation"]["max_per_store_difference"]),
        ("calendar_w0_materialized", validation["calendar"]["w0_materialized"]),
        ("status", "PASS"),
    ]
    for key, value in summary_rows:
        validation_sheet.append([key, value])
    validation_sheet.freeze_panes = "A2"

    workbook.save(path)


def _normalise_export_value(column: str, value: Any) -> str:
    if value is None or value == "":
        return ""
    if column in {"week_start", "week_end", "source_as_of_date", "forecast_as_of_date"}:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value)[:10]
    if column in {"actual_qty", "forecast_qty"}:
        return decimal_or_blank(quantize_qty(value))
    if column == "is_synthetic":
        return str(int(value))
    return str(value)


def validate_export_parity(csv_path: Path, xlsx_path: Path) -> dict[str, Any]:
    """Verify that the review workbook carries the CSV's logical rows."""

    from openpyxl import load_workbook

    with csv_path.open("r", encoding="utf-8", newline="") as handle:
        csv_rows = list(csv.DictReader(handle))
    _require_validation(
        len(csv_rows) == 5_120,
        f"CSV export has {len(csv_rows)} rows instead of 5,120",
    )
    csv_canonical = [
        tuple(_normalise_export_value(column, row[column]) for column in CSV_COLUMNS)
        for row in csv_rows
    ]

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    _require_validation(
        "Demand Store Week" in workbook.sheetnames,
        "XLSX is missing the canonical Demand Store Week sheet",
    )
    sheet = workbook["Demand Store Week"]
    values = sheet.iter_rows(values_only=True)
    header = tuple(next(values))
    _require_validation(
        header == CSV_COLUMNS,
        "XLSX canonical sheet columns differ from the CSV columns",
    )
    xlsx_canonical = [
        tuple(
            _normalise_export_value(column, value)
            for column, value in zip(CSV_COLUMNS, row)
        )
        for row in values
    ]
    workbook.close()
    _require_validation(
        len(xlsx_canonical) == 5_120,
        f"XLSX canonical sheet has {len(xlsx_canonical)} rows instead of 5,120",
    )
    _require_validation(
        csv_canonical == xlsx_canonical,
        "CSV and XLSX canonical rows differ",
    )
    return {
        "csv_rows": len(csv_canonical),
        "xlsx_rows": len(xlsx_canonical),
        "same_logical_rows": True,
    }


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return decimal_text(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Mapping):
        return {str(key): _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def write_manifest(manifest: Mapping[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_json_safe(dict(manifest)), indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def _fmt(value: Any, digits: int = 6) -> str:
    if isinstance(value, float):
        return f"{value:.{digits}f}"
    if isinstance(value, Decimal):
        return decimal_text(value)
    return str(value)


def _markdown_map(mapping: Mapping[str, Any]) -> str:
    return "\n".join(
        f"- **{key}:** `{_fmt(value)}`" for key, value in mapping.items()
    )


def render_generation_report(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    csv_path: Path,
    xlsx_path: Path,
    manifest_path: Path,
    test_result: str,
    reproducibility: Mapping[str, Any],
) -> str:
    counts = validation["row_counts"]
    w1 = validation["w1_reconciliation"]
    trend = validation["trend"]
    plausibility = validation["plausibility"]
    preflight = {
        "batch": snapshot.batch_id,
        "workbook_version": snapshot.workbook_version,
        "import_status": snapshot.import_status,
        "source_sha256": snapshot.source_sha256,
        "source_as_of_date": snapshot.source_as_of_date,
        "fact_rows": snapshot.fact_rows,
        "fact_distinct_stores": snapshot.fact_distinct_stores,
        "fact_distinct_items": snapshot.fact_distinct_items,
        "fact_dates": snapshot.fact_dates,
        "duplicate_groups": snapshot.duplicate_groups,
        "bad_ads": snapshot.bad_ads,
        "bad_forecast_7d": snapshot.bad_forecast_7d,
        "dim_store_rows": snapshot.dim_store_rows,
        "dim_item_rows": snapshot.dim_item_rows,
    }

    highest = ", ".join(
        f"{item['store_id']} ({item['trend']:.2%})"
        for item in plausibility["highest_store_trends"]
    )
    lowest = ", ".join(
        f"{item['store_id']} ({item['trend']:.2%})"
        for item in plausibility["lowest_store_trends"]
    )
    largest_increase = plausibility["largest_week_over_week_increase"]
    largest_decrease = plausibility["largest_week_over_week_decrease"]

    lines = [
        "# Demand Store-Week Generation Report",
        "",
        "## Generation metadata",
        "",
        "| Field | Value |",
        "|---|---|",
        f"| Generation version | `{params.generation_version}` |",
        f"| Generator version | `{params.generator_version}` |",
        f"| Seed | `{params.seed}` |",
        f"| Source revision | `{snapshot.source_revision}` |",
        f"| Source import batch | `{snapshot.batch_id}` |",
        f"| Source as-of date | `{snapshot.source_as_of_date.isoformat()}` |",
        f"| Input fingerprint | `{input_fp}` |",
        f"| Output fingerprint | `{output_fp}` |",
        f"| Total rows | `{counts['total']}` |",
        f"| Synthetic rows | `{counts['synthetic_total']}` |",
        f"| Source-derived rows | `{counts['w1_source_derived']}` |",
        "| Overall verdict | **READY FOR SQL LOAD WITH CAVEATS** |",
        "",
        "The candidate passed source preflight and all offline generation/validation checks. It remains caveated because W+1 is a rolling seven-day quantity re-anchored to the first full calendar week, and historical rows are synthetic without genuine sales ground truth. No Azure SQL object or row was created or modified.",
        "",
        "## 1. Source Preflight Result",
        "",
        "The generator executed SELECT-only Azure SQL queries against the approved v8.5 source family. No INSERT, UPDATE, DELETE, CREATE, ALTER, DROP, migration, or transaction write was used.",
        "",
        "All mandatory checks passed:",
        "",
        "- batch 23 is `COMPLETED`, `retail_facts_seed`, and `v8.5`.",
        f"- the batch metadata SHA matches `{snapshot.source_sha256}`;",
        f"- `retail.fact_inventory_daily` has `{snapshot.fact_rows}` rows across `{snapshot.fact_dates}` date;",
        f"- the fact contains `{snapshot.fact_distinct_stores}` stores and `{snapshot.fact_distinct_items}` items;",
        f"- `retail.dim_store` has `{snapshot.dim_store_rows}` rows and `retail.dim_item` has `{snapshot.dim_item_rows}` rows;",
        f"- duplicate `(store_key, item_key, cal_date)` groups: `{snapshot.duplicate_groups}`;",
        f"- invalid/null/negative ADS rows: `{snapshot.bad_ads}`; invalid/null/negative `forecast_7d` rows: `{snapshot.bad_forecast_7d}`;",
        "- all fact rows are tied to import batch 23 and the snapshot date is 2026-07-01.",
        "",
        "The preflight did not query or use v8.2 `StoreSkuSnapshot`, `Sku`, or workbook measures.",
        "",
        "## 2. Generator Design",
        "",
        "The generator materialises one row per store and period offset for offsets -16…-1 and +1…+16. It derives a store baseline as the average of the source store `forecast_7d` aggregate and `SUM(ads) × 7.45`. Historical quantities use that source-scale baseline with a bounded store-specific growth rate, two low-amplitude smooth seasonal waves, and seeded smoothed noise. Future quantities use the same shape anchored to the exact source W+1 value.",
        "",
        "The source-derived W+1 branch exits before any model factor is applied. W+1 is copied/quantized from the v8.5 source aggregate only; it is not smoothed, trended, seasonally adjusted, or randomly varied.",
        "",
        "The model is intentionally small and transparent. It does not use workbook Trend values, `agent_kpi_reference`, frontend Trend values, forecast accuracy constants, or SQL-side randomness.",
        "",
        "## 3. Inputs Used",
        "",
        "| Input | Source | Use |",
        "|---|---|---|",
        "| Store ID | `retail.dim_store.store_id` | Runtime population and output grain. |",
        "| Vertical | `retail.dim_store.vertical_id` | Stable store-specific trajectory hash and GRC scope. |",
        "| Store context | `size_index`, `health_index`, `footfall_index`, `cluster`, `channel` | Bounded growth/variation differences. |",
        "| ADS | `retail.fact_inventory_daily.ads` | Store weekly baseline component via DOW sum 7.45. |",
        "| Forecast 7d | `retail.fact_inventory_daily.forecast_7d` | Exact W+1 source anchor and baseline component. |",
        "| Calendar | Contract date arithmetic | Monday-Sunday period mapping. |",
        "",
        "`retail.dim_item` was preflighted at 800 rows but no SKU-level allocation was introduced because the output contract is store-week grain. No v8.2 source was used.",
        "",
        "## 4. Mathematical Generation Approach",
        "",
        "For each store, the generator computes:",
        "",
        "```text",
        "ads_weekly = SUM(v8.5 ADS) × 7.45",
        "baseline = (ads_weekly + source W+1) / 2",
        "growth_rate = bounded seeded store/context rate in [-1.2%, +1.2%] per week",
        "seasonality = two bounded sinusoidal components with 13- and 26-week periods",
        "noise = seeded smoothed recurrence: 0.78 × prior + 0.22 × new draw",
        "historical actual = baseline × growth × relative seasonality × relative noise",
        "future forecast = exact W+1 × growth × relative seasonality × relative noise",
        "```",
        "",
        "Final values are quantized once to six decimal places with `ROUND_HALF_UP`. The source W+1 row is quantized only for the export representation and is reconciled to the unrounded source aggregate within 0.0001 units.",
        "",
        "## 5. Seed / Version / Reproducibility",
        "",
        f"- generation version: `{params.generation_version}`",
        f"- generator version: `{params.generator_version}`",
        f"- fixed seed: `{params.seed}`",
        f"- input fingerprint: `{input_fp}`",
        f"- output fingerprint: `{output_fp}`",
        "",
        f"Identical rerun fingerprint: `{reproducibility['same_fingerprint']}`; identical six-decimal rows: `{reproducibility['same_rows']}`; changed-seed synthetic rows differ: `{reproducibility['changed_seed_synthetic_differs']}`; changed-seed W+1 rows remain identical: `{reproducibility['changed_seed_w1_same']}`.",
        "",
        "The output fingerprint is a SHA-256 digest of sorted canonical rows after final quantization. Volatile generation timestamps are excluded.",
        "",
        "## 6. Output Files",
        "",
        f"- CSV: `{csv_path}`",
        f"- XLSX review copy: `{xlsx_path}`",
        f"- Manifest: `{manifest_path}`",
        "",
        "The XLSX `Demand Store Week` sheet has the same 5,120 logical rows and canonical columns as the CSV. `Trend Summary` and `Validation Summary` are review-only sheets.",
        "",
        "## 7. Row Counts",
        "",
        "| Data block | Rows | Provenance |",
        "|---|---:|---|",
        f"| W-16…W-1 historical | `{counts['historical_synthetic']}` | Synthetic historical actual |",
        f"| W+1 | `{counts['w1_source_derived']}` | v8.5 source-derived |",
        f"| W+2…W+16 | `{counts['later_forecast_synthetic']}` | Synthetic future forecast |",
        f"| **Total** | **`{counts['total']}`** | Mixed provenance |",
        "",
        "## 8. W+1 Reconciliation",
        "",
        f"- maximum per-store difference: `{w1['max_per_store_difference']}` units.",
        f"- total source W+1: `{w1['total_source_w1']}`.",
        f"- total generated W+1: `{w1['total_generated_w1']}`.",
        f"- total difference: `{w1['total_difference']}`; tolerance `{w1['tolerance']}`.",
        f"- S001 source: `{w1['s001_source']}`; S001 generated: `{w1['s001_generated']}`.",
        "",
        "All 160 store anchors passed the 0.0001-unit tolerance. W+1 remained identical in the changed-seed negative control.",
        "",
        "## 9. Calendar Validation",
        "",
        "The generated calendar uses Asia/Jakarta business semantics, Monday-Sunday weeks, and source as-of 2026-07-01. W0 is not materialized.",
        "",
        "| Period | Week start | Week end | Result |",
        "|---|---|---|---|",
        f"| W-16 | `{validation['calendar']['w_minus_16']['week_start']}` | `{validation['calendar']['w_minus_16']['week_end']}` | PASS |",
        f"| W-1 | `{validation['calendar']['w_minus_1']['week_start']}` | `{validation['calendar']['w_minus_1']['week_end']}` | PASS |",
        f"| W+1 | `{validation['calendar']['w_plus_1']['week_start']}` | `{validation['calendar']['w_plus_1']['week_end']}` | PASS |",
        f"| W+16 | `{validation['calendar']['w_plus_16']['week_start']}` | `{validation['calendar']['w_plus_16']['week_end']}` | PASS |",
        "",
        "## 10. Demand Trend Results",
        "",
        "Trend is calculated after generation as aggregate W+1…W+4 forecast divided by aggregate W-4…W-1 actual, minus one. Individual store Trend percentages are not averaged.",
        "",
        "| Scope | Demand Trend |",
        "|---|---:|",
        *[f"| {scope} | `{value:.6%}` |" for scope, value in trend.items()],
        "",
        "GRC uses the stores whose `retail.dim_store.vertical_id = 'GRC'`. No workbook/reference Trend is an acceptance target.",
        "",
        "## 11. Plausibility Summary",
        "",
        f"- historical weekly store quantity min/median/max: `{plausibility['historical_weekly_store_quantity']['min']:.2f}` / `{plausibility['historical_weekly_store_quantity']['median']:.2f}` / `{plausibility['historical_weekly_store_quantity']['max']:.2f}`.",
        f"- future weekly store quantity min/median/max: `{plausibility['future_weekly_store_quantity']['min']:.2f}` / `{plausibility['future_weekly_store_quantity']['median']:.2f}` / `{plausibility['future_weekly_store_quantity']['max']:.2f}`.",
        f"- store Trend min/median/max: `{plausibility['store_trend']['min']:.2%}` / `{plausibility['store_trend']['median']:.2%}` / `{plausibility['store_trend']['max']:.2%}`.",
        f"- GRC Trend: `{trend['GRC']:.2%}`; all-store Trend: `{trend['ALL']:.2%}`.",
        f"- Trend counts (positive ≥1%, negative ≤-1%, near-zero otherwise): `{plausibility['trend_counts']}`;",
        f"- largest week-over-week increase: `{largest_increase['store_id']}` W{largest_increase['from_period']:+d}→W{largest_increase['to_period']:+d}, `{largest_increase['relative_change']:.2%}`;",
        f"- largest week-over-week decrease: `{largest_decrease['store_id']}` W{largest_decrease['from_period']:+d}→W{largest_decrease['to_period']:+d}, `{largest_decrease['relative_change']:.2%}`;",
        f"- five highest store Trends: {highest};",
        f"- five lowest store Trends: {lowest};",
        f"- suspiciously flat historical stores (CV < {FLAT_HISTORY_COEFFICIENT_OF_VARIATION:.3f}): `{plausibility['flat_history_stores']}`;",
        f"- unusually volatile stores (> {MAX_WEEK_OVER_WEEK_ABS_CHANGE:.0%} stored-period movement): `{plausibility['extreme_volatility_stores']}`.",
        "",
        "The defined plausibility thresholds are ±35% store Trend, 20% maximum adjacent stored-period movement, and coefficient of variation below 0.001 for a flat-history warning. The candidate did not breach a fail threshold.",
        "",
        "## 12. Validation Results",
        "",
        "| Validation | Result |",
        "|---|---|",
        "| Population | PASS |",
        "| Provenance counts | PASS |",
        "| Period set / no W0 | PASS |",
        "| Monday-Sunday dates | PASS |",
        "| Non-negative finite quantities | PASS |",
        "| Null/provenance shape | PASS |",
        "| Duplicate keys | PASS |",
        "| Per-store and total W+1 reconciliation | PASS |",
        "| Trend formula and aggregate-before-divide | PASS |",
        "| Reproducibility / changed-seed control | PASS |",
        "| CSV/XLSX canonical row parity | PASS |",
        "| Plausibility thresholds | PASS |",
        "",
        "## 13. Test Results",
        "",
        f"{test_result}",
        "",
        "## 14. Known Limitations",
        "",
        "- Historical quantities are synthetic and cannot be validated against genuine store sales, returns, cancellations, or stockout-censored demand.",
        "- W+1 is a rolling/modelled seven-day quantity re-anchored to 2026-07-06; it is not a target-dated calendar-week forecast.",
        "- The output is store-week grain and cannot truthfully answer category- or SKU-filtered Trend questions.",
        "- The current application has separate chain/fixture/frontend curve paths; this task did not wire or replace them.",
        "- The current sales-unit aggregation semantics are preserved without pack-factor or buying-unit conversion.",
        "",
        "## 15. SQL Load Readiness",
        "",
        "The candidate is **READY FOR SQL LOAD WITH CAVEATS** after human review of the exported artifacts. Task 6 may use the Task 2 proposed `synthetic.demand_generation`, `synthetic.demand_store_week`, and `retail.v_demand_store_week` designs, but this task did not create them and did not insert any rows. SQL load must reverify source lineage, preserve the manifest fingerprints, and validate the loaded rows before approval/current cutover.",
        "",
        "## 16. Next-Step Handoff",
        "",
        "Task 4 review may inspect the CSV/XLSX and this report for realism. Task 5 may review volatility, distribution, seasonality, spikes, and continuity without targeting any workbook Trend constant. Task 6 may create the additive SQL layer and load only the reviewed candidate; no backend/frontend wiring is authorized by this task.",
        "",
    ]
    return "\n".join(lines)


def write_generation_report(content: str, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(content, encoding="utf-8")


def build_manifest(
    *,
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    validation: Mapping[str, Any],
    input_fp: str,
    output_fp: str,
    generated_at: str,
    reproducibility: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "generation_version": params.generation_version,
        "generator_version": params.generator_version,
        "seed": params.seed,
        "source_revision": snapshot.source_revision,
        "source_sha256": snapshot.source_sha256,
        "source_import_batch_id": snapshot.batch_id,
        "source_as_of_date": snapshot.source_as_of_date.isoformat(),
        "business_timezone": BUSINESS_TIMEZONE,
        "input_fingerprint": input_fp,
        "output_fingerprint": output_fp,
        "generated_at": generated_at,
        "generator_parameters": params.manifest_dict(),
        "source_preflight": {
            "agent_name": snapshot.agent_name,
            "workbook_version": snapshot.workbook_version,
            "import_status": snapshot.import_status,
            "batch_total_rows": snapshot.batch_total_rows,
            "fact_rows": snapshot.fact_rows,
            "fact_distinct_stores": snapshot.fact_distinct_stores,
            "fact_distinct_items": snapshot.fact_distinct_items,
            "fact_dates": snapshot.fact_dates,
            "fact_min_date": snapshot.fact_min_date.isoformat(),
            "fact_max_date": snapshot.fact_max_date.isoformat(),
            "duplicate_groups": snapshot.duplicate_groups,
            "bad_ads": snapshot.bad_ads,
            "bad_forecast_7d": snapshot.bad_forecast_7d,
            "dim_store_rows": snapshot.dim_store_rows,
            "dim_item_rows": snapshot.dim_item_rows,
        },
        "row_counts": validation["row_counts"],
        "export_parity": validation["export_parity"],
        "w1_reconciliation": validation["w1_reconciliation"],
        "trend": validation["trend"],
        "plausibility": validation["plausibility"],
        "reproducibility": dict(reproducibility),
        "sql_changes_performed": False,
    }


def _reproducibility_check(
    snapshot: SourceSnapshot,
    params: GeneratorParameters,
    rows: Sequence[DemandRow],
    fingerprint: str,
) -> dict[str, Any]:
    rerun = generate_dataset(snapshot, params)
    rerun_fingerprint = output_fingerprint(rerun)
    same_rows = [canonical_row_text(row) for row in rows] == [
        canonical_row_text(row) for row in rerun
    ]
    negative_params = replace(params, seed=params.seed + 1)
    negative_rows = generate_dataset(snapshot, negative_params)
    synthetic_original = {
        (row.store_id, row.period_offset): row.actual_qty or row.forecast_qty
        for row in rows
        if row.is_synthetic
    }
    synthetic_negative = {
        (row.store_id, row.period_offset): row.actual_qty or row.forecast_qty
        for row in negative_rows
        if row.is_synthetic
    }
    w1_original = {
        row.store_id: row.forecast_qty for row in rows if row.period_offset == 1
    }
    w1_negative = {
        row.store_id: row.forecast_qty
        for row in negative_rows
        if row.period_offset == 1
    }
    return {
        "same_fingerprint": fingerprint == rerun_fingerprint,
        "same_rows": same_rows,
        "changed_seed_synthetic_differs": synthetic_original != synthetic_negative,
        "changed_seed_w1_same": w1_original == w1_negative,
        "rerun_output_fingerprint": rerun_fingerprint,
        "negative_control_seed": negative_params.seed,
        "negative_control_output_fingerprint": output_fingerprint(negative_rows),
    }


def run_generation(
    *,
    output_dir: Path,
    params: GeneratorParameters,
    test_result: str,
) -> dict[str, Any]:
    snapshot = load_source_snapshot()
    rows = generate_dataset(snapshot, params)
    validation = validate_dataset(rows, snapshot)
    output_fp = output_fingerprint(rows)
    input_fp = input_fingerprint(snapshot, params)
    reproducibility = _reproducibility_check(snapshot, params, rows, output_fp)
    _require_validation(
        reproducibility["same_fingerprint"]
        and reproducibility["same_rows"]
        and reproducibility["changed_seed_synthetic_differs"]
        and reproducibility["changed_seed_w1_same"],
        "Reproducibility validation failed",
    )

    csv_path = output_dir / "demand_store_week_poc_v1.csv"
    xlsx_path = output_dir / "demand_store_week_poc_v1.xlsx"
    manifest_path = output_dir / "demand_store_week_poc_v1_manifest.json"
    report_path = REPO_ROOT / "plans" / "demand-store-week-generation-report.md"
    write_csv(rows, csv_path)
    write_xlsx(rows, validation, xlsx_path)
    validation["export_parity"] = validate_export_parity(csv_path, xlsx_path)
    generated_at = datetime.now(timezone.utc).isoformat()
    manifest = build_manifest(
        snapshot=snapshot,
        params=params,
        validation=validation,
        input_fp=input_fp,
        output_fp=output_fp,
        generated_at=generated_at,
        reproducibility=reproducibility,
    )
    write_manifest(manifest, manifest_path)
    write_generation_report(
        render_generation_report(
            snapshot=snapshot,
            params=params,
            validation=validation,
            input_fp=input_fp,
            output_fp=output_fp,
            csv_path=csv_path,
            xlsx_path=xlsx_path,
            manifest_path=manifest_path,
            test_result=test_result,
            reproducibility=reproducibility,
        ),
        report_path,
    )
    return {
        "snapshot": snapshot,
        "rows": rows,
        "validation": validation,
        "input_fingerprint": input_fp,
        "output_fingerprint": output_fp,
        "reproducibility": reproducibility,
        "csv_path": csv_path,
        "xlsx_path": xlsx_path,
        "manifest_path": manifest_path,
        "report_path": report_path,
    }


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "artifacts",
        help="Local output directory (default: artifacts)",
    )
    parser.add_argument(
        "--seed",
        type=int,
        default=FIXED_SEED,
        help="Deterministic synthetic seed (default: documented fixed seed)",
    )
    parser.add_argument(
        "--generation-version",
        default=GENERATION_VERSION,
    )
    parser.add_argument(
        "--generator-version",
        default=GENERATOR_VERSION,
    )
    parser.add_argument(
        "--test-result",
        default="Focused generator tests: not run by generator CLI.",
        help="Text recorded in the generation report's Test Results section",
    )
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    params = GeneratorParameters(
        generation_version=args.generation_version,
        generator_version=args.generator_version,
        seed=args.seed,
    )
    try:
        result = run_generation(
            output_dir=args.output_dir,
            params=params,
            test_result=args.test_result,
        )
    except SourcePreflightError as exc:
        print(f"SOURCE PREFLIGHT FAILED: {exc}", file=sys.stderr)
        return 2
    except GenerationError as exc:
        print(f"GENERATION/VALIDATION FAILED: {exc}", file=sys.stderr)
        return 3
    print(
        json.dumps(
            {
                "report_path": str(result["report_path"]),
                "csv_path": str(result["csv_path"]),
                "xlsx_path": str(result["xlsx_path"]),
                "manifest_path": str(result["manifest_path"]),
                "output_fingerprint": result["output_fingerprint"],
                "row_counts": result["validation"]["row_counts"],
                "trend": result["validation"]["trend"],
                "w1_reconciliation": result["validation"]["w1_reconciliation"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
