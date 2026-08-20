"""Generate `resources/inbound_store_sku_16w_v1.csv` -- a 16-week arrival schedule.

Run it yourself:

    ./.venv/Scripts/python.exe scripts/generate_synthetic_inbound_16w.py

Input:  resources/demand_store_sku_32w_poc_v1.csv   (forecast_w1..forecast_w16)
        resources/dbtemp/schema_with_data.json      (lead_d, pack_factor)
Output: resources/inbound_store_sku_16w_v1.csv
        resources/inbound_store_sku_16w_v1_manifest.json
        resources/inbound_store_sku_16w_v1.xlsx

The .xlsx is the same rows as the CSV, for importing through the SSMS wizard
instead of running the seeding script. It is written here rather than by a
separate step so it cannot fall out of step with the CSV beside it -- a stale
workbook silently loaded into the live table would be worse than no workbook.

WHY THIS TABLE EXISTS
---------------------
No table in this warehouse records WHEN an inbound order arrives. The workbook
stores how much is on order per SKU (`open_po_qty`) and never a date, which is
stated outright in `REQUIREMENT_NOTE` and in A3's own chart comment. The
requirement chart therefore had to place every open PO on its SKU's lead day --
2, 4 or 7 -- and since the chart steps weekly, every route lands inside the
first week and cover is a flat line from W+1 onward. Cumulative demand rising
against a flat stock diverges without bound, which is why that board always
read "Cover runs out at W+1" no matter what was in scope.

This file invents the missing arrival calendar, and says so. It is synthetic
and labelled synthetic everywhere it surfaces.

HOW THE SCHEDULE IS BUILT
-------------------------
Demand-anchored: what arrives over the horizon is what the horizon is forecast
to sell, so arrivals oscillate ABOUT the demand line instead of drifting away
from it. That is the whole point -- an inbound stream that under-delivers by
construction would reproduce the runaway gap this replaces.

Batched, not trickled. Each route receives on its own cadence:

    direct (lead 2d)  every week      -- fresh, store-direct
    flow   (lead 4d)  every week      -- DC pick and pass, fast movers
    cross  (lead 7d)  every 2 weeks   -- DC consolidation across vendors

Cadence is a modelling choice: no workbook cell states a delivery interval, so
any set of intervals here is invented. Lead time bounds the fastest possible
reorder and nothing bounds the slowest. These were chosen so that arrivals per
week stay near demand per week -- close enough that the chart reads as supply
against demand rather than as two unrelated magnitudes sharing an axis. An
earlier 1/2/3 set put 12,500 of the 16,000 rows on the same fortnightly beat,
which made chain arrivals swing 13x between alternating weeks.

A delivery covers demand from its own week until the next one, rounded to
whole cases (`pack_factor`) because a purchase order buys cases, not units.

THE PHASE IS SHARED WITHIN A ROUTE, AND THAT IS DELIBERATE
----------------------------------------------------------
Every SKU on a route receives in the same weeks. Staggering the phase per SKU
would be equally defensible in isolation, but 16,000 independently-phased rows
average each other out completely: chain arrivals would come back flat and the
line would sit on top of demand saying nothing a single number could not.
Synchronising by route is also the more realistic of the two -- a DC
consolidation run ships on a fixed calendar, it does not re-randomise per item.
Here it is the cross route alone that batches, which is what puts the visible
rise and fall into an otherwise weekly stream.

There is no random component anywhere in this generator. The schedule is a
function of the demand curve and the route, so a rerun reproduces the file
byte for byte and the manifest records no seed.
"""

from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

REPO = Path(__file__).resolve().parents[1]

DEMAND_CSV = REPO / "resources" / "demand_store_sku_32w_poc_v1.csv"
WORKBOOK = REPO / "resources" / "dbtemp" / "schema_with_data.json"
TARGET_CSV = REPO / "resources" / "inbound_store_sku_16w_v1.csv"
TARGET_MANIFEST = REPO / "resources" / "inbound_store_sku_16w_v1_manifest.json"
TARGET_XLSX = REPO / "resources" / "inbound_store_sku_16w_v1.xlsx"

GENERATION_NAME = "inbound_store_sku_16w_v1"
TABLE_NAME = "inbound_store_sku_16w"
GENERATOR_VERSION = "inbound-store-sku-16w-generator-v1.0.0"

WEEKS = 16
FORECAST_COLUMNS = [f"forecast_w{n}" for n in range(1, WEEKS + 1)]
ARRIVAL_COLUMNS = [f"arrival_w{n}" for n in range(1, WEEKS + 1)]
COLUMNS = ["sku_id", "store_id", "route"] + ARRIVAL_COLUMNS

# Route selection is `route_for` from the dashboards, restated: the first route
# whose lead time the SKU does not exceed. Cadence is this file's own addition.
ROUTE_CADENCE: tuple[tuple[str, int, int], ...] = (
    # (route id, max lead days, weeks between deliveries)
    ("direct", 2, 1),
    ("flow", 4, 1),
    ("cross", 7, 2),
)

# The three gates this generator refuses to write a file without.
RATIO_BOUNDS = (0.98, 1.02)
MIN_COVER_AMPLITUDE = 0.15
MAX_DEMAND_SPREAD = 0.35


def route_and_cadence(lead_days: float) -> tuple[str, int]:
    for route, max_lead, cadence in ROUTE_CADENCE:
        if lead_days <= max_lead:
            return route, cadence
    route, _, cadence = ROUTE_CADENCE[-1]
    return route, cadence


def delivery_weeks(cadence: int) -> list[int]:
    """Weeks 1..16 this cadence receives in, counting from the first one.

    Shared by every SKU on the route -- see the module docstring on why the
    phase is not staggered.
    """
    return list(range(1, WEEKS + 1, cadence))


def load_demand() -> list[dict[str, Any]]:
    with DEMAND_CSV.open(newline="", encoding="utf-8-sig") as handle:
        return [
            {
                "sku_id": row["sku_id"],
                "store_id": row["store_id"],
                "forecast": [float(row[column]) for column in FORECAST_COLUMNS],
            }
            for row in csv.DictReader(handle)
        ]


def load_workbook() -> tuple[dict[str, dict[str, Any]], float]:
    payload = json.loads(WORKBOOK.read_text(encoding="utf-8"))
    tables: dict[str, list[dict[str, Any]]] = {}
    for table in payload["tables"]:
        names = [column["name"] for column in table["columns"]]
        tables[table["name"]] = [dict(zip(names, row)) for row in table["rows"]]

    sku_master = {row["sku_id"]: row for row in tables["sku_master"]}
    # Chain on-hand, the same figure `build_lines` puts on each line. Only the
    # running-position report below uses it; nothing is written from it.
    on_hand = sum(float(row["qty_on_hand"]) for row in tables["replenishment_detail"])
    return sku_master, on_hand


def build_rows(
    demand: list[dict[str, Any]], sku_master: dict[str, dict[str, Any]]
) -> list[dict[str, Any]]:
    rows = []
    for record in demand:
        sku = sku_master[record["sku_id"]]
        route, cadence = route_and_cadence(float(sku["lead_d"]))
        pack = float(sku["pack_factor"]) or 1.0
        weeks = delivery_weeks(cadence)
        forecast = record["forecast"]

        arrivals = [0.0] * WEEKS
        for position, week in enumerate(weeks):
            # This delivery carries the shelf to the next one.
            until = weeks[position + 1] if position + 1 < len(weeks) else WEEKS + 1
            need = sum(forecast[week - 1 : until - 1])
            # Whole cases. Nearest rather than always-up: rounding every one of
            # ~110,000 deliveries up would add half a case each and push total
            # inbound past the demand it is anchored to, which is the one thing
            # this schedule must not do.
            arrivals[week - 1] = round(need / pack) * pack

        rows.append(
            {
                "sku_id": record["sku_id"],
                "store_id": record["store_id"],
                "route": route,
                **{
                    column: f"{value:.6f}"
                    for column, value in zip(ARRIVAL_COLUMNS, arrivals)
                },
            }
        )
    return rows


def weekly_totals(
    rows: list[dict[str, Any]], demand: list[dict[str, Any]]
) -> tuple[list[float], list[float]]:
    """Chain arrivals per week and chain demand per week."""
    weekly_demand = [0.0] * WEEKS
    for record in demand:
        for index, value in enumerate(record["forecast"]):
            weekly_demand[index] += value

    weekly_arrivals = [0.0] * WEEKS
    for row in rows:
        for index, column in enumerate(ARRIVAL_COLUMNS):
            weekly_arrivals[index] += float(row[column])

    return weekly_arrivals, weekly_demand


def running_position(
    weekly_arrivals: list[float], weekly_demand: list[float], on_hand: float
) -> list[float]:
    """Stock left at the end of each week: opening + arrivals - demand.

    Not plotted. The chart draws arrivals against demand -- two flows, one
    scale -- and this is the stock underneath them. It is what actually answers
    "does the chain run out": a week where arrivals fall short of demand is
    ordinary and the shelf absorbs it; only an empty shelf is a shortfall.
    """
    position = []
    opening = on_hand
    for index in range(WEEKS):
        opening = opening + weekly_arrivals[index] - weekly_demand[index]
        position.append(opening)
        opening = max(0.0, opening)
    return position


def fingerprint(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_xlsx(rows: list[dict[str, Any]]) -> None:
    """The same rows as a workbook, for the SSMS import wizard.

    Arrival columns are written as numbers, not as the formatted strings the
    CSV carries. The wizard types each column from the cell values it sees, and
    a numeric column arriving as text lands in NVARCHAR and then fails the
    insert into DECIMAL(20,6).
    """
    from openpyxl import Workbook

    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(TABLE_NAME)
    sheet.append(COLUMNS)
    numeric = [name not in ("sku_id", "store_id", "route") for name in COLUMNS]
    for row in rows:
        sheet.append(
            [
                float(row[name]) if is_number else row[name]
                for name, is_number in zip(COLUMNS, numeric)
            ]
        )
    workbook.save(TARGET_XLSX)


def main() -> int:
    for path in (DEMAND_CSV, WORKBOOK):
        if not path.exists():
            print(f"FAIL  source not found: {path}")
            return 1

    demand = load_demand()
    sku_master, on_hand = load_workbook()

    unknown = sorted({r["sku_id"] for r in demand} - set(sku_master))[:5]
    if unknown:
        print(f"FAIL  demand CSV references SKUs the workbook does not hold: {unknown}")
        return 1

    rows = build_rows(demand, sku_master)

    total_forecast = sum(sum(record["forecast"]) for record in demand)
    total_arrivals = sum(
        float(row[column]) for row in rows for column in ARRIVAL_COLUMNS
    )
    ratio = total_arrivals / total_forecast if total_forecast else 0.0
    weekly_arrivals, weekly_demand = weekly_totals(rows, demand)
    position = running_position(weekly_arrivals, weekly_demand, on_hand)
    low, high = min(weekly_arrivals), max(weekly_arrivals)
    amplitude = (high / low - 1.0) if low else 0.0
    # How far the drawn line strays from the line beside it. This is what
    # decides whether the chart reads as one comparison or as two unrelated
    # magnitudes stacked on one axis.
    mean_demand = sum(weekly_demand) / WEEKS
    spread = max(abs(high / mean_demand - 1.0), abs(low / mean_demand - 1.0))

    print(f"  ..  {len(rows):,} rows, chain on-hand {on_hand:,.0f}")
    print(f"  ..  inbound {total_arrivals:,.0f} against demand {total_forecast:,.0f}")

    # GATE 1: no creep. Total inbound must match total demand across the
    # horizon, or cover drifts away from requirement exactly as it did before.
    low, high = RATIO_BOUNDS
    if not low <= ratio <= high:
        print(f"FAIL  inbound/demand ratio {ratio:.4f} outside [{low}, {high}]")
        return 1
    print(f"  ok  inbound/demand ratio {ratio:.4f} within [{low}, {high}]")

    # GATE 2: the line moves. A schedule arriving in equal weekly instalments
    # would satisfy gate 1 and draw a flat line on top of demand, which says
    # nothing a single number could not.
    if amplitude < MIN_COVER_AMPLITUDE:
        print(
            f"FAIL  arrivals peak-to-trough {amplitude:.1%}"
            f" below {MIN_COVER_AMPLITUDE:.0%}"
        )
        return 1
    print(f"  ok  arrivals peak-to-trough {amplitude:.1%}")

    # GATE 3: and it stays beside demand. Arrivals share an axis with demand
    # now, so a schedule that towers over it is unreadable however well it
    # satisfies the other two.
    if spread > MAX_DEMAND_SPREAD:
        print(
            f"FAIL  arrivals stray {spread:.1%} from weekly demand,"
            f" past {MAX_DEMAND_SPREAD:.0%}"
        )
        return 1
    print(f"  ok  arrivals stay within {spread:.1%} of weekly demand")

    empty = [index + 1 for index, value in enumerate(position) if value <= 0]
    print(
        f"  ..  stock runs out in week(s) {empty}"
        if empty
        else "  ..  stock holds every week"
    )

    with TARGET_CSV.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=COLUMNS, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)

    by_route: dict[str, int] = {}
    for row in rows:
        by_route[row["route"]] = by_route.get(row["route"], 0) + 1

    manifest = {
        "generation_name": GENERATION_NAME,
        "generator_version": GENERATOR_VERSION,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "business_timezone": "Asia/Jakarta",
        "output_row_count": len(rows),
        "quantity_precision": 6,
        "input_fingerprint": fingerprint(DEMAND_CSV),
        "output_fingerprint": fingerprint(TARGET_CSV),
        "column_contract": {
            "columns": COLUMNS,
            "identifier_columns": ["sku_id", "store_id", "route"],
            "arrival_columns": ARRIVAL_COLUMNS,
            "canonical_sort": ["sku_id", "store_id"],
            "numeric_precision": 6,
        },
        "provenance": {
            "arrival_w1_to_arrival_w16": (
                "synthetic. No table in this warehouse records an inbound"
                " arrival date; this schedule is invented to supply one."
            ),
            "quantities": (
                "demand-anchored: each delivery carries forecast_w* demand from"
                " its own week until the next delivery, rounded to whole"
                " pack_factor cases."
            ),
            "timing": (
                "route cadence, phase shared within a route -- direct and flow"
                " weekly, cross every 2 weeks. Invented: no workbook cell"
                " states a delivery interval."
            ),
            "lead_d_and_pack_factor": (
                "resources/dbtemp/schema_with_data.json sku_master"
            ),
        },
        "determinism": {
            "seed": None,
            "note": (
                "No random component. The schedule is a function of the demand"
                " curve and the SKU's route, so a rerun is byte-identical."
            ),
        },
        "plausibility": {
            "inbound_demand_ratio": ratio,
            "inbound_demand_ratio_bounds": list(RATIO_BOUNDS),
            "arrivals_peak_to_trough": amplitude,
            "arrivals_peak_to_trough_floor": MIN_COVER_AMPLITUDE,
            "arrivals_spread_from_demand": spread,
            "arrivals_spread_ceiling": MAX_DEMAND_SPREAD,
            "chain_arrivals_by_week": weekly_arrivals,
            "chain_demand_by_week": weekly_demand,
            "chain_position_by_week": position,
            "rows_by_route": by_route,
            "cadence_weeks": {route: cadence for route, _, cadence in ROUTE_CADENCE},
            "delivery_weeks": {
                route: delivery_weeks(cadence) for route, _, cadence in ROUTE_CADENCE
            },
        },
    }
    TARGET_MANIFEST.write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )

    write_xlsx(rows)

    routes = ", ".join(f"{name} {count:,}" for name, count in sorted(by_route.items()))
    print(f"  ok  {routes}")
    print(f"  ok  wrote {TARGET_CSV.name} ({TARGET_CSV.stat().st_size / 1024:.1f} KB)")
    print(f"  ok  wrote {TARGET_MANIFEST.name}")
    print(f"  ok  wrote {TARGET_XLSX.name} ({TARGET_XLSX.stat().st_size / 1024:.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
