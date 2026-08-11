"""Extract the retail workbook into `resources/dbtemp/schema_with_data.json`.

Run it yourself:

    cd backend
    ../.venv/Scripts/python.exe ../scripts/extract_workbook_schema.py

The workbook behind the Data Source page holds 49 sheets, of which 38 carry
data rather than prose. This script turns those into 30 tables — columns with
inferred types, primary keys, foreign keys and the rows themselves — in one
self-describing JSON file. Nothing is written to Postgres; the file is the
input for a later load, and the record of what the workbook actually contains.

Read-only against the workbook, and it never touches the Formula Manager's
store (`resources/dbtemp/formula.json`) sitting in the same directory.

Rows are positional arrays against the declared column list. Object-per-row
would be twice the bytes for the same content, and ENGINE_STORE alone is
16,000 rows.

The output is validated before it is written: every primary key must be
unique, every foreign key must resolve, and the five sheets the Formula
Manager cites must all have produced a table. A breach fails the run rather
than writing a file that looks fine.
"""

from __future__ import annotations

import json
import os
import sys
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Callable

import openpyxl

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "backend"))

from src.excel.workbook import workbook_path  # noqa: E402

OUTPUT = REPO / "resources" / "dbtemp" / "schema_with_data.json"
VERSION = 1

# The Formula Manager cites cells on exactly these sheets, across
# resources/dbtemp/formula.json and frontend/.../workedExamples.json. A run
# that drops any of them has broken the page's data lineage, so it fails.
FORMULA_SHEETS = {
    "ENGINE_STORE",
    "SKU_Master",
    "Workforce",
    "Constants",
    "Stores",
}

# Symbols that carry meaning in a header and would otherwise sanitise away.
# "OTIF %" must not collide with "OTIF", and "Δ" is a column of its own on
# the What-If sheets.
SYMBOL_WORDS = {
    "%": " pct",
    "Δ": " delta",
    "Σ": " sum",
    "&": " and",
    "/": " ",
    "·": " ",
    "-": " ",
}


@dataclass
class ForeignKey:
    columns: list[str]
    table: str
    references: list[str]


@dataclass
class SheetSpec:
    """One sheet's mapping to one table. The whole configuration lives here."""

    sheet: str
    table: str
    description: str
    header_row: int = 5
    primary_key: list[str] = field(default_factory=list)
    rename: dict[str, str] = field(default_factory=dict)
    foreign_keys: list[ForeignKey] = field(default_factory=list)
    # Sheets that end in a TOTAL row: the aggregate is kept, flagged, and its
    # foreign keys are exempted rather than the row being dropped.
    total_row_marker: str | None = None


@dataclass
class Table:
    name: str
    source_sheet: str
    header_row: int | None
    description: str
    columns: list[str]
    types: list[str]
    nullable: list[bool]
    source_headers: list[str | None]
    primary_key: list[str]
    foreign_keys: list[ForeignKey]
    rows: list[list[Any]]

    def index(self, column: str) -> int:
        return self.columns.index(column)

    def values(self, column: str) -> list[Any]:
        position = self.index(column)
        return [row[position] for row in self.rows]


# --------------------------------------------------------------------------
# Header and value handling
# --------------------------------------------------------------------------


def column_name(header: Any, position: int) -> str:
    """`Lead time (d)` -> `lead_time_d`, `OTIF %` -> `otif_pct`."""
    text = str(header).strip()
    for symbol, word in SYMBOL_WORDS.items():
        text = text.replace(symbol, word)
    cleaned = "".join(c if c.isalnum() else "_" for c in text).lower()
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    cleaned = cleaned.strip("_")
    if not cleaned:
        return f"col_{position + 1}"
    if cleaned[0].isdigit():
        return f"c_{cleaned}"
    return cleaned


def unique_names(names: list[str]) -> list[str]:
    """Suffix repeats so a duplicated header cannot collapse two columns."""
    seen: dict[str, int] = {}
    result = []
    for name in names:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


def json_value(value: Any) -> Any:
    """Make one cell JSON-safe without changing what it means."""
    if value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def column_type(values: list[Any]) -> str:
    """Infer one type per column. Mixed values stay text rather than coerce."""
    present = [v for v in values if v is not None]
    if not present:
        return "text"
    if all(isinstance(v, bool) for v in present):
        return "boolean"
    if all(isinstance(v, int) and not isinstance(v, bool) for v in present):
        return "integer"
    if all(
        isinstance(v, (int, float)) and not isinstance(v, bool) for v in present
    ):
        return "numeric"
    if all(isinstance(v, str) and _is_iso_date(v) for v in present):
        return "date"
    return "text"


def _is_iso_date(value: str) -> bool:
    try:
        datetime.fromisoformat(value)
    except ValueError:
        return False
    return len(value) >= 10


def read_grid(
    worksheet: Any, header_row: int
) -> tuple[list[str], list[str], list[tuple[int, list[Any]]]]:
    """Return sanitised names, the workbook's own captions, and the data rows.

    The captions are kept verbatim as each column's `source_header` — that is
    what maps a column back to the worksheet a reader is looking at.

    Trailing all-blank rows and columns beyond the last named header are
    dropped: openpyxl's max_row/max_column follow formatting, not content.
    """
    rows = list(worksheet.iter_rows(min_row=header_row, values_only=True))
    raw = rows[0]
    keep = [i for i, h in enumerate(raw) if h not in (None, "")]
    captions = [str(raw[i]).strip() for i in keep]
    headers = unique_names(
        [column_name(raw[i], position) for position, i in enumerate(keep)]
    )

    body: list[tuple[int, list[Any]]] = []
    for offset, row in enumerate(rows[1:], start=header_row + 1):
        values = [
            json_value(row[i]) if i < len(row) else None for i in keep
        ]
        if all(v is None for v in values):
            continue
        body.append((offset, values))
    return headers, captions, body


# --------------------------------------------------------------------------
# The configuration
# --------------------------------------------------------------------------

VERTICAL_ID_FK = ForeignKey(["vertical_id"], "verticals", ["vertical_id"])
VERTICAL_LABEL_FK = ForeignKey(
    ["vertical_label"], "verticals", ["dashboard_label"]
)
STORE_FK = ForeignKey(["store_id"], "stores", ["store_id"])
SKU_FK = ForeignKey(["sku_id"], "sku_master", ["sku_id"])
CAT_FK = ForeignKey(["cat_id"], "categories", ["cat_id"])
VENDOR_NAME_FK = ForeignKey(["vendor"], "vendors", ["vendor"])
BRAND_FK = ForeignKey(["brand"], "brand_performance", ["brand"])


def _agent_kpi(sheet: str, table: str, description: str) -> SheetSpec:
    """The nine A-sheets and their siblings: one row per vertical, by label."""
    return SheetSpec(
        sheet=sheet,
        table=table,
        description=description,
        primary_key=["vertical_label"],
        rename={"vertical": "vertical_label"},
        foreign_keys=[VERTICAL_LABEL_FK],
    )


SHEET_SPECS: list[SheetSpec] = [
    # -- dimensions ---------------------------------------------------------
    SheetSpec(
        sheet="Verticals",
        table="verticals",
        description=(
            "The 8 retail verticals. Each is a legal entity in the D365 "
            "mapping. `dashboard_label` is derived, not a workbook column."
        ),
        primary_key=["vertical_id"],
        rename={"id": "vertical_id"},
    ),
    SheetSpec(
        sheet="Stores",
        table="stores",
        description="160 stores, 20 per vertical.",
        primary_key=["store_id"],
        rename={"vertical": "vertical_id"},
        foreign_keys=[VERTICAL_ID_FK],
    ),
    SheetSpec(
        sheet="Categories",
        table="categories",
        description="160 merchandise categories, 20 per vertical.",
        primary_key=["cat_id"],
        rename={"vertical": "vertical_id"},
        foreign_keys=[VERTICAL_ID_FK],
    ),
    SheetSpec(
        sheet="Main Vendor",
        table="vendors",
        description=(
            "8 vendors. `vendor_account` is the D365 key; `vendor` is the "
            "display name every other sheet joins on."
        ),
        primary_key=["vendor_account"],
    ),
    SheetSpec(
        sheet="SKU_Master",
        table="sku_master",
        description="800 released products, 100 per vertical. Item master.",
        primary_key=["sku_id"],
        rename={"vertical": "vertical_id"},
        foreign_keys=[VERTICAL_ID_FK, CAT_FK, VENDOR_NAME_FK, BRAND_FK],
    ),
    # -- detail facts -------------------------------------------------------
    SheetSpec(
        sheet="ENGINE_STORE",
        table="engine_store",
        description=(
            "The per-SKU x per-store decision grid, 800 items x 20 stores per "
            "vertical. The Formula Manager's primary source."
        ),
        header_row=3,
        primary_key=["sku_id", "store_id"],
        rename={"store": "store_id", "vertical": "vertical_id", "cat": "cat_id"},
        foreign_keys=[SKU_FK, STORE_FK, VERTICAL_ID_FK, CAT_FK],
    ),
    SheetSpec(
        sheet="ENGINE",
        table="engine",
        description="Chain-net decision grid, one row per SKU.",
        primary_key=["sku_id"],
        rename={"sku": "sku_id", "vertical": "vertical_id", "cat": "cat_id"},
        foreign_keys=[SKU_FK, VERTICAL_ID_FK, CAT_FK, VENDOR_NAME_FK, BRAND_FK],
    ),
    SheetSpec(
        sheet="Trade Agreement",
        table="trade_agreements",
        description="Vendor price breaks: 3 offers per item across 8 vendors.",
        primary_key=["item", "vendor_account"],
        foreign_keys=[
            ForeignKey(["item"], "sku_master", ["sku_id"]),
            ForeignKey(["vendor_account"], "vendors", ["vendor_account"]),
            VENDOR_NAME_FK,
        ],
    ),
    SheetSpec(
        sheet="Replenishment Detail",
        table="replenishment_detail",
        description="Per-item reorder proposal with designated vs best price.",
        primary_key=["item"],
        rename={"category": "cat_id", "vertical": "vertical_id"},
        foreign_keys=[
            ForeignKey(["item"], "sku_master", ["sku_id"]),
            CAT_FK,
            VERTICAL_ID_FK,
            ForeignKey(["designated_vendor"], "vendors", ["vendor"]),
            ForeignKey(["best_price_vendor"], "vendors", ["vendor"]),
        ],
    ),
    SheetSpec(
        sheet="Workforce",
        table="workforce",
        description=(
            "Per-store roster vs required FTE. The last row is the chain "
            "TOTAL, kept with is_total set."
        ),
        primary_key=["store_id"],
        rename={"store": "store_id", "vertical": "vertical_id"},
        foreign_keys=[STORE_FK, VERTICAL_ID_FK],
        total_row_marker="TOTAL",
    ),
    SheetSpec(
        sheet="Promotion & Discount Detail",
        table="promotion_discount_detail",
        description="48 promotions with their D365 discount construct.",
        primary_key=["promo_id"],
        rename={"vertical": "vertical_label"},
        foreign_keys=[VERTICAL_LABEL_FK],
    ),
    SheetSpec(
        sheet="Brand Events",
        table="brand_events",
        description="23 store-level events and the demand lift each carries.",
        primary_key=["store_id"],
        rename={"store": "store_id", "vertical": "vertical_id"},
        foreign_keys=[STORE_FK, VERTICAL_ID_FK],
    ),
    # -- agent KPI aggregates ----------------------------------------------
    _agent_kpi("A1 Demand Forecasting", "a1_demand_forecasting", "A1 KPIs per vertical."),
    _agent_kpi("A2 Inventory Risk", "a2_inventory_risk", "A2 KPIs per vertical."),
    _agent_kpi("A3 Replenishment", "a3_replenishment", "A3 KPIs per vertical."),
    _agent_kpi("A4 Promotion", "a4_promotion", "A4 KPIs per vertical."),
    _agent_kpi("A5 Pricing & Markdown", "a5_pricing_markdown", "A5 KPIs per vertical."),
    _agent_kpi("A6 Assortment", "a6_assortment", "A6 KPIs per vertical."),
    _agent_kpi("A7 Workforce Optimizer", "a7_workforce_optimizer", "A7 KPIs per vertical."),
    _agent_kpi("A8 Vendor & Brand", "a8_vendor_brand", "A8 KPIs per vertical."),
    _agent_kpi("A9 AI Summary", "a9_ai_summary", "A9 consolidated value at stake."),
    _agent_kpi(
        "What-If · Per Agent",
        "what_if_per_agent",
        "Baseline vs +20% demand, per vertical.",
    ),
    SheetSpec(
        sheet="UOM & PO Summary",
        table="uom_po_summary",
        description=(
            "Purchase-order rollup per vertical. Keys on the vertical code, "
            "not the dashboard label."
        ),
        primary_key=["vertical_id"],
        rename={"vertical": "vertical_id"},
        foreign_keys=[VERTICAL_ID_FK],
    ),
    SheetSpec(
        sheet="Vertical Rollup",
        table="vertical_rollup",
        description=(
            "Every agent's headline figure per vertical. The last row is "
            "CHAIN TOTAL, kept with is_total set."
        ),
        primary_key=["vertical_id"],
        rename={"vertical": "vertical_id"},
        foreign_keys=[VERTICAL_ID_FK],
        total_row_marker="CHAIN TOTAL",
    ),
    SheetSpec(
        sheet="What-If Simulator",
        table="what_if_simulator",
        description="Baseline vs live-lever value, 3 metrics per vertical.",
        header_row=6,
        primary_key=["vertical_label", "metric"],
        rename={"vertical": "vertical_label"},
        foreign_keys=[VERTICAL_LABEL_FK],
    ),
    SheetSpec(
        sheet="Vendor Scorecard",
        table="vendor_scorecard",
        description="OTIF, fill, funding and risk per vendor.",
        primary_key=["vendor"],
        foreign_keys=[VENDOR_NAME_FK],
    ),
    SheetSpec(
        sheet="Brand Performance",
        table="brand_performance",
        description="12 brands. This is the brand dimension for SKUs.",
        primary_key=["brand"],
    ),
]

CHART_SHEETS = [
    "Command Center Charts",
    "A1 Charts",
    "A2 Charts",
    "A3 Charts",
    "A4 Charts",
    "A5 Charts",
    "A6 Charts",
    "A7 Charts",
    "A8 Charts",
]

# Sheets deliberately left out: prose, glossaries and the D365 integration
# reference. Each is one SHEET_SPECS entry away if it is ever wanted.
EXCLUDED_SHEETS = [
    "LISTING",
    "Cover & Storyline",
    "Formulas",
    "Terminology",
    "Data Sources",
    "ERP Approval Matrix",
    "Agentic Prompts",
    "Demo Script",
    "D365 Table Reference",
    "D365 Field Mapping",
    "D365 Worked Example",
]


# --------------------------------------------------------------------------
# Generic sheet -> table
# --------------------------------------------------------------------------


def build_table(worksheet: Any, spec: SheetSpec) -> Table:
    headers, captions, body = read_grid(worksheet, spec.header_row)
    columns = [spec.rename.get(name, name) for name in headers]
    source_headers: list[str | None] = list(captions)

    rows = [list(values) for _, values in body]

    if spec.total_row_marker is not None:
        columns.append("is_total")
        source_headers.append(None)
        key = 0
        for row, (_, values) in zip(rows, body):
            row.append(values[key] == spec.total_row_marker)

    columns.append("source_row")
    source_headers.append(None)
    for row, (source_row, _) in zip(rows, body):
        row.append(source_row)

    return finish_table(
        name=spec.table,
        source_sheet=spec.sheet,
        header_row=spec.header_row,
        description=spec.description,
        columns=columns,
        source_headers=source_headers,
        primary_key=spec.primary_key,
        foreign_keys=spec.foreign_keys,
        rows=rows,
    )


def finish_table(
    *,
    name: str,
    source_sheet: str,
    header_row: int | None,
    description: str,
    columns: list[str],
    source_headers: list[str | None],
    primary_key: list[str],
    foreign_keys: list[ForeignKey],
    rows: list[list[Any]],
) -> Table:
    types = [column_type([row[i] for row in rows]) for i in range(len(columns))]
    nullable = [
        any(row[i] is None for row in rows) for i in range(len(columns))
    ]
    return Table(
        name=name,
        source_sheet=source_sheet,
        header_row=header_row,
        description=description,
        columns=columns,
        types=types,
        nullable=nullable,
        source_headers=source_headers,
        primary_key=primary_key,
        foreign_keys=foreign_keys,
        rows=rows,
    )


# --------------------------------------------------------------------------
# The sheets that are not a plain grid
# --------------------------------------------------------------------------


def build_constants(workbook: Any) -> Table:
    """Constants is two blocks: model parameters, then WHAT-IF LEVERS.

    The worked examples cite Constants!B16 and Constants!B17, which are in the
    second block, so both are loaded and each row keeps its source cell.
    """
    worksheet = workbook["Constants"]
    rows: list[list[Any]] = []
    block = "model_parameter"
    for number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if number <= 5:
            continue
        label = row[0]
        if label in (None, ""):
            continue
        value = json_value(row[1]) if len(row) > 1 else None
        if value is None:
            # A lone label with no value is the block heading.
            block = "what_if_lever"
            continue
        rows.append([block, str(label).strip(), value, f"B{number}", number])

    return finish_table(
        name="constants",
        source_sheet="Constants",
        header_row=5,
        description=(
            "Model parameters and what-if levers. `source_cell` is the "
            "address the Formula Manager cites (Constants!B16 is the demand "
            "lever, B17 the promo lever)."
        ),
        columns=["block", "parameter", "value", "source_cell", "source_row"],
        source_headers=[None, "Parameter", "Value", None, None],
        primary_key=["block", "parameter"],
        foreign_keys=[],
        rows=rows,
    )


def build_time_series(workbook: Any) -> Table:
    """24 months x 8 vertical columns, unpivoted so the FK is expressible."""
    worksheet = workbook["Time Series 24mo"]
    grid = list(worksheet.iter_rows(min_row=5, values_only=True))
    header = grid[0]
    verticals = [h for h in header[1:] if h not in (None, "")]

    rows: list[list[Any]] = []
    for offset, row in enumerate(grid[1:], start=6):
        month = row[0]
        if month in (None, ""):
            continue
        for position, vertical in enumerate(verticals, start=1):
            rows.append(
                [str(month), str(vertical), json_value(row[position]), offset]
            )

    return finish_table(
        name="time_series_24mo",
        source_sheet="Time Series 24mo",
        header_row=5,
        description=(
            "24 months of GMV per vertical, unpivoted from 8 wide columns "
            "into one row per (month, vertical)."
        ),
        columns=["month", "vertical_id", "gmv", "source_row"],
        source_headers=["Month", None, None, None],
        primary_key=["month", "vertical_id"],
        foreign_keys=[VERTICAL_ID_FK],
        rows=rows,
    )


def build_chart_series(workbook: Any) -> Table:
    """The 9 chart sheets, which are stacked label/value blocks, not tables.

    Each block is a title row (column A only), a mini header naming the
    series, then label/value pairs. Nine near-identical two-column schemas
    would say less than one long table does.
    """
    rows: list[list[Any]] = []
    identifier = 0
    for sheet in CHART_SHEETS:
        worksheet = workbook[sheet]
        block_no = 0
        block_title = ""
        series_label = ""
        for number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            if number <= 4:
                continue
            label = row[0]
            if label in (None, ""):
                continue
            value = row[1] if len(row) > 1 else None
            if value in (None, ""):
                block_no += 1
                block_title = str(label).strip()
                continue
            if isinstance(value, str):
                series_label = str(label).strip()
                continue
            identifier += 1
            rows.append(
                [
                    identifier,
                    sheet,
                    block_no,
                    block_title,
                    series_label,
                    str(label).strip(),
                    json_value(value),
                    number,
                ]
            )

    return finish_table(
        name="chart_series",
        source_sheet=" + ".join(CHART_SHEETS),
        header_row=None,
        description=(
            "Every plotted point from the 9 chart sheets. Blocks are numbered "
            "within their sheet; (sheet, block_no, category_label) is the "
            "natural key."
        ),
        columns=[
            "id",
            "sheet",
            "block_no",
            "block_title",
            "series_label",
            "category_label",
            "value",
            "source_row",
        ],
        source_headers=[None, None, None, None, None, None, None, None],
        primary_key=["id"],
        foreign_keys=[],
        rows=rows,
    )


SPECIAL_BUILDERS: dict[str, Callable[[Any], Table]] = {
    "constants": build_constants,
    "time_series_24mo": build_time_series,
    "chart_series": build_chart_series,
}


def add_dashboard_labels(verticals: Table, workbook: Any) -> None:
    """Give `verticals` the label the A-sheets actually use.

    `Verticals.Short` says "Department Store" and "Online" where every A-sheet
    says "General Merch" and "Digital/Online". The A-sheets list verticals in
    workbook order, so the mapping is positional — and asserted, because a
    silent mis-pairing here would mislabel every KPI table.
    """
    labels = [
        row[0]
        for row in workbook["A1 Demand Forecasting"].iter_rows(
            min_row=6, values_only=True
        )
        if row[0] not in (None, "")
    ]
    if len(labels) != len(verticals.rows):
        raise SystemExit(
            f"FAIL  A1 lists {len(labels)} verticals, Verticals has "
            f"{len(verticals.rows)}. The positional label mapping is unsafe."
        )

    verticals.columns.insert(3, "dashboard_label")
    verticals.source_headers.insert(3, None)
    verticals.types.insert(3, "text")
    verticals.nullable.insert(3, False)
    for row, label in zip(verticals.rows, labels):
        row.insert(3, str(label))


# --------------------------------------------------------------------------
# Validation
# --------------------------------------------------------------------------


def validate(tables: dict[str, Table]) -> list[str]:
    problems: list[str] = []

    for table in tables.values():
        if not table.primary_key:
            problems.append(f"{table.name}: no primary key")
            continue
        positions = [table.index(column) for column in table.primary_key]
        keys = [tuple(row[i] for i in positions) for row in table.rows]
        if len(set(keys)) != len(keys):
            duplicates = len(keys) - len(set(keys))
            problems.append(
                f"{table.name}: primary key {table.primary_key} has "
                f"{duplicates} duplicate(s)"
            )
        if any(None in key for key in keys):
            problems.append(f"{table.name}: primary key has NULLs")

    for table in tables.values():
        total = (
            table.values("is_total")
            if "is_total" in table.columns
            else [False] * len(table.rows)
        )
        for key in table.foreign_keys:
            parent = tables.get(key.table)
            if parent is None:
                problems.append(
                    f"{table.name}: FK targets unknown table {key.table}"
                )
                continue
            for child_column, parent_column in zip(key.columns, key.references):
                if child_column not in table.columns:
                    problems.append(
                        f"{table.name}: FK column {child_column} does not exist"
                    )
                    continue
                allowed = set(parent.values(parent_column))
                child = table.values(child_column)
                orphans = {
                    value
                    for value, aggregate in zip(child, total)
                    if not aggregate and value is not None and value not in allowed
                }
                if orphans:
                    sample = ", ".join(sorted(str(v) for v in orphans)[:4])
                    problems.append(
                        f"{table.name}.{child_column} -> {key.table}."
                        f"{parent_column}: {len(orphans)} orphan(s) [{sample}]"
                    )

    covered = {table.source_sheet for table in tables.values()}
    missing = FORMULA_SHEETS - covered
    if missing:
        problems.append(
            "Formula Manager sheets missing from the output: "
            + ", ".join(sorted(missing))
        )

    return problems


# --------------------------------------------------------------------------
# Output
# --------------------------------------------------------------------------


def serialise(tables: list[Table], source: Path) -> str:
    """Indented structure, but one line per data row.

    Letting json.dumps(indent=4) format the rows too would put every single
    cell on its own line: 18 MB rather than 6 MB, for the same content and
    worse readability. So the rows are rendered compactly and spliced in.
    """
    placeholders = {}
    payload = {
        "version": VERSION,
        "source_workbook": source.name,
        "generated_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "table_count": len(tables),
        "row_count": sum(len(table.rows) for table in tables),
        "excluded_sheets": EXCLUDED_SHEETS,
        "tables": [
            {
                "name": table.name,
                "source_sheet": table.source_sheet,
                "header_row": table.header_row,
                "description": table.description,
                "row_count": len(table.rows),
                "columns": [
                    {
                        "name": name,
                        "source_header": source_header,
                        "type": declared,
                        "nullable": is_nullable,
                    }
                    for name, source_header, declared, is_nullable in zip(
                        table.columns,
                        table.source_headers,
                        table.types,
                        table.nullable,
                    )
                ],
                "primary_key": table.primary_key,
                "foreign_keys": [
                    {
                        "columns": key.columns,
                        "references": {
                            "table": key.table,
                            "columns": key.references,
                        },
                    }
                    for key in table.foreign_keys
                ],
                "rows": placeholders.setdefault(
                    table.name, f"\x00rows:{table.name}\x00"
                ),
            }
            for table in tables
        ],
    }

    body = json.dumps(payload, indent=4, ensure_ascii=False)
    for table in tables:
        marker = json.dumps(placeholders[table.name])
        if not table.rows:
            body = body.replace(marker, "[]")
            continue
        # 12 spaces: "rows" sits four levels deep inside the table object.
        lines = ",\n".join(
            "                " + json.dumps(row, ensure_ascii=False)
            for row in table.rows
        )
        body = body.replace(marker, "[\n" + lines + "\n            ]")
    return body


def write_atomically(path: Path, body: str) -> None:
    """Temp file + os.replace, matching src/formulas/repository.py.

    formula.json lives in this directory. A half-written 6 MB file next to it
    is not something the Formula Manager should ever be able to observe.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    handle, temporary = tempfile.mkstemp(
        dir=str(path.parent), prefix=f".{path.name}.", suffix=".tmp"
    )
    try:
        with os.fdopen(handle, "w", encoding="utf-8", newline="\n") as stream:
            stream.write(body + "\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, path)
    except BaseException:
        if os.path.exists(temporary):
            os.unlink(temporary)
        raise


def main() -> int:
    source = workbook_path()
    if not source.exists():
        print(f"FAIL  workbook not found: {source}")
        return 1

    print(f"Reading {source.name}")
    workbook = openpyxl.load_workbook(source, read_only=True, data_only=True)

    tables: dict[str, Table] = {}
    order: list[str] = []

    for spec in SHEET_SPECS:
        if spec.sheet not in workbook.sheetnames:
            print(f"FAIL  sheet not in workbook: {spec.sheet}")
            return 1
        table = build_table(workbook[spec.sheet], spec)
        tables[table.name] = table
        order.append(table.name)

    add_dashboard_labels(tables["verticals"], workbook)

    for name, builder in SPECIAL_BUILDERS.items():
        table = builder(workbook)
        tables[name] = table
        order.append(name)

    workbook.close()

    problems = validate(tables)
    if problems:
        print("\nFAIL  the extract does not hold together:\n")
        for problem in problems:
            print(f"  - {problem}")
        return 1

    ordered = [tables[name] for name in order]
    body = serialise(ordered, source)
    write_atomically(OUTPUT, body)

    print(f"\nWrote {OUTPUT.relative_to(REPO)}  "
          f"({len(body) / 1_048_576:.2f} MB)\n")
    print(f"{'table':<28} {'sheet':<30} {'rows':>7}  {'cols':>4}  fks")
    print("-" * 84)
    for table in ordered:
        sheet = table.source_sheet
        if len(sheet) > 29:
            sheet = sheet[:26] + "..."
        print(
            f"{table.name:<28} {sheet:<30} {len(table.rows):>7}  "
            f"{len(table.columns):>4}  {len(table.foreign_keys)}"
        )
    print("-" * 84)
    print(
        f"{len(ordered):<28} {'':<30} "
        f"{sum(len(t.rows) for t in ordered):>7}\n"
    )
    print(
        "Every primary key is unique, every foreign key resolves, and all "
        f"{len(FORMULA_SHEETS)} Formula Manager sheets are covered."
    )
    print(f"Excluded as prose or reference: {len(EXCLUDED_SHEETS)} sheets.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
