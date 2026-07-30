"""Mark the QC workbook with what was fixed on 30 July 2026.

Writes a new file rather than overwriting: openpyxl drops the x14 data
validation extension on save, so the original keeps its dropdowns intact.
"""

from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

SRC = "AI_Finance_Forum_Technical_Priorities_Bilingual_2026-07-29.xlsx..xlsx"
DST = "AI_Finance_Forum_Technical_Priorities_FIX-STATUS_2026-07-30.xlsx"

FIXED = "Fixed - awaiting retest"
FIXED_ID = "Sudah diperbaiki - menunggu uji ulang"
PARTIAL = "In progress"
PARTIAL_ID = "Sedang dikerjakan"

GREEN = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(color="006100", bold=True)
AMBER = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT = Font(color="9C6500", bold=True)
HEAD = PatternFill("solid", fgColor="DDEBF7")

C1 = "53daa2c"
C2 = "f448e66"

# QC id -> (fully fixed?, evidence note)
NOTES: dict[str, tuple[bool, str]] = {
    "QC-001": (True, (
        f"Fixed 30 Jul 2026 (branch bug-fix-trial, commit {C1}). "
        "category_breakdowns.amount_at_risk_idr_mn was probed as "
        "amount_idr_mn, so _row_get returned None and every value became 0.0 "
        "with no error. Verified on live DB batch 17: bars 3,800 / 3,050 / "
        "900 / 95 sum to 7,845 = the Flagged this cycle card."
    )),
    "QC-024": (True, (
        f"Fixed 30 Jul 2026 (commit {C1}). Same root cause as QC-001; the "
        "default view 'categories' now renders four bars on open."
    )),
    "QC-014": (True, (
        f"Fixed 30 Jul 2026 (commit {C1}). The KPI now reads "
        "summary.items_flagged (10). The 22 came from len(anomalies), which "
        "is LIMIT 30 and includes rows that are not flagged."
    )),
    "QC-015": (False, (
        f"Partly fixed 30 Jul 2026 (commit {C1}). The dashboard no longer "
        "contradicts itself: rows with is_direct_loss = false are out of the "
        "category bar and mix donut, so both sum to 7,845, and the 1,950 is "
        "named in the chart note. The alert wording has NOT been checked yet."
    )),
    "QC-007": (True, (
        f"Fixed 30 Jul 2026 (commit {C1}). The card showed "
        "high_risk_provision (2,500) under a label reading 'exposure'. It now "
        "shows the High tier exposure (5,000), which equals the chart bar, "
        "with the provision moved to the caption."
    )),
    "QC-013": (True, (
        f"Fixed 30 Jul 2026 (commit {C1}). The bar is now labelled 'Current "
        "95/90%' and reuses the Total protected card value (7,557.5) instead "
        "of recomputing it at 95/95 (7,577.5)."
    )),
    "QC-027": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Python and IEEE round half to "
        "even, so round(36.25, 1) is 36.2. Added _round_half_up and routed "
        "_fmt and _pct through it. Product GM% is now 23.8 / 36.3 / 16.7."
    )),
    "QC-009": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). The target now comes from the "
        "budget row (7,200 / 46,000 = 15.65%) and is threaded through the "
        "card caption, gauge label, simulator baseline and the recalculate "
        "endpoint. NOTE: formula check #18 expects 9.2/15 = 0.6133; with one "
        "target it becomes 9.2/15.65 = 0.586, so that check needs updating."
    )),
    "QC-039": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Removed the 'illustrative' flags, "
        "the db_kpis_count / db_profit_count / db_variance_count debug "
        "counters, and reworded the debug-flavoured notes. A test now fails "
        "if any of them come back."
    )),
    "QC-046": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Score = 50% at-risk amount + 35% "
        "worst severity + 15% flag count, relative to the worst vendor in the "
        "batch. Live batch 17: Osaka 100, Shenzhen 64, Bahan Baku 56, "
        "Taipei 33."
    )),
    "QC-029": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). One USD scale across the KPI row: "
        "3.3 M USD beside 2.0 M USD, instead of 3.3 M USD beside 2,000,000 "
        "USD which hid that 2.0 is less than 3.3."
    )),
    "QC-034": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). The zero 'Base' bar is gone. The "
        "chart now compares Do nothing 1,777.05 against Hedge 2.0M 700.05 - "
        "both non-zero, and the first equals the FX loss card."
    )),
    "QC-036": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Finance side:top was "
        "byte-identical to view:product. It now shows the gross margin pool "
        "in IDR (4,550 / 4,930 / 2,300 = 11,780) rather than repeating the "
        "rate."
    )),
    "QC-033": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Both charts now order [Now, "
        "Target]. Aligning them turned the pair into exact duplicates, so "
        "side:bottom was changed to Overdue vs expected recovery."
    )),
    "QC-028": (True, (
        f"Fixed 30 Jul 2026 (commit {C2}). Added a lower_is_better flag so "
        "enrichment inverts the ratio. DSO progress is now 0.8246 (47/57) "
        "instead of the clamped 1.2."
    )),
}

# 03_Formula_Checks: check number -> (verdict now, note)
FORMULA = {
    11: ("PASS", "Was 36.2 (banker's rounding). Now 23.8 / 36.3 / 16.7 - QC-027."),
    15: ("PASS", "One target everywhere: 15.7% from the budget row - QC-009."),
    18: ("REVIEW", (
        "Expected value is now stale. With a single target the progress is "
        "9.2 / 15.65 = 0.586, not 9.2 / 15 = 0.6133. The formula "
        "(value / target) is unchanged; only the constant moved."
    )),
    35: ("PASS", "Card now shows the High tier exposure 5,000 = chart bar - QC-007."),
    37: ("PASS", (
        "Still PASS. The chart now shows overbilling as one 900 category; "
        "400 recoverable + 500 blocked is that same 900 split by payment "
        "status. Both decompositions total 7,845."
    )),
    41: ("PASS", "Verified live: 6,370 / 7,060 / 7,557.5."),
    43: ("PASS", "Bar reuses the card value, so the 20 mn gap cannot recur - QC-013."),
    44: ("PASS", "KPI reads summary.items_flagged; both surfaces say 10 - QC-014."),
    45: ("PARTLY", (
        "Dashboard side resolved: control weaknesses are out of the category "
        "charts and the 1,950 is named in the note. Alert wording still to "
        "be checked - QC-015."
    )),
    46: ("PASS", "Score now weights amount and severity - QC-046."),
}


def style(cell, fixed: bool) -> None:
    cell.fill = GREEN if fixed else AMBER
    cell.font = GREEN_FONT if fixed else AMBER_FONT


def append_note(cell, note: str) -> None:
    existing = str(cell.value).strip() if cell.value else ""
    cell.value = f"{existing} | {note}" if existing else note
    cell.alignment = Alignment(wrap_text=True, vertical="top")


def main() -> None:
    wb = openpyxl.load_workbook(SRC, data_only=False)
    touched = {"findings": 0, "priorities": 0, "formula": 0}

    # --- 01_Findings (EN) and 01B_Findings_ID -------------------------------
    for sheet, status_fixed, status_partial in (
        ("01_Findings", FIXED, PARTIAL),
        ("01B_Findings_ID", FIXED_ID, PARTIAL_ID),
    ):
        ws = wb[sheet]
        for row in range(5, ws.max_row + 1):
            qc = ws[f"A{row}"].value
            if qc not in NOTES:
                continue
            fixed, note = NOTES[qc]
            cell = ws[f"N{row}"]
            cell.value = status_fixed if fixed else status_partial
            style(cell, fixed)
            append_note(ws[f"Q{row}"], note)
            touched["findings"] += 1

    # --- Technical Priorities EN / ID ---------------------------------------
    for sheet in ("Technical Priorities EN", "Technical Priorities ID"):
        ws = wb[sheet]
        header = ws.cell(39, 11)
        header.value = "Fix status 30 Jul 2026"
        header.fill = HEAD
        header.font = Font(bold=True)
        ws.column_dimensions["K"].width = 62
        for row in range(40, ws.max_row + 1):
            qc = ws.cell(row, 1).value
            if qc not in NOTES:
                continue
            fixed, note = NOTES[qc]
            status = ws.cell(row, 6)
            status.value = (
                (FIXED if sheet.endswith("EN") else FIXED_ID)
                if fixed
                else (PARTIAL if sheet.endswith("EN") else PARTIAL_ID)
            )
            style(status, fixed)
            target = ws.cell(row, 11)
            target.value = note
            target.alignment = Alignment(wrap_text=True, vertical="top")
            touched["priorities"] += 1

    # --- 03_Formula_Checks ---------------------------------------------------
    ws = wb["03_Formula_Checks"]
    for col, title, width in ((8, "Retest 30 Jul 2026", 20), (9, "What changed", 70)):
        cell = ws.cell(4, col)
        cell.value = title
        cell.fill = HEAD
        cell.font = Font(bold=True)
        ws.column_dimensions[cell.column_letter].width = width
    for row in range(5, ws.max_row + 1):
        number = ws.cell(row, 1).value
        try:
            number = int(number)
        except (TypeError, ValueError):
            continue
        if number not in FORMULA:
            continue
        verdict, note = FORMULA[number]
        cell = ws.cell(row, 8)
        cell.value = verdict
        style(cell, verdict == "PASS")
        detail = ws.cell(row, 9)
        detail.value = note
        detail.alignment = Alignment(wrap_text=True, vertical="top")
        touched["formula"] += 1

    # --- Manual tests blocked by the blank-chart defect ----------------------
    # The Result column is deliberately left alone: a human still has to run
    # these on the app. Only the blocking reason is updated.
    blocked_by_charts = {
        "MT-001": "Agent switcher",
        "MT-002": "View tabs",
    }
    for sheet, note in (
        ("02_Manual_Test_Checklist", (
            "30 Jul 2026: the blank Leakage charts that caused this Partial "
            "are fixed (QC-001 / QC-024). Ready to re-run."
        )),
        ("02B_Manual_Test_Checklist_ID", (
            "30 Jul 2026: chart Leakage kosong yang menyebabkan Sebagian "
            "sudah diperbaiki (QC-001 / QC-024). Siap diuji ulang."
        )),
    ):
        ws = wb[sheet]
        for row in range(5, ws.max_row + 1):
            if ws.cell(row, 1).value not in blocked_by_charts:
                continue
            cell = ws.cell(row, 12)  # Notes
            append_note(cell, note)
            touched["manual_tests"] = touched.get("manual_tests", 0) + 1

    wb.save(DST)
    print(f"wrote {DST}")
    print(f"  findings rows marked : {touched['findings']} (EN + ID)")
    print(f"  priorities rows      : {touched['priorities']} (EN + ID)")
    print(f"  formula checks       : {touched['formula']}")


if __name__ == "__main__":
    main()
