"""Load `Dataset_AI_Finance_Forum_V1.0` into the `newdata` schema.

Run it yourself:

    cd backend
    ../.venv/Scripts/python.exe ../scripts/import_new_dataset.py

The script is idempotent: every table is dropped and recreated, so re-running it
is the reset. `newdata` is therefore disposable, and nothing hand-authored may
be stored there — a recommendation, an edited row or an alert threshold written
into these tables is destroyed by the next run. `DROP SCHEMA newdata CASCADE`
undoes the whole import.

Nothing outside `newdata` is written except one `audit.import_batches` row: one
batch for the whole workbook, which is the point of the exercise. The old four
schemas are untouched and the running app does not read anything created here.

What it prints is a load report — sheet, table, row count, and every column that
had to fall back to text. Check the load against that rather than trusting it.
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy import text

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "backend"))

from src.db.db import get_engine  # noqa: E402

WORKBOOK = REPO / "Dataset_AI_Finance_Forum_V1.0_20260728.xlsx"
SCHEMA = "newdata"

# Prose and formula sheets. 90_Reconciliation holds Excel formulas rather than
# data; its 14 checks are re-expressed as SQL in verify_new_dataset.py.
SKIP_SHEETS = {"00_README", "LISTING", "90_Reconciliation"}

# Dimension tables get a primary key. Facts do not yet: foreign keys and
# indexes are worth adding once the query shapes are known, not before.
DIMENSION_KEYS = {
    "dim_legal_entity": "legal_entity_id",
    "dim_store": "store_id",
    "dim_category": "category_id",
    "dim_item": "item_id",
    "dim_customer": "customer_id",
    "dim_vendor": "vendor_id",
}

INSERT_CHUNK = 1000

ISO_DATE = re.compile(r"\d{4}-\d{2}-\d{2}")


def table_name(sheet: str) -> str:
    """`20_FACT_Sales` -> `fact_sales`. The sheet number is ordering, not meaning."""
    name = sheet.split("_", 1)[1] if sheet[:1].isdigit() else sheet
    return "".join(c if c.isalnum() else "_" for c in name).strip("_").lower()


def column_name(header: Any, position: int) -> str:
    """Sanitise a header into a column name.

    Headers that are pure punctuation — `91_Filter_Requirements` leads with a
    `#` counter — sanitise to nothing, so they fall back to their position
    rather than producing an unquotable empty identifier.
    """
    cleaned = "".join(
        c if c.isalnum() else "_" for c in str(header).strip()
    ).strip("_").lower()
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    if not cleaned or cleaned[0].isdigit():
        cleaned = f"col_{position + 1}" if not cleaned else f"c_{cleaned}"
    return cleaned


def unique_names(headers: list[str]) -> list[str]:
    """Suffix repeats so a duplicated header cannot collapse two columns."""
    seen: dict[str, int] = {}
    result = []
    for name in headers:
        if name in seen:
            seen[name] += 1
            result.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            result.append(name)
    return result


def header_row(worksheet: Any) -> int | None:
    """Find the first row that looks like a header.

    The offset is not constant across the workbook — row 5 on the dimension
    sheets and FACT_Sales, row 6 on the derived packs and 50_Agent_KPIs — and
    each sheet carries a title and a description above it. Hardcoding either
    number silently loads one group of sheets wrong, so the header is found by
    looking for the first row wide enough to be one.
    """
    for index, row in enumerate(
        worksheet.iter_rows(min_row=1, max_row=20, values_only=True),
        start=1,
    ):
        filled = [cell for cell in row if cell not in (None, "")]
        if len(filled) >= 3:
            return index
    return None


def sql_type(values: list[Any]) -> tuple[str, bool]:
    """Infer one Postgres type for a column. Returns (type, fell_back_to_text).

    A column whose values disagree becomes text rather than being coerced. That
    is reported, because a numeric column arriving as text is the failure mode
    worth catching early — it does not error, it just makes later SUMs
    impossible.
    """
    present = [v for v in values if v not in (None, "")]
    if not present:
        return "TEXT", False

    if all(isinstance(v, bool) for v in present):
        return "BOOLEAN", False
    # The workbook writes dates as ISO strings, not as Excel dates. Leaving
    # them as text would work for sorting but not for period arithmetic, and
    # `month` is the dimension this whole migration exists to get.
    if all(isinstance(v, str) and ISO_DATE.fullmatch(v) for v in present):
        return "DATE", False
    if all(isinstance(v, datetime) for v in present):
        return "TIMESTAMP", False
    if all(isinstance(v, date) and not isinstance(v, datetime) for v in present):
        return "DATE", False
    if all(isinstance(v, (date, datetime)) for v in present):
        return "TIMESTAMP", False
    if all(isinstance(v, int) and not isinstance(v, bool) for v in present):
        return "BIGINT", False
    if all(
        isinstance(v, (int, float, Decimal)) and not isinstance(v, bool)
        for v in present
    ):
        return "NUMERIC(24,6)", False

    mixed = len({type(v) for v in present}) > 1
    return "TEXT", mixed


def cell_value(value: Any, declared: str) -> Any:
    if value == "":
        return None
    if declared == "TEXT" and value is not None and not isinstance(value, str):
        return str(value)
    return value


def read_sheet(worksheet: Any) -> tuple[list[str], list[list[Any]]] | None:
    start = header_row(worksheet)
    if start is None:
        return None

    rows = list(worksheet.iter_rows(min_row=start, values_only=True))
    if not rows:
        return None

    raw_headers = rows[0]
    keep = [i for i, h in enumerate(raw_headers) if h not in (None, "")]
    headers = unique_names(
        [column_name(raw_headers[i], position) for position, i in enumerate(keep)]
    )

    body = []
    for row in rows[1:]:
        values = [row[i] if i < len(row) else None for i in keep]
        if all(v in (None, "") for v in values):
            continue
        body.append(values)

    return headers, body


def create_and_load(
    connection: Any,
    table: str,
    headers: list[str],
    body: list[list[Any]],
) -> list[str]:
    columns = []
    coerced = []
    for position, name in enumerate(headers):
        declared, fell_back = sql_type([row[position] for row in body])
        columns.append((name, declared))
        if fell_back:
            coerced.append(name)

    # Sheet order is meaning, not decoration: 51_Finance_PL reads Net revenue,
    # COGS, Gross margin, Opex, EBITDA, and a P&L shuffled by the planner is
    # wrong even when every figure is right. SQL guarantees no order without
    # ORDER BY, so the row's position is carried as a column.
    definitions = ['"_row_order" INTEGER NOT NULL']
    definitions += [f'"{name}" {declared}' for name, declared in columns]
    key = DIMENSION_KEYS.get(table)
    if key and key in headers:
        definitions.append(f'PRIMARY KEY ("{key}")')

    connection.execute(text(f'DROP TABLE IF EXISTS {SCHEMA}."{table}" CASCADE'))
    connection.execute(
        text(f'CREATE TABLE {SCHEMA}."{table}" ({", ".join(definitions)})')
    )

    if body:
        placeholders = ", ".join(
            [":_row_order"] + [f":{name}" for name, _ in columns]
        )
        quoted = ", ".join(
            ['"_row_order"'] + [f'"{name}"' for name, _ in columns]
        )
        statement = text(
            f'INSERT INTO {SCHEMA}."{table}" ({quoted}) VALUES ({placeholders})'
        )
        declared_by_name = dict(columns)
        for start in range(0, len(body), INSERT_CHUNK):
            chunk = body[start:start + INSERT_CHUNK]
            connection.execute(
                statement,
                [
                    {
                        "_row_order": start + offset + 1,
                        **{
                            name: cell_value(row[position], declared_by_name[name])
                            for position, name in enumerate(headers)
                        },
                    }
                    for offset, row in enumerate(chunk)
                ],
            )

    return coerced


def main() -> int:
    if not WORKBOOK.exists():
        print(f"FAIL  workbook not found: {WORKBOOK}")
        return 1

    print(f"Reading {WORKBOOK.name}")
    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)

    sheets = [ws for ws in workbook.worksheets if ws.title not in SKIP_SHEETS]
    parsed: list[tuple[str, str, list[str], list[list[Any]]]] = []
    for worksheet in sheets:
        result = read_sheet(worksheet)
        if result is None:
            print(f"SKIP  {worksheet.title} — no header row found")
            continue
        headers, body = result
        parsed.append((worksheet.title, table_name(worksheet.title), headers, body))

    if not parsed:
        print("FAIL  no loadable sheets")
        return 1

    # data_only reads the values Excel cached at its last save. A workbook
    # written by a generator and never opened in Excel has no cached values, so
    # every computed cell reads as None. Say so rather than loading empty
    # tables that look like a successful import.
    total_rows = sum(len(body) for _, _, _, body in parsed)
    if total_rows == 0:
        print(
            "FAIL  every sheet read as empty. The workbook has no cached "
            "values — open it in Excel and save it, then re-run."
        )
        return 1

    engine = get_engine()
    report: list[tuple[str, str, int, list[str]]] = []

    with engine.begin() as connection:
        connection.execute(text(f"CREATE SCHEMA IF NOT EXISTS {SCHEMA}"))

        for sheet, table, headers, body in parsed:
            coerced = create_and_load(connection, table, headers, body)
            report.append((sheet, table, len(body), coerced))

        batch_id = connection.execute(
            text(
                """
                INSERT INTO audit.import_batches (
                    agent_name, workbook_name, workbook_version, workbook_path,
                    import_status, imported_by, completed_at,
                    total_sheets, total_rows
                ) VALUES (
                    'new_dataset', :workbook_name, 'V1.0', :workbook_path,
                    'COMPLETED', 'import_new_dataset.py', CURRENT_TIMESTAMP,
                    :total_sheets, :total_rows
                )
                RETURNING id
                """
            ),
            {
                "workbook_name": WORKBOOK.name,
                "workbook_path": str(WORKBOOK),
                "total_sheets": len(report),
                "total_rows": total_rows,
            },
        ).scalar_one()

    print(f"\nLoaded into schema {SCHEMA}, import batch {batch_id}\n")
    print(f"{'sheet':<28} {'table':<26} {'rows':>7}")
    print("-" * 63)
    for sheet, table, rows, _ in report:
        print(f"{sheet:<28} {table:<26} {rows:>7}")
    print("-" * 63)
    print(f"{'':<28} {len(report):<26} {total_rows:>7}\n")

    fallbacks = [(t, c) for _, t, _, c in report if c]
    if fallbacks:
        print("Columns held as TEXT because their values disagreed:")
        for table, columns in fallbacks:
            print(f"  {table}: {', '.join(columns)}")
        print()

    print(
        "Note: simulator levers, alerts and recommendations are not in this "
        "workbook. Anything using them is still reading the old schemas."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
