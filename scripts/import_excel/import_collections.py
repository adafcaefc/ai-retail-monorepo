from __future__ import annotations

import sys
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session


BACKEND_ROOT = Path(__file__).resolve().parents[2]

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))


from src.db.db import get_session_factory
from src.db.models import (
    CollectionAssumption,
    CollectionDsoCashImpact,
    CollectionRecommendation,
    CollectionRiskScore,
    CollectionRiskTierExposure,
    CollectionWorklistItem,
    CustomerCreditAging,
    ImportBatch,
)


WORKBOOK_NAME = (
    "03C_Collections Credit Agent Demo Aug 2026 "
    "Data Engine v1.0 20260707.xlsx"
)

WORKBOOK_PATH = BACKEND_ROOT / "data" / WORKBOOK_NAME

WORKBOOK_VERSION = "v1.0-20260707"
AGENT_NAME = "collections_credit_agent"

EXPECTED_SHEETS = [
    "01 Assumptions",
    "02 Customer Credit & Aging",
    "03 Risk Scoring",
    "04 DSO & Cash Impact",
    "05 Collections Worklist",
    "06 Recommendation",
]


def as_text(value: Any) -> str | None:
    if value is None:
        return None

    text_value = str(value).strip()

    if not text_value:
        return None

    return text_value


def as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return Decimal(int(value))

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    try:
        cleaned_value = str(value).replace(",", "").strip()
        return Decimal(cleaned_value)
    except (TypeError, ValueError):
        return None


def as_integer(value: Any) -> int | None:
    if value is None or value == "":
        return None

    if isinstance(value, bool):
        return int(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def as_boolean(value: Any) -> bool:
    text_value = as_text(value)

    if text_value is None:
        return False

    return text_value.lower() in {
        "yes",
        "true",
        "1",
    }


def validate_workbook(workbook: Any) -> None:
    missing_sheets = [
        sheet_name
        for sheet_name in EXPECTED_SHEETS
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise RuntimeError(
            "Missing required worksheets: "
            + ", ".join(missing_sheets)
        )


def find_row_by_first_cell(
    worksheet: Any,
    expected_value: str,
) -> int:
    for row_number in range(
        1,
        worksheet.max_row + 1,
    ):
        cell_value = as_text(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if cell_value == expected_value:
            return row_number

    raise RuntimeError(
        f"Could not find row {expected_value!r} "
        f"in sheet {worksheet.title!r}."
    )


def get_label_value(
    worksheet: Any,
    label: str,
) -> Any:
    row_number = find_row_by_first_cell(
        worksheet,
        label,
    )

    return worksheet.cell(
        row=row_number,
        column=2,
    ).value


def create_import_batch(
    session: Session,
) -> int:
    imported_by = session.scalar(
        select(func.current_user())
    )

    import_batch = ImportBatch(
        agent_name=AGENT_NAME,
        workbook_name=WORKBOOK_NAME,
        workbook_version=WORKBOOK_VERSION,
        workbook_path=str(WORKBOOK_PATH),
        import_status="STARTED",
        imported_by=imported_by,
        total_sheets=len(EXPECTED_SHEETS),
        total_rows=0,
        metadata_json={
            "source": "Excel Data Engine",
            "data_type": "illustrative_demo_data",
            "scope": (
                "Collections and Credit "
                "Intelligence Agent"
            ),
        },
    )

    session.add(import_batch)
    session.flush()

    if import_batch.id is None:
        raise RuntimeError(
            "Failed to create Collections import batch."
        )

    return int(import_batch.id)


def import_assumptions(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    group_names = {
        "SALES & DSO TARGET": "SALES_AND_DSO_TARGET",
        "RISK SCORING MODEL  (points, max 100)": (
            "RISK_SCORING_MODEL"
        ),
        (
            "RECOVERY & PROVISION  "
            "(illustrative placeholders)"
        ): "RECOVERY_AND_PROVISION",
    }

    unit_mapping = {
        "Annual credit sales (IDR mn)": "IDR million",
        "Days in year": "days",
        "Target DSO (days)": "days",
        "Spot USD / IDR (for export customers)": (
            "IDR per USD"
        ),
        "DBT weight (days beyond terms)": "points",
        "Overdue severity weight (61+ days share)": "points",
        "Credit utilization weight (balance / limit)": (
            "points"
        ),
        "Trend points: Worsening": "points",
        "Trend points: Stable": "points",
        "Trend points: Improving": "points",
        "Dispute points (Yes)": "points",
        "DBT days for maximum points": "days",
        "Tier threshold: Medium (min score)": "score",
        "Tier threshold: High (min score)": "score",
        "Low-risk recovery of overdue (%)": "percentage",
        "Medium-risk recovery of overdue (%)": "percentage",
        "High-risk recovery of overdue (%)": "percentage",
        "High-risk loss rate (ECL proxy, %)": "percentage",
    }

    current_group: str | None = None
    records: list[CollectionAssumption] = []

    for row_number in range(
        1,
        worksheet.max_row + 1,
    ):
        assumption_name = as_text(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        assumption_value = worksheet.cell(
            row=row_number,
            column=2,
        ).value

        notes = worksheet.cell(
            row=row_number,
            column=4,
        ).value

        if assumption_name is None:
            continue

        if assumption_name in group_names:
            current_group = group_names[assumption_name]
            continue

        if assumption_name not in unit_mapping:
            continue

        if current_group is None:
            raise RuntimeError(
                "Assumption group was not found for "
                f"{assumption_name!r}."
            )

        numeric_value = as_decimal(assumption_value)
        text_value = None

        if numeric_value is None:
            text_value = as_text(assumption_value)

        records.append(
            CollectionAssumption(
                import_batch_id=import_batch_id,
                assumption_group=current_group,
                assumption_name=assumption_name,
                numeric_value=numeric_value,
                text_value=text_value,
                unit=unit_mapping[assumption_name],
                notes=as_text(notes),
            )
        )

    session.add_all(records)

    return len(records)


def import_customer_credit_aging(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    header_row = find_row_by_first_cell(
        worksheet,
        "ID",
    )

    records: list[CustomerCreditAging] = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        customer_id = as_text(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if customer_id is None:
            continue

        if customer_id == "TOTAL":
            break

        if not customer_id.startswith("C"):
            continue

        customer_name = as_text(
            worksheet.cell(row_number, 2).value
        )

        customer_segment = as_text(
            worksheet.cell(row_number, 3).value
        )

        payment_terms = as_text(
            worksheet.cell(row_number, 4).value
        )

        currency = as_text(
            worksheet.cell(row_number, 5).value
        )

        payment_trend = as_text(
            worksheet.cell(row_number, 8).value
        )

        if customer_name is None:
            raise RuntimeError(
                f"Customer name is missing at Excel row "
                f"{row_number}."
            )

        if currency not in {
            "IDR",
            "USD",
        }:
            raise RuntimeError(
                f"Invalid currency at Excel row "
                f"{row_number}: {currency!r}"
            )

        if payment_trend not in {
            "Improving",
            "Stable",
            "Worsening",
        }:
            raise RuntimeError(
                f"Invalid payment trend at Excel row "
                f"{row_number}: {payment_trend!r}"
            )

        records.append(
            CustomerCreditAging(
                import_batch_id=import_batch_id,
                customer_id=customer_id,
                customer_name=customer_name,
                customer_segment=customer_segment,
                payment_terms=payment_terms,
                currency=currency,
                credit_limit_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            6,
                        ).value
                    )
                    or Decimal("0")
                ),
                days_beyond_terms=(
                    as_integer(
                        worksheet.cell(
                            row_number,
                            7,
                        ).value
                    )
                    or 0
                ),
                payment_trend=payment_trend,
                has_dispute=as_boolean(
                    worksheet.cell(
                        row_number,
                        9,
                    ).value
                ),
                on_time_percentage=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            10,
                        ).value
                    )
                    or Decimal("0")
                ),
                current_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            11,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_1_30_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            12,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_31_60_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            13,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_61_90_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            14,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_90_plus_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            15,
                        ).value
                    )
                    or Decimal("0")
                ),
                total_ar_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            16,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            17,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_percentage=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            18,
                        ).value
                    )
                    or Decimal("0")
                ),
                credit_utilization=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            19,
                        ).value
                    )
                    or Decimal("0")
                ),
            )
        )

    session.add_all(records)

    return len(records)


def import_risk_scores(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    header_row = find_row_by_first_cell(
        worksheet,
        "ID",
    )

    records: list[CollectionRiskScore] = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        customer_id = as_text(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if customer_id is None:
            continue

        if not customer_id.startswith("C"):
            continue

        customer_name = as_text(
            worksheet.cell(row_number, 2).value
        )

        payment_trend = as_text(
            worksheet.cell(row_number, 8).value
        )

        risk_tier = as_text(
            worksheet.cell(row_number, 16).value
        )

        risk_rank = as_integer(
            worksheet.cell(row_number, 17).value
        )

        if customer_name is None:
            raise RuntimeError(
                f"Customer name is missing at Excel row "
                f"{row_number}."
            )

        if payment_trend not in {
            "Improving",
            "Stable",
            "Worsening",
        }:
            raise RuntimeError(
                f"Invalid payment trend at Excel row "
                f"{row_number}: {payment_trend!r}"
            )

        if risk_tier not in {
            "Low",
            "Medium",
            "High",
        }:
            raise RuntimeError(
                f"Invalid risk tier at Excel row "
                f"{row_number}: {risk_tier!r}"
            )

        if risk_rank is None or risk_rank <= 0:
            raise RuntimeError(
                f"Invalid risk rank at Excel row "
                f"{row_number}: {risk_rank!r}"
            )

        records.append(
            CollectionRiskScore(
                import_batch_id=import_batch_id,
                customer_id=customer_id,
                customer_name=customer_name,
                balance_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            3,
                        ).value
                    )
                    or Decimal("0")
                ),
                utilization_percentage=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            4,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_61_plus_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            5,
                        ).value
                    )
                    or Decimal("0")
                ),
                overdue_severity_percentage=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            6,
                        ).value
                    )
                    or Decimal("0")
                ),
                days_beyond_terms=(
                    as_integer(
                        worksheet.cell(
                            row_number,
                            7,
                        ).value
                    )
                    or 0
                ),
                payment_trend=payment_trend,
                has_dispute=as_boolean(
                    worksheet.cell(
                        row_number,
                        9,
                    ).value
                ),
                dbt_points=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            10,
                        ).value
                    )
                    or Decimal("0")
                ),
                severity_points=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            11,
                        ).value
                    )
                    or Decimal("0")
                ),
                utilization_points=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            12,
                        ).value
                    )
                    or Decimal("0")
                ),
                trend_points=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            13,
                        ).value
                    )
                    or Decimal("0")
                ),
                dispute_points=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            14,
                        ).value
                    )
                    or Decimal("0")
                ),
                risk_score=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            15,
                        ).value
                    )
                    or Decimal("0")
                ),
                risk_tier=risk_tier,
                risk_rank=risk_rank,
            )
        )

    session.add_all(records)

    return len(records)


def import_dso_cash_impact(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    record = CollectionDsoCashImpact(
        import_batch_id=import_batch_id,
        total_ar_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Total AR outstanding (IDR mn)",
                )
            )
            or Decimal("0")
        ),
        current_ar_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Not yet due (current)",
                )
            )
            or Decimal("0")
        ),
        overdue_ar_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Overdue (past due)",
                )
            )
            or Decimal("0")
        ),
        overdue_percentage=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Overdue as % of AR",
                )
            )
            or Decimal("0")
        ),
        annual_credit_sales_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Annual credit sales (IDR mn)",
                )
            )
            or Decimal("0")
        ),
        daily_credit_sales_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Daily credit sales (IDR mn)",
                )
            )
            or Decimal("0")
        ),
        current_dso_days=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Current DSO (days)",
                )
            )
            or Decimal("0")
        ),
        target_dso_days=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Target DSO (days)",
                )
            )
            or Decimal("0")
        ),
        dso_gap_days=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "DSO gap (days)",
                )
            )
            or Decimal("0")
        ),
        cash_freed_at_target_idr_mn=(
            as_decimal(
                get_label_value(
                    worksheet,
                    "Cash freed if target met (IDR mn)",
                )
            )
            or Decimal("0")
        ),
        high_risk_provision_idr_mn=as_decimal(
            get_label_value(
                worksheet,
                (
                    "Illustrative provision on high-risk "
                    "(ECL proxy, IDR mn)"
                ),
            )
        ),
    )

    session.add(record)

    return 1


def import_risk_tier_exposure(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    header_row = find_row_by_first_cell(
        worksheet,
        "Tier",
    )

    records: list[CollectionRiskTierExposure] = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        risk_tier = as_text(
            worksheet.cell(row_number, 1).value
        )

        if risk_tier == "Total":
            break

        if risk_tier not in {
            "Low",
            "Medium",
            "High",
        }:
            continue

        records.append(
            CollectionRiskTierExposure(
                import_batch_id=import_batch_id,
                risk_tier=risk_tier,
                customer_count=(
                    as_integer(
                        worksheet.cell(
                            row_number,
                            2,
                        ).value
                    )
                    or 0
                ),
                exposure_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            3,
                        ).value
                    )
                    or Decimal("0")
                ),
                percentage_of_ar=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            4,
                        ).value
                    )
                    or Decimal("0")
                ),
                notes=as_text(
                    worksheet.cell(
                        row_number,
                        5,
                    ).value
                ),
            )
        )

    session.add_all(records)

    return len(records)


def import_worklist(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    header_row = find_row_by_first_cell(
        worksheet,
        "Rank",
    )

    records: list[CollectionWorklistItem] = []

    for row_number in range(
        header_row + 1,
        worksheet.max_row + 1,
    ):
        priority_rank = as_integer(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if priority_rank is None:
            continue

        customer_name = as_text(
            worksheet.cell(row_number, 2).value
        )

        risk_tier = as_text(
            worksheet.cell(row_number, 6).value
        )

        if customer_name is None:
            raise RuntimeError(
                f"Worklist customer name is missing "
                f"at Excel row {row_number}."
            )

        if risk_tier not in {
            "Low",
            "Medium",
            "High",
        }:
            raise RuntimeError(
                f"Invalid worklist risk tier at Excel row "
                f"{row_number}: {risk_tier!r}"
            )

        records.append(
            CollectionWorklistItem(
                import_batch_id=import_batch_id,
                priority_rank=priority_rank,
                customer_name=customer_name,
                customer_segment=as_text(
                    worksheet.cell(
                        row_number,
                        3,
                    ).value
                ),
                overdue_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            4,
                        ).value
                    )
                    or Decimal("0")
                ),
                oldest_aging_bucket=as_text(
                    worksheet.cell(
                        row_number,
                        5,
                    ).value
                ),
                risk_tier=risk_tier,
                risk_score=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            7,
                        ).value
                    )
                    or Decimal("0")
                ),
                recommended_action=(
                    as_text(
                        worksheet.cell(
                            row_number,
                            8,
                        ).value
                    )
                    or "No recommendation available."
                ),
                recovery_percentage=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            9,
                        ).value
                    )
                    or Decimal("0")
                ),
                expected_recovery_idr_mn=(
                    as_decimal(
                        worksheet.cell(
                            row_number,
                            10,
                        ).value
                    )
                    or Decimal("0")
                ),
            )
        )

    session.add_all(records)

    return len(records)


def normalize_recommendation_title(
    value: str,
) -> str:
    return value.lstrip("-• ").strip()


def import_recommendations(
    session: Session,
    worksheet: Any,
    import_batch_id: int,
) -> int:
    heading_mapping = {
        "1) ACCELERATE  (pull cash in, lower DSO)": (
            "ACCELERATE"
        ),
        "2) CONTAIN  (stop the bleed on risky accounts)": (
            "CONTAIN"
        ),
        "3) PREVENT  (make it stick)": "PREVENT",
    }

    current_type: str | None = None
    recommendation_order = 0
    records: list[CollectionRecommendation] = []

    for row_number in range(
        1,
        worksheet.max_row + 1,
    ):
        first_cell = as_text(
            worksheet.cell(
                row=row_number,
                column=1,
            ).value
        )

        if first_cell is None:
            continue

        if first_cell in heading_mapping:
            current_type = heading_mapping[first_cell]
            continue

        is_action = (
            first_cell.startswith("-")
            or first_cell.startswith("•")
        )

        if not is_action or current_type is None:
            continue

        recommendation_order += 1

        action_title = normalize_recommendation_title(
            first_cell
        )

        records.append(
            CollectionRecommendation(
                import_batch_id=import_batch_id,
                recommendation_type=current_type,
                recommendation_order=(
                    recommendation_order
                ),
                action_title=action_title,
                action_description=(
                    as_text(
                        worksheet.cell(
                            row_number,
                            3,
                        ).value
                    )
                    or action_title
                ),
                expected_impact=as_text(
                    worksheet.cell(
                        row_number,
                        5,
                    ).value
                ),
                requires_approval=True,
                approval_route=(
                    "CFO + Credit / Finance Manager"
                ),
            )
        )

    session.add_all(records)

    return len(records)


def mark_batch_completed(
    session: Session,
    import_batch_id: int,
    total_rows: int,
) -> None:
    import_batch = session.get(
        ImportBatch,
        import_batch_id,
    )

    if import_batch is None:
        raise RuntimeError(
            f"Import batch not found: {import_batch_id}"
        )

    import_batch.import_status = "COMPLETED"
    import_batch.completed_at = datetime.now(
        timezone.utc
    )
    import_batch.total_rows = total_rows


def record_failed_import(
    error_message: str,
) -> None:
    with get_session_factory().begin() as session:
        imported_by = session.scalar(
            select(func.current_user())
        )

        session.add(
            ImportBatch(
                agent_name=AGENT_NAME,
                workbook_name=WORKBOOK_NAME,
                workbook_version=WORKBOOK_VERSION,
                workbook_path=str(WORKBOOK_PATH),
                import_status="FAILED",
                imported_by=imported_by,
                total_sheets=len(EXPECTED_SHEETS),
                total_rows=0,
                error_message=error_message[:5000],
                metadata_json={
                    "source": "Excel Data Engine",
                    "data_type": (
                        "illustrative_demo_data"
                    ),
                    "scope": (
                        "Collections and Credit "
                        "Intelligence Agent"
                    ),
                },
            )
        )


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK_PATH}"
        )

    print(f"Opening workbook: {WORKBOOK_PATH.name}")

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=True,
        read_only=True,
    )

    validate_workbook(workbook)

    row_counts: dict[str, int] = {}

    try:
        with get_session_factory().begin() as session:
            import_batch_id = create_import_batch(
                session
            )

            print(
                f"Import batch created: "
                f"{import_batch_id}"
            )

            row_counts["assumptions"] = (
                import_assumptions(
                    session,
                    workbook["01 Assumptions"],
                    import_batch_id,
                )
            )

            row_counts["customer_credit_aging"] = (
                import_customer_credit_aging(
                    session,
                    workbook[
                        "02 Customer Credit & Aging"
                    ],
                    import_batch_id,
                )
            )

            row_counts["risk_scores"] = (
                import_risk_scores(
                    session,
                    workbook["03 Risk Scoring"],
                    import_batch_id,
                )
            )

            row_counts["dso_cash_impact"] = (
                import_dso_cash_impact(
                    session,
                    workbook["04 DSO & Cash Impact"],
                    import_batch_id,
                )
            )

            row_counts["risk_tier_exposure"] = (
                import_risk_tier_exposure(
                    session,
                    workbook["04 DSO & Cash Impact"],
                    import_batch_id,
                )
            )

            row_counts["worklist"] = (
                import_worklist(
                    session,
                    workbook["05 Collections Worklist"],
                    import_batch_id,
                )
            )

            row_counts["recommendations"] = (
                import_recommendations(
                    session,
                    workbook["06 Recommendation"],
                    import_batch_id,
                )
            )

            total_rows = sum(row_counts.values())

            mark_batch_completed(
                session,
                import_batch_id,
                total_rows,
            )

        print("")
        print(
            "Collections workbook import completed."
        )
        print(f"Import batch ID: {import_batch_id}")

        for table_name, row_count in row_counts.items():
            print(
                f"{table_name}: {row_count} rows"
            )

        print(
            f"Total imported rows: "
            f"{sum(row_counts.values())}"
        )

    except Exception as error:
        print("")
        print("Import failed.")
        print(f"Error: {error}")

        try:
            record_failed_import(str(error))
        except Exception as audit_error:
            print(
                "Failed to record import error: "
                f"{audit_error}"
            )

        raise

    finally:
        workbook.close()


if __name__ == "__main__":
    main()