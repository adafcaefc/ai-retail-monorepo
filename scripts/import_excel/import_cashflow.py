from __future__ import annotations

import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

import psycopg
from openpyxl import load_workbook
from psycopg.types.json import Jsonb


BACKEND_ROOT = Path(__file__).resolve().parents[2]

if str(BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(BACKEND_ROOT))


from src.common.env import config


WORKBOOK_NAME = (
    "03B_Cash Flow Agent Demo Aug 2026 "
    "Data Engine v1.0 20260707.xlsx"
)

WORKBOOK_PATH = BACKEND_ROOT / "data" / WORKBOOK_NAME

WORKBOOK_VERSION = "v1.0-20260707"
AGENT_NAME = "cashflow_agent"

EXPECTED_SHEETS = [
    "02 Assumptions",
    "03 AR Collections",
    "04 AP USD Payables",
    "05 Other Outflows",
    "06 Cash Forecast 13W",
    "07 FX Scenarios",
    "08 Recommendation",
]


def get_connection_string() -> str:
    database_url = config.DATABASE_URL

    if not database_url:
        raise RuntimeError(
            "DATABASE_URL is not configured in the .env file."
        )

    return database_url.replace(
        "postgresql+psycopg://",
        "postgresql://",
        1,
    )


def as_date(value: Any) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    return None


def as_decimal(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None

    if isinstance(value, Decimal):
        return value

    if isinstance(value, bool):
        return Decimal(int(value))

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    try:
        cleaned_value = str(value).replace(",", "")
        return Decimal(cleaned_value)
    except Exception:
        return None


def as_integer(value: Any) -> int | None:
    if value is None or value == "":
        return None

    return int(value)


def as_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value).strip()

    if not text:
        return None

    return text


def validate_workbook(workbook: Any) -> None:
    missing_sheets = [
        sheet_name
        for sheet_name in EXPECTED_SHEETS
        if sheet_name not in workbook.sheetnames
    ]

    if missing_sheets:
        raise RuntimeError(
            "Missing required worksheets: "
            + ", ".join(missing_sheets)
        )


def create_import_batch(
    cursor: psycopg.Cursor[Any],
) -> int:
    cursor.execute(
        """
        INSERT INTO audit.import_batches (
            agent_name,
            workbook_name,
            workbook_version,
            workbook_path,
            import_status,
            imported_by,
            total_sheets,
            total_rows,
            metadata
        )
        VALUES (
            %s,
            %s,
            %s,
            %s,
            'STARTED',
            CURRENT_USER,
            %s,
            0,
            %s
        )
        RETURNING id
        """,
        (
            AGENT_NAME,
            WORKBOOK_NAME,
            WORKBOOK_VERSION,
            str(WORKBOOK_PATH),
            len(EXPECTED_SHEETS),
            Jsonb(
                {
                    "source": "Excel Data Engine",
                    "data_type": "illustrative_demo_data",
                    "scope": "Cash Flow Intelligence Agent",
                }
            ),
        ),
    )

    row = cursor.fetchone()

    if row is None:
        raise RuntimeError("Failed to create import batch.")

    return int(row[0])


def import_assumptions(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    row_groups = {
        5: "COMPANY_PROFILE",
        6: "COMPANY_PROFILE",
        7: "COMPANY_PROFILE",
        8: "COMPANY_PROFILE",
        11: "FORECAST_SETUP",
        12: "FORECAST_SETUP",
        13: "FORECAST_SETUP",
        16: "FX_ASSUMPTIONS",
        17: "FX_ASSUMPTIONS",
        18: "FX_ASSUMPTIONS",
        19: "FX_ASSUMPTIONS",
        20: "FX_ASSUMPTIONS",
        21: "FX_ASSUMPTIONS",
        24: "POLICY_AND_LIQUIDITY",
        25: "POLICY_AND_LIQUIDITY",
        26: "POLICY_AND_LIQUIDITY",
        27: "POLICY_AND_LIQUIDITY",
    }

    units = {
        11: "date",
        12: "weeks",
        13: "text",
        16: "IDR per USD",
        17: "IDR per USD",
        18: "percentage",
        19: "percentage",
        20: "IDR per USD",
        21: "IDR per USD",
        24: "IDR million",
        25: "IDR million",
    }

    inserted_rows = 0

    for row_number, assumption_group in row_groups.items():
        assumption_name = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        value = worksheet.cell(
            row=row_number,
            column=2,
        ).value

        notes = worksheet.cell(
            row=row_number,
            column=4,
        ).value

        numeric_value = None
        text_value = None
        date_value = None

        if isinstance(value, datetime):
            date_value = value.date()
        elif isinstance(value, date):
            date_value = value
        elif isinstance(value, (int, float, Decimal)):
            numeric_value = as_decimal(value)
        else:
            text_value = as_text(value)

        cursor.execute(
            """
            INSERT INTO cashflow.assumptions (
                import_batch_id,
                assumption_group,
                assumption_name,
                numeric_value,
                text_value,
                date_value,
                unit,
                notes
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            """,
            (
                import_batch_id,
                assumption_group,
                as_text(assumption_name),
                numeric_value,
                text_value,
                date_value,
                units.get(row_number),
                as_text(notes),
            ),
        )

        inserted_rows += 1

    return inserted_rows


def import_ar_collections(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    inserted_rows = 0

    for row_number in range(5, 31):
        invoice_number = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        if not invoice_number:
            continue

        cursor.execute(
            """
            INSERT INTO cashflow.ar_collections (
                import_batch_id,
                invoice_number,
                customer_name,
                customer_segment,
                invoice_date,
                payment_terms_days,
                due_date,
                original_week,
                expected_week,
                currency,
                amount_idr_mn,
                usd_amount,
                idr_value_mn,
                delay_flag,
                notes
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s
            )
            """,
            (
                import_batch_id,
                as_text(invoice_number),
                as_text(worksheet.cell(row_number, 2).value),
                as_text(worksheet.cell(row_number, 3).value),
                as_date(worksheet.cell(row_number, 4).value),
                as_integer(worksheet.cell(row_number, 5).value),
                as_date(worksheet.cell(row_number, 6).value),
                as_integer(worksheet.cell(row_number, 7).value),
                as_integer(worksheet.cell(row_number, 8).value),
                as_text(worksheet.cell(row_number, 9).value),
                as_decimal(worksheet.cell(row_number, 10).value),
                as_decimal(worksheet.cell(row_number, 11).value),
                as_decimal(worksheet.cell(row_number, 12).value),
                as_text(worksheet.cell(row_number, 13).value),
                as_text(worksheet.cell(row_number, 14).value),
            ),
        )

        inserted_rows += 1

    return inserted_rows


def import_ap_payables(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    inserted_rows = 0

    for row_number in range(5, 28):
        bill_number = worksheet.cell(
            row=row_number,
            column=1,
        ).value

        if not bill_number:
            continue

        deferrable_value = as_text(
            worksheet.cell(row_number, 11).value
        )

        is_deferrable = (
            deferrable_value is not None
            and deferrable_value.lower() == "yes"
        )

        cursor.execute(
            """
            INSERT INTO cashflow.ap_payables (
                import_batch_id,
                bill_number,
                vendor_name,
                category,
                payment_terms_days,
                due_date,
                payment_week,
                currency,
                amount_idr_mn,
                usd_amount,
                idr_value_mn,
                is_deferrable,
                notes
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s
            )
            """,
            (
                import_batch_id,
                as_text(bill_number),
                as_text(worksheet.cell(row_number, 2).value),
                as_text(worksheet.cell(row_number, 3).value),
                as_integer(worksheet.cell(row_number, 4).value),
                as_date(worksheet.cell(row_number, 5).value),
                as_integer(worksheet.cell(row_number, 6).value),
                as_text(worksheet.cell(row_number, 7).value),
                as_decimal(worksheet.cell(row_number, 8).value),
                as_decimal(worksheet.cell(row_number, 9).value),
                as_decimal(worksheet.cell(row_number, 10).value),
                is_deferrable,
                as_text(worksheet.cell(row_number, 12).value),
            ),
        )

        inserted_rows += 1

    return inserted_rows


def import_other_outflows(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    inserted_rows = 0

    for row_number in range(5, 10):
        category = as_text(
            worksheet.cell(row_number, 1).value
        )

        if not category:
            continue

        for week_number in range(1, 14):
            amount = worksheet.cell(
                row=row_number,
                column=week_number + 1,
            ).value

            cursor.execute(
                """
                INSERT INTO cashflow.other_outflows (
                    import_batch_id,
                    category,
                    week_number,
                    amount_idr_mn
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s
                )
                """,
                (
                    import_batch_id,
                    category,
                    week_number,
                    as_decimal(amount) or Decimal("0"),
                ),
            )

            inserted_rows += 1

    return inserted_rows


def import_weekly_forecast(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    inserted_rows = 0

    row_mapping = {
        "week_start": 5,
        "week_end": 6,
        "customer_collections": 8,
        "total_inflows": 9,
        "vendor_payments_idr": 11,
        "vendor_payments_usd": 12,
        "payroll": 13,
        "rent_utilities_opex": 14,
        "taxes": 15,
        "loan_repayment": 16,
        "total_outflows": 17,
        "net_cash_flow": 19,
        "opening_cash": 20,
        "closing_cash": 21,
        "minimum_buffer": 22,
        "headroom": 23,
        "status": 24,
    }

    for week_number in range(1, 14):
        column_number = week_number + 1

        cursor.execute(
            """
            INSERT INTO cashflow.weekly_forecast (
                import_batch_id,
                week_number,
                week_start,
                week_end,
                customer_collections_idr_mn,
                total_inflows_idr_mn,
                vendor_payments_idr_mn,
                vendor_payments_usd_idr_mn,
                payroll_idr_mn,
                rent_utilities_opex_idr_mn,
                taxes_idr_mn,
                loan_repayment_idr_mn,
                total_outflows_idr_mn,
                net_cash_flow_idr_mn,
                opening_cash_idr_mn,
                closing_cash_idr_mn,
                minimum_buffer_idr_mn,
                headroom_idr_mn,
                status
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s
            )
            """,
            (
                import_batch_id,
                week_number,
                as_date(
                    worksheet.cell(
                        row_mapping["week_start"],
                        column_number,
                    ).value
                ),
                as_date(
                    worksheet.cell(
                        row_mapping["week_end"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["customer_collections"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["total_inflows"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["vendor_payments_idr"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["vendor_payments_usd"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["payroll"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["rent_utilities_opex"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["taxes"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["loan_repayment"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["total_outflows"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["net_cash_flow"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["opening_cash"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["closing_cash"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["minimum_buffer"],
                        column_number,
                    ).value
                ),
                as_decimal(
                    worksheet.cell(
                        row_mapping["headroom"],
                        column_number,
                    ).value
                ),
                as_text(
                    worksheet.cell(
                        row_mapping["status"],
                        column_number,
                    ).value
                ),
            ),
        )

        inserted_rows += 1

    return inserted_rows


def import_fx_scenarios(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    inserted_rows = 0

    net_usd_exposure = as_decimal(
        worksheet.cell(7, 2).value
    )

    for row_number in range(12, 16):
        cursor.execute(
            """
            INSERT INTO cashflow.fx_scenarios (
                import_batch_id,
                scenario_name,
                usd_exposure,
                fx_rate_idr_per_usd,
                movement_vs_spot,
                fx_cash_impact_idr_mn,
                notes,
                is_recommended
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            """,
            (
                import_batch_id,
                as_text(worksheet.cell(row_number, 1).value),
                net_usd_exposure,
                as_decimal(worksheet.cell(row_number, 2).value),
                as_decimal(worksheet.cell(row_number, 3).value),
                as_decimal(worksheet.cell(row_number, 4).value),
                as_text(worksheet.cell(row_number, 5).value),
                False,
            ),
        )

        inserted_rows += 1

    for row_number in range(29, 33):
        scenario_name = as_text(
            worksheet.cell(row_number, 1).value
        )

        is_recommended = (
            scenario_name is not None
            and "RECOMMENDED" in scenario_name.upper()
        )

        cursor.execute(
            """
            INSERT INTO cashflow.fx_scenarios (
                import_batch_id,
                scenario_name,
                action_description,
                usd_exposure,
                downside_avoided_idr_mn,
                premium_idr_mn,
                liquidity_effect,
                confidence_label,
                is_recommended
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            """,
            (
                import_batch_id,
                scenario_name,
                as_text(worksheet.cell(row_number, 2).value),
                net_usd_exposure,
                as_decimal(worksheet.cell(row_number, 3).value),
                as_decimal(worksheet.cell(row_number, 4).value),
                as_text(worksheet.cell(row_number, 5).value),
                as_text(worksheet.cell(row_number, 6).value),
                is_recommended,
            ),
        )

        inserted_rows += 1

    return inserted_rows


def import_recommendations(
    cursor: psycopg.Cursor[Any],
    worksheet: Any,
    import_batch_id: int,
) -> int:
    recommendation_rows = [
        (14, "LIQUIDITY"),
        (15, "LIQUIDITY"),
        (16, "LIQUIDITY"),
        (19, "FX"),
        (20, "FX"),
        (21, "GOVERNANCE"),
    ]

    inserted_rows = 0

    for recommendation_order, item in enumerate(
        recommendation_rows,
        start=1,
    ):
        row_number, recommendation_type = item

        action_title = as_text(
            worksheet.cell(row_number, 1).value
        )

        action_description = as_text(
            worksheet.cell(row_number, 3).value
        )

        expected_impact = as_text(
            worksheet.cell(row_number, 5).value
        )

        if not action_title:
            continue

        if not action_description:
            action_description = action_title

        cursor.execute(
            """
            INSERT INTO cashflow.recommendations (
                import_batch_id,
                recommendation_type,
                recommendation_order,
                action_title,
                action_description,
                expected_impact,
                assumptions,
                risks,
                requires_approval,
                approval_route
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s,
                %s
            )
            """,
            (
                import_batch_id,
                recommendation_type,
                recommendation_order,
                action_title,
                action_description,
                expected_impact,
                Jsonb([]),
                Jsonb([]),
                True,
                "CFO + Treasury Manager",
            ),
        )

        inserted_rows += 1

    return inserted_rows


def mark_batch_completed(
    cursor: psycopg.Cursor[Any],
    import_batch_id: int,
    total_rows: int,
) -> None:
    cursor.execute(
        """
        UPDATE audit.import_batches
        SET
            import_status = 'COMPLETED',
            completed_at = CURRENT_TIMESTAMP,
            total_rows = %s
        WHERE id = %s
        """,
        (
            total_rows,
            import_batch_id,
        ),
    )


def record_failed_import(error_message: str) -> None:
    with psycopg.connect(
        get_connection_string()
    ) as connection:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO audit.import_batches (
                    agent_name,
                    workbook_name,
                    workbook_version,
                    workbook_path,
                    import_status,
                    imported_by,
                    total_sheets,
                    total_rows,
                    error_message,
                    metadata
                )
                VALUES (
                    %s,
                    %s,
                    %s,
                    %s,
                    'FAILED',
                    CURRENT_USER,
                    %s,
                    0,
                    %s,
                    %s
                )
                """,
                (
                    AGENT_NAME,
                    WORKBOOK_NAME,
                    WORKBOOK_VERSION,
                    str(WORKBOOK_PATH),
                    len(EXPECTED_SHEETS),
                    error_message[:5000],
                    Jsonb(
                        {
                            "source": "Excel Data Engine",
                            "data_type": "illustrative_demo_data",
                        }
                    ),
                ),
            )

        connection.commit()


def main() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            f"Workbook not found: {WORKBOOK_PATH}"
        )

    print(f"Opening workbook: {WORKBOOK_PATH.name}")

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=True,
        read_only=True,
    )

    validate_workbook(workbook)

    row_counts: dict[str, int] = {}

    try:
        with psycopg.connect(
            get_connection_string()
        ) as connection:
            with connection.cursor() as cursor:
                import_batch_id = create_import_batch(cursor)

                print(
                    f"Import batch created: {import_batch_id}"
                )

                row_counts["assumptions"] = import_assumptions(
                    cursor,
                    workbook["02 Assumptions"],
                    import_batch_id,
                )

                row_counts["ar_collections"] = (
                    import_ar_collections(
                        cursor,
                        workbook["03 AR Collections"],
                        import_batch_id,
                    )
                )

                row_counts["ap_payables"] = import_ap_payables(
                    cursor,
                    workbook["04 AP USD Payables"],
                    import_batch_id,
                )

                row_counts["other_outflows"] = (
                    import_other_outflows(
                        cursor,
                        workbook["05 Other Outflows"],
                        import_batch_id,
                    )
                )

                row_counts["weekly_forecast"] = (
                    import_weekly_forecast(
                        cursor,
                        workbook["06 Cash Forecast 13W"],
                        import_batch_id,
                    )
                )

                row_counts["fx_scenarios"] = (
                    import_fx_scenarios(
                        cursor,
                        workbook["07 FX Scenarios"],
                        import_batch_id,
                    )
                )

                row_counts["recommendations"] = (
                    import_recommendations(
                        cursor,
                        workbook["08 Recommendation"],
                        import_batch_id,
                    )
                )

                total_rows = sum(row_counts.values())

                mark_batch_completed(
                    cursor,
                    import_batch_id,
                    total_rows,
                )

            connection.commit()

        print("")
        print("Cash Flow workbook import completed.")
        print(f"Import batch ID: {import_batch_id}")

        for table_name, row_count in row_counts.items():
            print(f"{table_name}: {row_count} rows")

        print(
            f"Total imported rows: {sum(row_counts.values())}"
        )

    except Exception as error:
        print("")
        print("Import failed.")
        print(f"Error: {error}")

        try:
            record_failed_import(str(error))
        except Exception as audit_error:
            print(
                "Failed to record import error: "
                f"{audit_error}"
            )

        raise

    finally:
        workbook.close()


if __name__ == "__main__":
    main()